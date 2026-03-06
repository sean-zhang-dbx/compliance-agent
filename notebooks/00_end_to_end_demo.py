# Databricks notebook source
# MAGIC %md
# MAGIC # GSK Controls Evidence Review Agent — End-to-End Demo
# MAGIC
# MAGIC This notebook walks through the **complete lifecycle** of an automated FRMC control test:
# MAGIC
# MAGIC | Step | What Happens | Key Output |
# MAGIC |------|-------------|------------|
# MAGIC | **0. Setup** | Install deps, configure paths | Environment ready |
# MAGIC | **1. Generate Data** | Create synthetic projects with realistic evidence | 6 project directories |
# MAGIC | **2. Explore Inputs** | Inspect engagement metadata, workbooks, evidence files | Understand what the agent sees |
# MAGIC | **3. Agent Walkthrough** | Call each tool step-by-step to show intermediate outputs | Tool-level I/O |
# MAGIC | **4. Full Agent Run** | Execute the complete agent on a project end-to-end | Final report |
# MAGIC | **5. Review Outputs** | Inspect the report, MLflow traces, and email notification | Audit-ready deliverables |

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 0: Setup & Configuration

# COMMAND ----------

# MAGIC %pip install mlflow>=3.1 databricks-langchain langgraph>=0.3.4 langchain-core openpyxl PyMuPDF fpdf2 Pillow requests --quiet
# MAGIC dbutils.library.restartPython()

# COMMAND ----------

import os, json, sys
from pathlib import Path
from pprint import pprint

dbutils.widgets.text("catalog", "", "UC Catalog")

notebook_path = dbutils.notebook.entry_point.getDbutils().notebook().getContext().notebookPath().get()
workspace_root = "/Workspace" + str(Path(notebook_path).parent.parent)

sys.path.insert(0, workspace_root)

os.environ["LLM_ENDPOINT"] = "databricks-claude-opus-4-6"
os.environ["VISION_LLM_ENDPOINT"] = "databricks-claude-sonnet-4-6"
os.environ["FAST_LLM_ENDPOINT"] = "databricks-claude-haiku-4-5"
os.environ["UC_CATALOG"] = dbutils.widgets.get("catalog") if dbutils.widgets.get("catalog") else spark.conf.get("spark.databricks.unityCatalog.defaultCatalog", "main")
os.environ["PROJECTS_LOCAL_PATH"] = f"{workspace_root}/sample_data/projects"

import mlflow
mlflow.langchain.autolog()
experiment_name = f"/Users/{spark.conf.get('spark.databricks.notebook.userName', 'demo')}/gsk-compliance-agent-demo"
mlflow.set_experiment(experiment_name)

print(f"Workspace root: {workspace_root}")
print(f"Projects path:  {os.environ['PROJECTS_LOCAL_PATH']}")
print(f"MLflow experiment: {experiment_name}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 1: Generate Sample Data
# MAGIC
# MAGIC We generate **6 realistic control testing projects**, each with:
# MAGIC - `engagement.json` — metadata, rules, testing attributes, evidence manifest
# MAGIC - `engagement_workbook.xlsx` — population data, sampling config, testing table
# MAGIC - `evidence/` — PDFs, screenshots, emails, photos (varies by control type)
# MAGIC
# MAGIC | Project | Control | Domain | Evidence Types |
# MAGIC |---------|---------|--------|---------------|
# MAGIC | `p2p_028` | Payment Proposal Review | Accounts Payable | PDFs (invoices, policy) |
# MAGIC | `itg_015` | User Access Review | IT General Controls | PDFs, Screenshots |
# MAGIC | `fin_042` | Journal Entry Review | Financial Reporting | PDFs, Emails (.eml) |
# MAGIC | `hr_003` | Segregation of Duties | HR / IT Controls | _(workbook only)_ |
# MAGIC | `rev_019` | Revenue Cutoff Testing | Revenue Recognition | PDFs, Photos |
# MAGIC | `env_007` | Environmental Inspection | EHS | Embedded Excel images |

# COMMAND ----------

exec(open(f"{workspace_root}/sample_data/generate_all_projects.py").read())

projects_dir = Path(os.environ["PROJECTS_LOCAL_PATH"])
print("Generated projects:")
for d in sorted(projects_dir.iterdir()):
    if d.is_dir():
        files = list(d.rglob("*"))
        file_types = set(f.suffix for f in files if f.is_file())
        print(f"  📁 {d.name}/  ({len(files)} files: {', '.join(sorted(file_types))})")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 2: Explore the Inputs
# MAGIC
# MAGIC Before running the agent, let's see exactly what it receives as input.

# COMMAND ----------

# MAGIC %md
# MAGIC ### 2a. Engagement Metadata
# MAGIC The engagement JSON is the **instruction file** — it tells the agent what control to test, what rules apply, and what evidence to examine.

# COMMAND ----------

project = "p2p_028"
engagement_path = projects_dir / project / "engagement.json"
engagement = json.loads(engagement_path.read_text())

print(f"{'='*60}")
print(f"ENGAGEMENT: {engagement['number']} — {engagement['name']}")
print(f"{'='*60}")
print(f"  Control ID:     {engagement['control_objective']['control_id']}")
print(f"  Control Name:   {engagement['control_objective']['control_name']}")
print(f"  Domain:         {engagement['control_objective']['domain']}")
print(f"  Lead:           {engagement['engagement_lead']}")
print(f"  Start Date:     {engagement['planned_start']}")
print()
print("  RULES (from engagement JSON):")
rules = engagement["control_objective"]["rules"]
for k, v in rules.items():
    print(f"    {k}: {v}")
print()
print("  TESTING ATTRIBUTES:")
for attr in engagement["testing_attributes"]:
    print(f"    [{attr['ref']}] {attr['name']}  (applies to: {attr['applies_to']})")
print()
print("  EVIDENCE FILES:")
for ef in engagement["evidence_files"]:
    print(f"    📄 {ef['path']}  (type={ef['type']}, focus={ef['focus']})")
print()
print("  NOTIFICATION EMAILS:")
emails = engagement.get("notification_emails", {})
for k, v in emails.items():
    print(f"    {k}: {v}")

# COMMAND ----------

# MAGIC %md
# MAGIC ### 2b. Engagement Workbook (Excel)
# MAGIC The workbook contains the **population data, sampling methodology, and testing table**. Let's see what's inside.

# COMMAND ----------

import openpyxl, io

wb_path = projects_dir / project / "engagement_workbook.xlsx"
wb = openpyxl.load_workbook(str(wb_path), data_only=True)

print(f"Workbook: {wb_path.name}")
print(f"Tabs: {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(c is not None for c in r)]
    print(f"  📋 {sheet_name}: {len(non_empty)} rows × {ws.max_column} cols")
    if non_empty:
        headers = non_empty[0]
        print(f"     Headers: {[str(h)[:30] for h in headers if h]}")
        if len(non_empty) > 1:
            print(f"     First row: {[str(c)[:20] for c in non_empty[1] if c]}")
    print()

# COMMAND ----------

# MAGIC %md
# MAGIC ### 2c. Evidence Files
# MAGIC PDFs, screenshots, emails — these are the supporting documents the agent reviews.

# COMMAND ----------

evidence_dir = projects_dir / project / "evidence"
if evidence_dir.exists():
    for f in sorted(evidence_dir.iterdir()):
        size_kb = f.stat().st_size / 1024
        print(f"  {'📄' if f.suffix=='.pdf' else '📧' if f.suffix=='.eml' else '🖼️'} {f.name}  ({size_kb:.1f} KB)")
else:
    print("  No evidence directory found for this project")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 3: Agent Walkthrough — Tool by Tool
# MAGIC
# MAGIC Now let's **manually call each tool** the agent uses, in order, to see the intermediate inputs and outputs at every step.

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 1: `list_projects` → Discover available projects

# COMMAND ----------

from agent.tools import (
    list_projects, load_engagement, parse_workbook,
    review_document, execute_test, compile_results,
    save_report, send_email
)

result = list_projects.invoke({})
projects_data = json.loads(result)
print(f"Found {projects_data['count']} projects:\n")
for p in projects_data["projects"]:
    print(f"  [{p.get('control_id', '?')}] {p.get('control_name', p['project_dir'])}  ({p.get('domain', '')})")

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 2: `load_engagement` → Read the instruction file
# MAGIC
# MAGIC **Input:** project directory name  
# MAGIC **Output:** Full engagement metadata with rules, attributes, and evidence list

# COMMAND ----------

result = load_engagement.invoke({"project_path": project})
eng = json.loads(result)

print(f"Control: {eng['control_objective']['control_id']} — {eng['control_objective']['control_name']}")
print(f"Domain:  {eng['control_objective']['domain']}")
print(f"Rules:   {json.dumps(eng['control_objective']['rules'], indent=2)}")
print(f"\nTesting Attributes: {len(eng['testing_attributes'])}")
print(f"Evidence Files:     {len(eng['evidence_files'])}")

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 3: `parse_workbook` → Read the Excel workbook
# MAGIC
# MAGIC **Input:** project directory name  
# MAGIC **Output:** Tab names, population data, sampling config, selected sample, testing attributes, embedded image detection

# COMMAND ----------

result = parse_workbook.invoke({"file_path": project})
wb_data = json.loads(result)

print(f"Tabs: {wb_data['tab_names']}")
print(f"Has embedded images: {wb_data['has_embedded_images']}")
print(f"Sampling config: {json.dumps(wb_data['sampling_config'], indent=2)}")
print(f"\nSelected sample ({len(wb_data['selected_sample'])} items):")
for i, item in enumerate(wb_data["selected_sample"][:3]):
    print(f"  [{i+1}] {json.dumps(item)}")
if len(wb_data["selected_sample"]) > 3:
    print(f"  ... and {len(wb_data['selected_sample'])-3} more")

print(f"\nTesting attributes ({len(wb_data['testing_attributes'])}):")
for attr in wb_data["testing_attributes"]:
    print(f"  [{attr['ref']}] {attr['attribute'][:80]}")

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 4: `review_document` → Analyze a PDF with the LLM
# MAGIC
# MAGIC **Input:** file path + control context + focus area  
# MAGIC **Output:** LLM analysis of the document contents

# COMMAND ----------

evidence_files = eng["evidence_files"]
if evidence_files:
    first_evidence = evidence_files[0]
    evidence_path = str(projects_dir / project / first_evidence["path"])
    
    print(f"Reviewing: {first_evidence['path']}")
    print(f"Type: {first_evidence['type']}")
    print(f"Focus: {first_evidence['focus']}")
    print("=" * 60)
    
    result = review_document.invoke({
        "file_path": evidence_path,
        "context": f"Control {eng['control_objective']['control_id']} - {eng['control_objective']['control_name']}",
        "focus_area": first_evidence["focus"],
    })
    doc_review = json.loads(result)
    print(f"\nFile size: {doc_review['file_size_bytes']} bytes")
    print(f"\n--- LLM Analysis ---\n{doc_review['analysis'][:2000]}")

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 5: `execute_test` → Run a test attribute against a sample item
# MAGIC
# MAGIC **Input:** test reference, attribute, procedure, control context, sample item, evidence summary  
# MAGIC **Output:** Pass/Fail result with narrative and exception details

# COMMAND ----------

if wb_data["testing_attributes"] and wb_data["selected_sample"]:
    test_attr = wb_data["testing_attributes"][0]
    sample_item = wb_data["selected_sample"][0]
    
    print(f"Running test [{test_attr['ref']}]: {test_attr['attribute'][:80]}")
    print(f"Against sample: {json.dumps(sample_item)[:120]}")
    print("=" * 60)
    
    result = execute_test.invoke({
        "test_ref": test_attr["ref"],
        "attribute": test_attr["attribute"],
        "procedure": test_attr.get("procedure", ""),
        "control_context": json.dumps(eng["control_objective"]),
        "sample_item_json": json.dumps(sample_item),
        "evidence_summary": doc_review.get("analysis", "")[:2000] if evidence_files else "",
    })
    test_result = json.loads(result)
    
    print(f"\nTest Ref: {test_result['test_ref']}")
    print(f"\n--- LLM Analysis ---\n{test_result['llm_analysis'][:2000]}")

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 6: `compile_results` → Generate the final report
# MAGIC
# MAGIC **Input:** All test results, control metadata, sampling info  
# MAGIC **Output:** Structured markdown report with executive summary, results table, exception details, and issue template

# COMMAND ----------

pop_size = 0
for tab_name, tab_info in wb_data.get("tabs", {}).items():
    if "population" in tab_name.lower() or "journal" in tab_name.lower() or "cutoff" in tab_name.lower():
        pop_size = tab_info.get("row_count", 0)
        break
if not pop_size:
    pop_size = int(wb_data.get("sampling_config", {}).get("Population Size", 0))

report = compile_results.invoke({
    "control_id": eng["control_objective"]["control_id"],
    "control_name": eng["control_objective"]["control_name"],
    "engagement_number": eng["number"],
    "domain": eng["control_objective"]["domain"],
    "population_size": pop_size,
    "sample_size": len(wb_data["selected_sample"]),
    "testing_attributes_json": json.dumps(wb_data["testing_attributes"]),
    "test_results_json": json.dumps([test_result]),
    "rules_json": json.dumps(eng["control_objective"]["rules"]),
})

print("=" * 60)
print("COMPILED REPORT (first 3000 chars)")
print("=" * 60)
print(report[:3000])

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 7: `save_report` → Persist the report
# MAGIC
# MAGIC **Input:** project path + report content  
# MAGIC **Output:** Saved file paths (local + UC volume)

# COMMAND ----------

result = save_report.invoke({
    "project_path": project,
    "report_content": report,
    "report_format": "both",
})
save_result = json.loads(result)
print(f"Status: {save_result['status']}")
print(f"Report length: {save_result['report_length']} chars")
print(f"Saved to:")
for f in save_result["files"]:
    print(f"  📄 {f}")

# COMMAND ----------

# MAGIC %md
# MAGIC ### Tool 8: `send_email` → Notify stakeholders
# MAGIC
# MAGIC **Input:** recipient, subject, body (HTML)  
# MAGIC **Output:** Sent status (or simulated .eml if SMTP not configured)

# COMMAND ----------

notification = eng.get("notification_emails", {})
to_addr = notification.get("report_to", "sean.zhang@databricks.com")

result = send_email.invoke({
    "to": to_addr,
    "subject": f"[{eng['control_objective']['control_id']}] Controls Evidence Review Complete",
    "body": f"<h2>{eng['control_objective']['control_id']} - Review Complete</h2><p>The automated review for engagement {eng['number']} is complete.</p>",
    "importance": "high",
    "project_path": project,
})
email_result = json.loads(result)
print(f"Email status: {email_result['status']}")
print(f"To: {email_result['to']}")
print(f"Method: {email_result['method']}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 4: Full Agent Run (End-to-End)
# MAGIC
# MAGIC Now let's run the **complete agent** on a different project. The agent autonomously:
# MAGIC 1. Loads the engagement instructions
# MAGIC 2. Parses the workbook
# MAGIC 3. Reviews all evidence files
# MAGIC 4. Executes all testing attributes
# MAGIC 5. Compiles the final report
# MAGIC 6. Saves and emails the results
# MAGIC
# MAGIC All with **MLflow tracing** capturing every tool call.

# COMMAND ----------

from agent.agent import AGENT
from mlflow.types.responses import ResponsesAgentRequest

full_project = "fin_042"

request = ResponsesAgentRequest(
    input=[{"role": "user", "content": f"Review project {full_project}. Execute all testing attributes and produce the full report."}]
)

with mlflow.start_run(run_name=f"full_review_{full_project}"):
    response = AGENT.predict(request)

# COMMAND ----------

for item in response.output:
    if hasattr(item, "content"):
        for block in item.content:
            if hasattr(block, "text") and block.text:
                print(block.text[:5000])
                print("\n" + "=" * 60 + "\n")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 5: Review Outputs
# MAGIC
# MAGIC ### 5a. Saved Reports

# COMMAND ----------

for proj_dir in sorted(projects_dir.iterdir()):
    if proj_dir.is_dir():
        reports = list(proj_dir.glob("report_*.md"))
        if reports:
            for r in reports:
                size_kb = r.stat().st_size / 1024
                print(f"  📝 {proj_dir.name}/{r.name}  ({size_kb:.1f} KB)")

# COMMAND ----------

# MAGIC %md
# MAGIC ### 5b. MLflow Traces
# MAGIC
# MAGIC Every tool call is captured in MLflow. Navigate to the experiment to see:
# MAGIC - **Trace timeline** — which tools were called and in what order
# MAGIC - **Span details** — inputs and outputs of each tool invocation
# MAGIC - **Token usage** — how many LLM tokens were consumed
# MAGIC - **Latency breakdown** — where time was spent

# COMMAND ----------

traces = mlflow.search_traces(
    experiment_names=[experiment_name],
    max_results=10,
)

if len(traces) > 0:
    display(traces[["request_id", "timestamp_ms", "status", "execution_time_ms", "tags"]].head(10))
    print(f"\nTotal traces: {len(traces)}")
    print(f"View full traces at: {experiment_name}")
else:
    print("No traces found yet. Run Step 4 first.")

# COMMAND ----------

# MAGIC %md
# MAGIC ### 5c. Simulated Email Outputs

# COMMAND ----------

for proj_dir in sorted(projects_dir.iterdir()):
    if proj_dir.is_dir():
        emls = list(proj_dir.glob("sent_email_*.eml"))
        if emls:
            for eml_file in emls:
                print(f"  📧 {proj_dir.name}/{eml_file.name}")
                content = eml_file.read_text()
                for line in content.split("\n")[:6]:
                    print(f"     {line}")
                print()

# COMMAND ----------

# MAGIC %md
# MAGIC ## Summary
# MAGIC
# MAGIC ### What We Demonstrated
# MAGIC
# MAGIC | Component | Technology | Purpose |
# MAGIC |-----------|-----------|---------|
# MAGIC | **Agent Orchestration** | LangGraph | Stateful tool-calling loop |
# MAGIC | **LLM Reasoning** | Databricks Foundation Model API (Claude Opus 4.6 / Haiku 4.5) | Document analysis, test execution, report generation |
# MAGIC | **Vision Analysis** | Claude Sonnet 4.6 (multimodal) | Screenshot, photo, and embedded image analysis |
# MAGIC | **Data Storage** | Unity Catalog Volumes | Project files, evidence, reports |
# MAGIC | **Observability** | MLflow Tracing | Full trace of every tool call with inputs/outputs |
# MAGIC | **Deployment** | Databricks Apps | FastAPI backend + React frontend |
# MAGIC | **Notifications** | Gmail SMTP | Email report delivery |
# MAGIC
# MAGIC ### Inputs → Outputs Flow
# MAGIC
# MAGIC ```
# MAGIC INPUTS                          AGENT WORKFLOW                    OUTPUTS
# MAGIC ┌──────────────────┐     ┌─────────────────────────┐     ┌──────────────────────┐
# MAGIC │ engagement.json  │────▶│ 1. Load instructions    │     │ Markdown report       │
# MAGIC │  - control rules │     │ 2. Parse workbook       │────▶│  - Executive summary  │
# MAGIC │  - test attrs    │     │ 3. Review evidence      │     │  - Results table      │
# MAGIC │  - evidence list │     │ 4. Execute tests (A-F)  │     │  - Exception details  │
# MAGIC ├──────────────────┤     │ 5. Compile results      │     │  - Issue template     │
# MAGIC │ workbook.xlsx    │────▶│ 6. Save report          │     ├──────────────────────┤
# MAGIC │  - population    │     │ 7. Email notification   │     │ JSON summary          │
# MAGIC │  - sample        │     └─────────────────────────┘     │  - assessment rating  │
# MAGIC │  - testing table │                                     │  - has_exceptions     │
# MAGIC ├──────────────────┤                                     ├──────────────────────┤
# MAGIC │ evidence/        │                                     │ Email notification    │
# MAGIC │  - PDFs          │                                     │  - report_to          │
# MAGIC │  - screenshots   │                                     │  - exceptions_to      │
# MAGIC │  - emails (.eml) │                                     ├──────────────────────┤
# MAGIC │  - photos        │                                     │ MLflow traces         │
# MAGIC │  - embedded imgs │                                     │  - full tool timeline │
# MAGIC └──────────────────┘                                     │  - token usage        │
# MAGIC                                                          └──────────────────────┘
# MAGIC ```
