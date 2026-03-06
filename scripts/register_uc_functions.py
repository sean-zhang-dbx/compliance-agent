#!/usr/bin/env python3
"""
Register all 18 compliance agent tools as Unity Catalog Python functions.

Config values (LLM endpoints, volume paths) are baked into function bodies
at registration time.  Runtime context (run_id, project_dir, app_base_url)
is passed as explicit parameters where needed.

Usage:
    python scripts/register_uc_functions.py --catalog MY_CATALOG
    python scripts/register_uc_functions.py --catalog MY_CATALOG --schema my_schema
"""

import argparse
import json
import sys
import textwrap
import time as _time


# ═══════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════

def _cfg(body: str, cfg: dict) -> str:
    """Replace {CFG_XXX} placeholders with actual config values."""
    for key, value in cfg.items():
        body = body.replace("{CFG_" + key.upper() + "}", str(value))
    return body


def _build_sql(catalog: str, schema: str, name: str, defn: dict, cfg: dict) -> str:
    fqn = f"`{catalog}`.`{schema}`.`{name}`"
    params = defn["params"]
    returns = defn.get("returns", "STRING")
    comment = defn.get("comment", "").replace("'", "''")
    deps = defn.get("deps", [])
    body = _cfg(defn["body"], cfg)

    parts = [f"CREATE OR REPLACE FUNCTION {fqn}({params})"]
    parts.append(f"RETURNS {returns}")
    if comment:
        parts.append(f"COMMENT '{comment}'")
    parts.append("LANGUAGE PYTHON")
    if deps:
        parts.append(f"ENVIRONMENT (dependencies = '{json.dumps(deps)}', environment_version = 'None')")
    parts.append(f"AS $$\n{body}\n$$")
    return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════
# TIER 1 — Pure logic (stdlib only)
# ═══════════════════════════════════════════════════════════════════════

TIER1 = {}

TIER1["announce_plan"] = dict(
    params="steps STRING",
    comment="Announce the structured assessment plan to the user. Call ONCE after loading engagement. Args: steps – JSON array of plan step objects each with id, label, and optional detail.",
    body=textwrap.dedent("""\
        import json
        parsed = json.loads(steps)
        return json.dumps({"status": "plan_announced", "step_count": len(parsed), "steps": parsed})
    """),
)

TIER1["ask_user"] = dict(
    params="question STRING, options STRING",
    comment="Ask the user for clarification when uncertain. Args: question – the question text; options – optional comma-separated suggested answers (pass empty string if none).",
    body=textwrap.dedent("""\
        import json
        result = {"type": "user_question", "question": question}
        if options:
            result["suggested_options"] = [o.strip() for o in options.split(",")]
        return json.dumps(result, indent=2)
    """),
)

TIER1["generate_test_plan"] = dict(
    params="engagement_json STRING, workbook_json STRING",
    comment="Generate a deterministic test plan from engagement and workbook data. Call AFTER load_engagement and parse_workbook. Returns JSON with ordered test_plan list and total_tests count.",
    body=textwrap.dedent("""\
        import json

        engagement = json.loads(engagement_json) if isinstance(engagement_json, str) else engagement_json
        workbook = json.loads(workbook_json) if isinstance(workbook_json, str) else workbook_json

        attributes = engagement.get("testing_attributes", [])
        rules = engagement.get("control_objective", {}).get("rules", {})
        threshold = rules.get("threshold_gbp", rules.get("threshold_usd", rules.get("threshold", 0)))

        selected_sample = workbook.get("selected_sample", [])
        wb_attributes = workbook.get("testing_attributes", [])

        attr_procedures = {}
        for wa in wb_attributes:
            attr_procedures[wa.get("ref", "")] = wa.get("procedure", wa.get("attribute", ""))

        test_plan = []
        for attr in attributes:
            ref = attr.get("ref", "?")
            name = attr.get("name", "")
            applies_to = attr.get("applies_to", "all")
            procedure = attr_procedures.get(ref, name)

            if applies_to == "control_level":
                test_plan.append({
                    "test_ref": ref,
                    "attribute": name,
                    "procedure": procedure,
                    "applies_to": applies_to,
                    "sample_item_json": json.dumps({
                        "_type": "population_level",
                        "population_size": workbook.get("sampling_config", {}).get(
                            "Population Size",
                            workbook.get("sampling_config", {}).get("Total Population", "unknown")),
                    }),
                })
            elif applies_to in ("all",):
                for item in selected_sample:
                    test_plan.append({
                        "test_ref": ref,
                        "attribute": name,
                        "procedure": procedure,
                        "applies_to": applies_to,
                        "sample_item_json": json.dumps(item),
                    })
                if not selected_sample:
                    test_plan.append({
                        "test_ref": ref,
                        "attribute": name,
                        "procedure": procedure,
                        "applies_to": applies_to,
                        "sample_item_json": json.dumps({"_type": "no_sample_available"}),
                    })
            else:
                matched_items = []
                for item in selected_sample:
                    amount_str = ""
                    for k, v in item.items():
                        kl = k.lower()
                        if any(w in kl for w in ["amount", "value", "total", "gbp", "usd"]):
                            amount_str = str(v).replace(",", "").replace("\\u00a3", "").replace("$", "").strip()
                            break
                    if applies_to == "above_threshold" and threshold:
                        try:
                            if float(amount_str) >= float(threshold):
                                matched_items.append(item)
                        except (ValueError, TypeError):
                            matched_items.append(item)
                    else:
                        matched_items.append(item)

                if matched_items:
                    for item in matched_items:
                        test_plan.append({
                            "test_ref": ref,
                            "attribute": name,
                            "procedure": procedure,
                            "applies_to": applies_to,
                            "sample_item_json": json.dumps(item),
                        })
                else:
                    test_plan.append({
                        "test_ref": ref,
                        "attribute": name,
                        "procedure": procedure,
                        "applies_to": applies_to,
                        "sample_item_json": json.dumps({"_type": "no_applicable_items", "filter": applies_to}),
                    })

        return json.dumps({
            "test_plan": test_plan,
            "total_tests": len(test_plan),
            "attributes_count": len(attributes),
            "sample_size": len(selected_sample),
            "instruction": "Execute EVERY entry in test_plan using execute_test. Do NOT skip any.",
        }, indent=2)
    """),
)

TIER1["aggregate_test_results"] = dict(
    params="batch_results_json STRING",
    comment="Deterministically aggregate per-sample test results into per-attribute summaries. Call AFTER batch_execute_tests. Returns JSON array of per-attribute results for fill_workbook and compile_results.",
    body=textwrap.dedent("""\
        import json

        SAMPLE_ID_FIELDS = [
            "JE_Number", "Order_No", "Location_ID", "User_ID", "Document_No",
            "Inspection_ID", "Sample_No", "Row_No", "Location_Name",
        ]

        def extract_sample_id(sample):
            for field in SAMPLE_ID_FIELDS:
                val = sample.get(field)
                if val:
                    return str(val)
            return json.dumps(sample, sort_keys=True)[:80]

        def parse_llm_analysis(raw):
            text = raw.strip()
            if text.startswith("```"):
                text = text.strip("`").lstrip("json").lstrip("\\n")
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                return {"result": "Unknown", "narrative": raw[:500]}

        batch = json.loads(batch_results_json) if isinstance(batch_results_json, str) else batch_results_json
        raw_results = batch.get("results", batch) if isinstance(batch, dict) else batch

        by_ref = {}
        for entry in raw_results:
            ref = entry.get("test_ref", "?").strip().upper()
            sample = entry.get("sample_item", {})
            if isinstance(sample, str):
                try:
                    sample = json.loads(sample)
                except json.JSONDecodeError:
                    sample = {}

            parsed = parse_llm_analysis(entry.get("llm_analysis", "{}"))
            sample_id = extract_sample_id(sample)
            result_lower = parsed.get("result", "Unknown").strip().lower()

            if ref not in by_ref:
                by_ref[ref] = {"results": [], "narratives": [], "sample_ids": [], "exceptions": []}

            bucket = by_ref[ref]
            bucket["results"].append(result_lower)
            bucket["narratives"].append(f"[{sample_id}]: {parsed.get('narrative', '')}")
            bucket["sample_ids"].append(sample_id)

            if result_lower == "fail":
                bucket["exceptions"].append({
                    "description": parsed.get("exception") or parsed.get("narrative", "Test failed"),
                    "severity": parsed.get("severity", "Medium"),
                    "sample_id": sample_id,
                })

        aggregated = []
        for ref in sorted(by_ref.keys()):
            bucket = by_ref[ref]
            results_set = set(bucket["results"])

            if results_set == {"pass"}:
                agg_result = "Pass"
            elif "fail" in results_set:
                agg_result = "Fail"
            elif results_set <= {"not applicable"}:
                agg_result = "Not Applicable"
            else:
                agg_result = "Partial"

            merged_exceptions = []
            if bucket["exceptions"]:
                all_affected = sorted({e["sample_id"] for e in bucket["exceptions"]})
                descriptions = []
                severity = "Medium"
                severity_rank = {"critical": 4, "high": 3, "medium": 2, "low": 1}
                for exc in bucket["exceptions"]:
                    descriptions.append(exc["description"])
                    exc_sev = exc.get("severity", "Medium")
                    if severity_rank.get(exc_sev.lower(), 0) > severity_rank.get(severity.lower(), 0):
                        severity = exc_sev

                merged_exceptions = [{
                    "description": "; ".join(dict.fromkeys(descriptions)),
                    "severity": severity,
                    "affected_samples": all_affected,
                    "root_cause": "See test narrative",
                    "remediation": "Review and remediate",
                }]

            aggregated.append({
                "ref": ref,
                "result": agg_result,
                "narrative": "\\n".join(bucket["narratives"]),
                "sample_items_tested": bucket["sample_ids"],
                "exceptions": merged_exceptions,
            })

        return json.dumps(aggregated, indent=2)
    """),
)


# ═══════════════════════════════════════════════════════════════════════
# TIER 2 — Volume I/O (databricks-sdk)
# ═══════════════════════════════════════════════════════════════════════

TIER2 = {}

TIER2["list_projects"] = dict(
    params="",
    comment="List all available control testing projects. Call FIRST to discover projects. Returns JSON with project list including name, control_id, control_name, and domain.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        w = WorkspaceClient()
        projects = []
        try:
            items = w.files.list_directory_contents(PROJECTS_BASE_PATH)
            for item in items:
                if item.is_directory:
                    name = item.path.rstrip("/").split("/")[-1]
                    info = {"project_dir": name, "source": "uc_volume"}
                    try:
                        resp = w.files.download(f"{PROJECTS_BASE_PATH}/{name}/engagement.json")
                        eng = json.loads(resp.contents.read().decode("utf-8"))
                        co = eng.get("control_objective", {})
                        info.update({
                            "engagement_number": eng.get("number", ""),
                            "engagement_name": eng.get("name", ""),
                            "control_id": co.get("control_id", ""),
                            "control_name": co.get("control_name", ""),
                            "domain": co.get("domain", ""),
                        })
                    except Exception:
                        pass
                    projects.append(info)
        except Exception:
            pass
        return json.dumps({"projects": projects, "count": len(projects)}, indent=2)
    """),
)

TIER2["load_engagement"] = dict(
    params="project_path STRING",
    comment="Load engagement metadata and instructions from a project. Args: project_path – project dir name (e.g. p2p_028) or full path to engagement.json. Returns JSON with control_objective, testing_attributes, instructions, evidence_files.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        w = WorkspaceClient()
        if project_path.endswith(".json"):
            file_path = project_path
        else:
            file_path = f"{PROJECTS_BASE_PATH}/{project_path}/engagement.json"
        resp = w.files.download(file_path)
        engagement = json.loads(resp.contents.read().decode("utf-8"))
        return json.dumps(engagement, indent=2)
    """),
)

TIER2["save_report"] = dict(
    params="project_path STRING, report_content STRING, report_format STRING, control_id STRING, control_name STRING, run_id STRING, project_dir STRING, app_base_url STRING",
    comment="Save the final report to the project directory in UC Volumes. Args: project_path – project dir name; report_content – full markdown report; report_format – markdown or both; control_id – control ID for filename; control_name – control name for filename; run_id – current run ID; project_dir – project directory; app_base_url – app base URL for building report links. Returns JSON with saved file paths and report_url.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import re
        from datetime import datetime
        from databricks.sdk import WorkspaceClient
        import io

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        w = WorkspaceClient()

        name_slug = re.sub(r"[^a-zA-Z0-9]+", "_", (control_name or "report")).strip("_").lower()[:50]
        ctrl_prefix = (control_id or "report").replace("-", "_").upper()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"{ctrl_prefix}_{name_slug}_{timestamp}"

        saved_files = []
        volume_url = None
        proj = project_dir or project_path

        if run_id and proj:
            vol_path = f"{PROJECTS_BASE_PATH}/{proj}/runs/{run_id}/{base_name}.md"
            try:
                w.files.upload(vol_path, io.BytesIO(report_content.encode("utf-8")), overwrite=True)
                saved_files.append(vol_path)
                volume_url = vol_path
            except Exception:
                pass

        report_url = ""
        if app_base_url and proj and run_id:
            report_url = f"{app_base_url.rstrip('/')}/api/artifacts/{proj}/{run_id}/{base_name}.md"

        return json.dumps({
            "status": "saved",
            "files": saved_files,
            "volume_url": volume_url,
            "report_url": report_url,
            "filename": f"{base_name}.md",
            "report_length": len(report_content),
        }, indent=2)
    """),
)


# ═══════════════════════════════════════════════════════════════════════
# TIER 3 — Excel (openpyxl + databricks-sdk)
# ═══════════════════════════════════════════════════════════════════════

TIER3 = {}

TIER3["parse_workbook"] = dict(
    params="file_path STRING",
    comment="Parse the engagement workbook (XLSX) and extract all tabs. Also detects embedded images. Args: file_path – path to workbook or project directory name. Returns JSON with tab_names, tabs data, sampling_config, selected_sample, testing_attributes, and has_embedded_images.",
    deps=["openpyxl", "databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import io
        import zipfile
        import openpyxl
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        w = WorkspaceClient()

        if not file_path.endswith(".xlsx"):
            file_path = f"{PROJECTS_BASE_PATH}/{file_path}/engagement_workbook.xlsx"

        resp = w.files.download(file_path)
        wb_content = resp.contents.read()

        wb = openpyxl.load_workbook(io.BytesIO(wb_content), data_only=True)
        tabs = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            tabs[sheet_name] = rows

        wb2 = openpyxl.load_workbook(io.BytesIO(wb_content))
        image_counts = {}
        total_images = 0
        for sheet_name in wb2.sheetnames:
            ws = wb2[sheet_name]
            count = len(ws._images)
            if count > 0:
                image_counts[sheet_name] = count
                total_images += count

        zip_image_count = 0
        try:
            with zipfile.ZipFile(io.BytesIO(wb_content), "r") as zf:
                zip_image_count = sum(1 for n in zf.namelist() if n.startswith("xl/media/"))
        except Exception:
            pass

        result = {
            "tab_names": list(tabs.keys()),
            "tabs": {},
            "sampling_config": {},
            "selected_sample": [],
            "testing_attributes": [],
            "has_embedded_images": total_images > 0 or zip_image_count > 0,
            "image_counts_by_tab": image_counts,
            "total_embedded_images": max(total_images, zip_image_count),
        }

        for tab_name, rows in tabs.items():
            if not rows:
                result["tabs"][tab_name] = {"headers": [], "row_count": 0, "data_preview": []}
                continue
            first_data_row = 0
            for i, row in enumerate(rows):
                non_empty = [c for c in row if c and c.strip()]
                if len(non_empty) >= 2:
                    first_data_row = i
                    break
            headers = rows[first_data_row] if first_data_row < len(rows) else []
            data_rows = rows[first_data_row + 1:] if first_data_row + 1 < len(rows) else []
            non_empty_rows = [r for r in data_rows if any(c.strip() for c in r)]
            result["tabs"][tab_name] = {
                "headers": headers,
                "row_count": len(non_empty_rows),
                "data_preview": non_empty_rows[:30],
            }

        for tab_name, rows in tabs.items():
            lower = tab_name.lower()
            if "sampling" in lower or "sample" in lower:
                in_config = True
                in_sample = False
                sample_headers = []
                for row in rows:
                    if row[0] and "Selected Sample" in row[0]:
                        in_config = False
                        in_sample = True
                        continue
                    if in_config and row[0] and row[0].strip() and row[1] and row[1].strip():
                        key = row[0].strip()
                        if key not in ("Sampling", "Sampling Methodology"):
                            result["sampling_config"][key] = row[1].strip()
                    if in_sample and not sample_headers and any(c.strip() for c in row):
                        sample_headers = [h for h in row if h.strip()]
                        continue
                    if in_sample and sample_headers and any(c.strip() for c in row):
                        item = {}
                        for i, h in enumerate(sample_headers):
                            if i < len(row):
                                item[h] = row[i]
                        if any(v.strip() for v in item.values()):
                            result["selected_sample"].append(item)
            if "testing" in lower and "table" in lower:
                for row in rows:
                    if len(row) >= 2 and row[0].strip() and len(row[0].strip()) <= 2:
                        ref = row[0].strip()
                        if ref.isalpha() and ref.isupper():
                            result["testing_attributes"].append({
                                "ref": ref,
                                "attribute": row[1].strip() if len(row) > 1 else "",
                                "procedure": row[2].strip() if len(row) > 2 else "",
                                "answer": row[3].strip() if len(row) > 3 else "",
                            })
        return json.dumps(result, indent=2)
    """),
)

TIER3["extract_workbook_images"] = dict(
    params="file_path STRING, context STRING",
    comment="Extract and analyze images embedded inside an Excel workbook using the vision LLM. Call AFTER parse_workbook if has_embedded_images is true. Args: file_path – path to workbook or project dir; context – control context for the LLM.",
    deps=["openpyxl", "databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import io
        import base64
        import zipfile
        import openpyxl
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        VISION_LLM_ENDPOINT = "{CFG_VISION_LLM_ENDPOINT}"
        w = WorkspaceClient()

        if not file_path.endswith(".xlsx"):
            file_path = f"{PROJECTS_BASE_PATH}/{file_path}/engagement_workbook.xlsx"

        resp = w.files.download(file_path)
        wb_bytes = resp.contents.read()
        images_found = []

        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for idx, img in enumerate(ws._images):
                try:
                    anchor_cell = f"{img.anchor._from.col}:{img.anchor._from.row}" if hasattr(img.anchor, '_from') else "unknown"
                except Exception:
                    anchor_cell = "unknown"
                images_found.append({"sheet": sheet_name, "anchor": anchor_cell, "source": "openpyxl", "data": img._data(), "index": idx})

        if not images_found:
            try:
                with zipfile.ZipFile(io.BytesIO(wb_bytes), "r") as zf:
                    for name in sorted(zf.namelist()):
                        if name.startswith("xl/media/"):
                            images_found.append({"sheet": "archive", "anchor": name, "source": "zip_media", "data": zf.read(name), "index": len(images_found)})
            except Exception:
                pass

        ctx = context or "GSK FRMC control testing"
        analyses = []
        for img_info in images_found:
            b64 = base64.b64encode(img_info["data"]).decode("utf-8")
            header = img_info["data"][:8]
            mime = "image/png" if header[:4] == b'\\x89PNG' else ("image/jpeg" if header[:2] == b'\\xff\\xd8' else "image/png")

            prompt = (
                f"You are analyzing an image extracted from an Excel workbook for {ctx}.\\n"
                f"This image was found in sheet '{img_info['sheet']}' at position '{img_info['anchor']}'.\\n\\n"
                f"Analyze the image and report:\\n"
                f"1. What does this image show? (dashboard, screenshot, photo, chart)\\n"
                f"2. Key data visible (readings, statuses, dates, measurements)\\n"
                f"3. Any EXCEPTIONS (values exceeding limits, red flags, non-compliance)\\n"
                f"4. Relevance to the control being tested\\n"
            )

            response = w.serving_endpoints.query(
                name=VISION_LLM_ENDPOINT,
                messages=[{"role": "user", "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                ]}],
                max_tokens=4096,
                temperature=0,
            )
            analysis_text = response.choices[0].message.content

            analyses.append({
                "sheet": img_info["sheet"],
                "anchor": img_info["anchor"],
                "image_format": mime,
                "size_bytes": len(img_info["data"]),
                "analysis": analysis_text,
            })

        return json.dumps({"file_path": file_path, "images_extracted": len(analyses), "analyses": analyses}, indent=2)
    """),
)

TIER3["fill_workbook"] = dict(
    params="project_path STRING, test_results_json STRING, control_id STRING, run_id STRING, project_dir STRING, app_base_url STRING",
    comment="Fill the engagement workbook with test results and exceptions. Args: project_path – project dir name; test_results_json – JSON array of test results from aggregate_test_results; control_id – control ID for naming; run_id – current run ID; project_dir – project directory; app_base_url – app base URL. Returns JSON with saved file paths and workbook_url.",
    deps=["openpyxl", "databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import io
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        w = WorkspaceClient()

        wb_path = f"{PROJECTS_BASE_PATH}/{project_path}/engagement_workbook.xlsx"
        resp = w.files.download(wb_path)
        wb_content = resp.contents.read()
        wb = openpyxl.load_workbook(io.BytesIO(wb_content))

        test_results = json.loads(test_results_json)
        results_by_ref = {}
        for tr in test_results:
            ref = tr.get("ref", "").strip().upper()
            if ref:
                if ref not in results_by_ref:
                    results_by_ref[ref] = tr
                else:
                    existing = results_by_ref[ref]
                    existing["narrative"] = f"{existing.get('narrative', '')}\\n{tr.get('narrative', '')}".strip()
                    if tr.get("result", "").lower() == "fail":
                        existing["result"] = "Fail"
                    for exc in tr.get("exceptions", []):
                        existing.setdefault("exceptions", []).append(exc)

        pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        partial_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        header_font = Font(bold=True, size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")

        testing_sheet = None
        for sn in wb.sheetnames:
            if "testing" in sn.lower() and "table" in sn.lower():
                testing_sheet = wb[sn]
                break
        if not testing_sheet:
            for sn in wb.sheetnames:
                if "testing" in sn.lower():
                    testing_sheet = wb[sn]
                    break

        if testing_sheet:
            header_row = None
            ref_col = None
            answer_col = None
            procedure_col = None
            for row_idx, row in enumerate(testing_sheet.iter_rows(min_row=1, max_row=testing_sheet.max_row), start=1):
                for cell in row:
                    val = str(cell.value or "").lower().strip()
                    if val == "ref":
                        header_row = row_idx
                        ref_col = cell.column
                    elif val in ("answer", "result"):
                        answer_col = cell.column
                    elif val == "procedure":
                        procedure_col = cell.column
            if header_row and ref_col:
                if not answer_col:
                    answer_col = (procedure_col or ref_col) + 2
                    testing_sheet.cell(row=header_row, column=answer_col, value="Answer").font = header_font
                result_col = answer_col + 1
                testing_sheet.cell(row=header_row, column=result_col, value="Result").font = header_font
                for row_idx in range(header_row + 1, testing_sheet.max_row + 1):
                    ref_val = str(testing_sheet.cell(row=row_idx, column=ref_col).value or "").strip().upper()
                    if ref_val in results_by_ref:
                        tr = results_by_ref[ref_val]
                        result_text = tr.get("result", "")
                        narrative = tr.get("narrative", "")
                        result_cell = testing_sheet.cell(row=row_idx, column=result_col)
                        result_cell.value = result_text
                        result_cell.alignment = wrap_align
                        if result_text.lower() == "pass":
                            result_cell.fill = pass_fill
                        elif result_text.lower() == "fail":
                            result_cell.fill = fail_fill
                        else:
                            result_cell.fill = partial_fill
                        answer_cell = testing_sheet.cell(row=row_idx, column=answer_col)
                        answer_cell.value = narrative
                        answer_cell.alignment = wrap_align
                testing_sheet.column_dimensions[openpyxl.utils.get_column_letter(answer_col)].width = 60
                testing_sheet.column_dimensions[openpyxl.utils.get_column_letter(result_col)].width = 15

        issue_sheet = None
        for sn in wb.sheetnames:
            if "issue" in sn.lower():
                issue_sheet = wb[sn]
                break

        agg = {}
        if issue_sheet:
            for tr in test_results:
                ref = tr.get("ref", "?").strip().upper()
                if tr.get("result", "").lower() not in ("fail", "partial"):
                    continue
                exceptions = tr.get("exceptions", [])
                raw_samples = tr.get("sample_items_tested", [])
                if isinstance(raw_samples, str):
                    raw_samples = [raw_samples]
                samples_set = set(str(s) for s in raw_samples if s)
                if ref not in agg:
                    best_exc = exceptions[0] if exceptions else {}
                    agg[ref] = {
                        "description": best_exc.get("description", tr.get("narrative", "Test failed")),
                        "severity": best_exc.get("severity", "Medium"),
                        "samples": samples_set,
                        "root_cause": best_exc.get("root_cause", "See test narrative"),
                        "remediation": best_exc.get("remediation", "Review and remediate"),
                    }
                else:
                    entry = agg[ref]
                    entry["samples"] |= samples_set
                    for exc in exceptions:
                        if exc.get("description"):
                            entry["description"] = f"{entry['description']}; {exc['description']}"

            issue_counter = 1
            next_row = 3
            for ref in sorted(agg.keys()):
                entry = agg[ref]
                issue_id = f"ISS-{control_id or 'CTRL'}-{issue_counter:03d}"
                issue_sheet.cell(row=next_row, column=1, value=issue_id)
                issue_sheet.cell(row=next_row, column=2, value=ref)
                issue_sheet.cell(row=next_row, column=3, value=entry.get("severity", "Medium"))
                desc_text = entry.get("description", "")[:32000]
                desc_cell = issue_sheet.cell(row=next_row, column=4, value=desc_text)
                desc_cell.alignment = wrap_align
                issue_sheet.cell(row=next_row, column=5, value=", ".join(sorted(entry.get("samples", set()))))
                issue_sheet.cell(row=next_row, column=6, value=entry.get("root_cause", ""))
                issue_sheet.cell(row=next_row, column=7, value=entry.get("remediation", ""))
                issue_sheet.cell(row=next_row, column=8, value="")
                issue_sheet.cell(row=next_row, column=9, value="")
                issue_sheet.cell(row=next_row, column=10, value="Open")
                next_row += 1
                issue_counter += 1
            if issue_counter == 1:
                issue_sheet.cell(row=3, column=1, value="No exceptions identified")
                issue_sheet.cell(row=3, column=3, value="N/A")
                issue_sheet.cell(row=3, column=10, value="Closed")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ctrl_prefix = (control_id or "workbook").replace("-", "_").upper()
        output_filename = f"{ctrl_prefix}_completed_{timestamp}.xlsx"

        output_buf = io.BytesIO()
        wb.save(output_buf)
        output_bytes = output_buf.getvalue()

        saved_files = []
        volume_url = None
        proj = project_dir or project_path
        if run_id and proj:
            vol_path = f"{PROJECTS_BASE_PATH}/{proj}/runs/{run_id}/{output_filename}"
            try:
                w.files.upload(vol_path, io.BytesIO(output_bytes), overwrite=True)
                saved_files.append(vol_path)
                volume_url = vol_path
            except Exception:
                pass

        workbook_url = ""
        if app_base_url and proj and run_id:
            workbook_url = f"{app_base_url.rstrip('/')}/api/artifacts/{proj}/{run_id}/{output_filename}"

        return json.dumps({
            "status": "saved",
            "filename": output_filename,
            "files": saved_files,
            "volume_url": volume_url,
            "workbook_url": workbook_url,
            "attrs_filled": len(results_by_ref),
            "exceptions_logged": len(agg),
        }, indent=2)
    """),
)


# ═══════════════════════════════════════════════════════════════════════
# TIER 4 — LLM-powered tools (databricks-sdk)
# ═══════════════════════════════════════════════════════════════════════

TIER4 = {}

TIER4["review_document"] = dict(
    params="file_path STRING, context STRING, focus_area STRING",
    comment="Review a supporting PDF document using the LLM. Args: file_path – path to the PDF; context – control context from engagement; focus_area – what to look for (pass empty string for general review). Returns JSON with analysis.",
    deps=["PyMuPDF", "databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        FAST_LLM_ENDPOINT = "{CFG_FAST_LLM_ENDPOINT}"
        w = WorkspaceClient()

        if not file_path.startswith("/"):
            file_path = f"{PROJECTS_BASE_PATH}/{file_path}"

        resp = w.files.download(file_path)
        content = resp.contents.read()
        ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
        text_content = None

        if ext == "pdf":
            try:
                import fitz
                doc = fitz.open(stream=content, filetype="pdf")
                text_content = ""
                for page in doc:
                    text_content += page.get_text() + "\\n"
                doc.close()
            except ImportError:
                text_content = f"[PDF at {file_path}, {len(content)} bytes]"
        else:
            text_content = content.decode("utf-8", errors="replace")

        focus = focus_area or "general compliance review"
        ctx = context or "GSK FRMC control testing"
        prompt = (
            f"You are reviewing a document for {ctx}.\\n"
            f"Focus area: {focus}\\n\\n"
            f"Extract and report:\\n"
            f"1. Document type and purpose\\n"
            f"2. Key data points (amounts, dates, document numbers, user IDs)\\n"
            f"3. Any approvals, sign-offs, or authorizations visible\\n"
            f"4. Any EXCEPTIONS noted (policy violations, missing items, overdue actions)\\n"
            f"5. Any thresholds or limits mentioned and whether they are breached\\n"
            f"6. Summary of findings relevant to {focus}\\n\\n"
            f"Document text content:\\n\\n{text_content[:10000]}"
        )

        llm_response = w.serving_endpoints.query(
            name=FAST_LLM_ENDPOINT,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=4096, temperature=0,
        )

        return json.dumps({
            "file_path": file_path,
            "document_type": ext,
            "file_size_bytes": len(content),
            "review_focus": focus,
            "analysis": llm_response.choices[0].message.content,
        }, indent=2)
    """),
)

TIER4["review_screenshot"] = dict(
    params="file_path STRING, context STRING, focus_area STRING",
    comment="Review a screenshot or photo using the vision LLM. Args: file_path – path to image (PNG, JPG); context – control context; focus_area – what to look for (pass empty string for general). Returns JSON with analysis.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import base64
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        VISION_LLM_ENDPOINT = "{CFG_VISION_LLM_ENDPOINT}"
        w = WorkspaceClient()

        if not file_path.startswith("/"):
            file_path = f"{PROJECTS_BASE_PATH}/{file_path}"

        resp = w.files.download(file_path)
        content = resp.contents.read()
        ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else "png"

        b64 = base64.b64encode(content).decode("utf-8")
        mime = "image/png" if ext == "png" else "image/jpeg"

        focus = focus_area or "general visual inspection"
        ctx = context or "GSK FRMC control testing"
        prompt = (
            f"You are reviewing a screenshot/photo for {ctx}.\\n"
            f"Focus area: {focus}\\n\\n"
            f"Analyze the image and report:\\n"
            f"1. What the screenshot/photo shows\\n"
            f"2. Key data visible (dates, statuses, user IDs, names, quantities)\\n"
            f"3. Any status indicators (completed, overdue, pending, flagged)\\n"
            f"4. Any EXCEPTIONS visible (red flags, overdue items, policy violations)\\n"
            f"5. Summary of findings relevant to {focus}\\n"
        )

        llm_response = w.serving_endpoints.query(
            name=VISION_LLM_ENDPOINT,
            messages=[{"role": "user", "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
            ]}],
            max_tokens=4096, temperature=0,
        )

        return json.dumps({
            "file_path": file_path,
            "image_type": ext,
            "file_size_bytes": len(content),
            "review_focus": focus,
            "analysis": llm_response.choices[0].message.content,
        }, indent=2)
    """),
)

TIER4["analyze_email"] = dict(
    params="file_path STRING, context STRING, focus_area STRING",
    comment="Parse and analyze an email file (.eml) for compliance evidence. Args: file_path – path to .eml file; context – control context; focus_area – what to look for. Returns JSON with email metadata and analysis.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import email
        from email import policy as email_policy
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        FAST_LLM_ENDPOINT = "{CFG_FAST_LLM_ENDPOINT}"
        w = WorkspaceClient()

        if not file_path.startswith("/"):
            file_path = f"{PROJECTS_BASE_PATH}/{file_path}"

        resp = w.files.download(file_path)
        content = resp.contents.read()

        msg = email.message_from_bytes(content, policy=email_policy.default)
        email_from = str(msg.get("From", ""))
        email_to = str(msg.get("To", ""))
        email_subject = str(msg.get("Subject", ""))
        email_date = str(msg.get("Date", ""))

        body_text = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    body_text = part.get_content()
                    break
        else:
            body_text = msg.get_content()

        focus = focus_area or "authorization and approval"
        ctx = context or "GSK FRMC control testing"
        prompt = (
            f"You are analyzing an email for {ctx}.\\n"
            f"Focus area: {focus}\\n\\n"
            f"Email metadata:\\n"
            f"- From: {email_from}\\n"
            f"- To: {email_to}\\n"
            f"- Subject: {email_subject}\\n"
            f"- Date: {email_date}\\n\\n"
            f"Email body:\\n{body_text[:5000]}\\n\\n"
            f"Analyze and report:\\n"
            f"1. Who is the sender and their role/authority level?\\n"
            f"2. Who is the recipient?\\n"
            f"3. What action is being requested or confirmed?\\n"
            f"4. Is this a proper authorization/approval?\\n"
            f"5. Any EXCEPTIONS: self-approval, unauthorized approver, missing info, policy violations\\n"
            f"6. Summary of compliance findings\\n"
        )

        llm_response = w.serving_endpoints.query(
            name=FAST_LLM_ENDPOINT,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=4096, temperature=0,
        )

        return json.dumps({
            "file_path": file_path,
            "email_from": email_from,
            "email_to": email_to,
            "email_subject": email_subject,
            "email_date": email_date,
            "review_focus": focus,
            "analysis": llm_response.choices[0].message.content,
        }, indent=2)
    """),
)

TIER4["execute_test"] = dict(
    params="test_ref STRING, attribute STRING, procedure STRING, control_context STRING, sample_item_json STRING, evidence_summary STRING",
    comment="Execute a specific testing attribute against a SINGLE sample item. Args: test_ref – attribute ref (A, B, C...); attribute – the testing question; procedure – testing procedure; control_context – JSON with control_id, rules; sample_item_json – JSON of sample item; evidence_summary – summary of reviewed docs. Returns JSON with test_ref, sample_item, and llm_analysis.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import time
        from databricks.sdk import WorkspaceClient

        LLM_ENDPOINT = "{CFG_LLM_ENDPOINT}"
        w = WorkspaceClient()

        sample_item = json.loads(sample_item_json) if sample_item_json else {}

        # Pre-checks
        findings = []
        attr_lower = attribute.lower()
        try:
            ctx = json.loads(control_context) if isinstance(control_context, str) else control_context
        except Exception:
            ctx = {}
        rules = ctx.get("rules", {})

        preparer = sample_item.get("Preparer", "")
        approver = sample_item.get("Approver", "")
        if preparer and approver and ("self-approval" in attr_lower or "dual auth" in attr_lower):
            if preparer.strip().lower() == approver.strip().lower():
                findings.append(f"DATA CHECK FAIL: Preparer ({preparer}) = Approver ({approver}). Self-approval detected.")
            else:
                findings.append(f"DATA CHECK PASS: Preparer ({preparer}) != Approver ({approver}). Dual authorization verified.")

        threshold = rules.get("threshold_gbp", rules.get("threshold_usd", 0))
        if threshold and ("threshold" in attr_lower or "above" in attr_lower):
            for key in ("Amount_GBP", "Amount_in_GBP", "Amount"):
                if key in sample_item:
                    try:
                        amt = float(str(sample_item[key]).replace(",", "").replace("\\u00a3", "").replace("$", ""))
                        if amt >= float(threshold):
                            findings.append(f"DATA CHECK: Amount ({amt:,.2f}) >= threshold ({threshold:,}). Finance Director review required.")
                        else:
                            findings.append(f"DATA CHECK: Amount ({amt:,.2f}) < threshold ({threshold:,}). Below threshold.")
                    except (ValueError, TypeError):
                        pass
                    break

        if "supporting doc" in attr_lower or "documentation" in attr_lower:
            sup = sample_item.get("Supporting_Doc", sample_item.get("supporting_doc", ""))
            if isinstance(sup, str) and not sup.strip():
                findings.append("DATA CHECK FAIL: Supporting_Doc field is empty.")
            elif sup:
                findings.append(f"DATA CHECK PASS: Supporting document referenced: {sup}")

        pre_checks = "\\n".join(findings) if findings else ""

        evidence_with_checks = evidence_summary[:3000]
        if pre_checks:
            evidence_with_checks = f"**Automated Data Pre-Checks:**\\n{pre_checks}\\n\\n{evidence_with_checks}"

        TEST_PROMPT = (
            f"Execute testing attribute {test_ref} for the control described below.\\n\\n"
            f"**Control Context**:\\n{control_context[:2000]}\\n\\n"
            f"**Testing Attribute**: {attribute}\\n"
            f"**Testing Procedure**: {procedure}\\n\\n"
            f"**Sample Item Being Tested**:\\n{sample_item_json[:2000]}\\n\\n"
            f"**Available Evidence from Reviewed Documents**:\\n{evidence_with_checks}\\n\\n"
            "Perform the test and respond with a JSON object containing these fields:\\n"
            '- "result": "Pass" or "Fail" or "Not Applicable"\\n'
            '- "narrative": A detailed narrative. Start with "Yes, ..." or "No, ..." or "Not Applicable, ...". Reference specific identifiers.\\n'
            '- "exception": null if Pass, or a description of the exception if Fail.\\n'
            '- "severity": null if Pass, or "Critical" / "High" / "Medium" / "Low" if Fail.\\n'
            '- "confidence": "High", "Medium", or "Low".\\n'
            '- "confidence_rationale": One sentence explaining why.\\n\\n'
            "Respond ONLY with the JSON object, no other text."
        )

        max_retries = 4
        backoff_base = 8
        response = None
        for attempt in range(max_retries + 1):
            try:
                response = w.serving_endpoints.query(
                    name=LLM_ENDPOINT,
                    messages=[{"role": "user", "content": TEST_PROMPT}],
                    max_tokens=4096, temperature=0,
                )
                break
            except Exception as e:
                err_str = str(e)
                is_rate_limit = "429" in err_str or "REQUEST_LIMIT_EXCEEDED" in err_str or "rate limit" in err_str.lower()
                if is_rate_limit and attempt < max_retries:
                    time.sleep(backoff_base * (2 ** attempt))
                    continue
                raise

        return json.dumps({
            "test_ref": test_ref,
            "sample_item": sample_item,
            "llm_analysis": response.choices[0].message.content,
        }, indent=2)
    """),
)

TIER4["compile_results"] = dict(
    params="control_id STRING, control_name STRING, engagement_number STRING, domain STRING, population_size INT, sample_size INT, testing_attributes_json STRING, test_results_json STRING, rules_json STRING",
    comment="Compile all test results into the final assessment report. Returns formatted markdown report.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import time
        from databricks.sdk import WorkspaceClient

        LLM_ENDPOINT = "{CFG_LLM_ENDPOINT}"
        w = WorkspaceClient()

        REPORT_PROMPT = (
            f"Generate the final Controls Evidence Review Report based on completed testing.\\n\\n"
            f"**Control**: {control_id} - {control_name}\\n"
            f"**Domain**: {domain}\\n"
            f"**Engagement**: {engagement_number}\\n"
            f"**Population Size**: {population_size}\\n"
            f"**Sample Size**: {sample_size}\\n\\n"
            f"**Testing Attributes**:\\n{testing_attributes_json[:3000]}\\n\\n"
            f"**Control-Specific Rules**:\\n{rules_json[:2000]}\\n\\n"
            f"**Test Results**:\\n{test_results_json[:8000]}\\n\\n"
            "Produce a structured report with these sections:\\n\\n"
            "## 1. Executive Summary\\n2-3 sentences summarising the overall assessment.\\n\\n"
            "## 2. Testing Scope and Methodology\\nDescribe the population, sampling approach, and testing attributes.\\n\\n"
            "## 3. Control-Level Summary\\nFor each testing attribute, provide the narrative summary.\\n\\n"
            "## 4. Results Summary Table\\n| Attribute | Result | Confidence | Severity | Exceptions |\\n|-----------|--------|------------|----------|------------|\\n\\n"
            "## 5. Exception Details\\nFor each exception: Issue ID, Attribute, Severity, Description, Affected Samples, Root Cause, Remediation.\\n\\n"
            "## 6. Overall Control Assessment\\n**Effective** / **Effective with Exceptions** / **Ineffective** with justification.\\n\\n"
            "If any test result has **Low** confidence, recommend those attributes for manual re-review.\\n\\n"
            "## 7. Issue Template (Ready to Paste)\\n| Issue_ID | Testing_Attribute | Severity | Description | Affected_Samples | Root_Cause | Remediation | Owner | Due_Date | Status |"
        )

        max_retries = 4
        backoff_base = 8
        response = None
        for attempt in range(max_retries + 1):
            try:
                response = w.serving_endpoints.query(
                    name=LLM_ENDPOINT,
                    messages=[{"role": "user", "content": REPORT_PROMPT}],
                    max_tokens=8192, temperature=0,
                )
                break
            except Exception as e:
                err_str = str(e)
                is_rate_limit = "429" in err_str or "REQUEST_LIMIT_EXCEEDED" in err_str or "rate limit" in err_str.lower()
                if is_rate_limit and attempt < max_retries:
                    time.sleep(backoff_base * (2 ** attempt))
                    continue
                raise

        return response.choices[0].message.content
    """),
)


# ═══════════════════════════════════════════════════════════════════════
# TIER 5 — Email (smtplib + databricks-sdk)
# ═══════════════════════════════════════════════════════════════════════

TIER5 = {}

TIER5["send_email"] = dict(
    params="to STRING, subject STRING, body STRING, cc STRING, importance STRING, project_path STRING, report_url STRING, attach_workbook STRING, run_id STRING, project_dir STRING",
    comment="Send a professionally formatted email notification with the report. Args: to – recipient(s) comma-separated; subject – email subject; body – markdown-style body text; cc – CC recipients; importance – low/normal/high; project_path – project dir; report_url – URL to the report; attach_workbook – true/false; run_id – current run ID; project_dir – project directory. Returns JSON with status.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import re
        import io
        import base64
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.mime.application import MIMEApplication
        from email.message import EmailMessage
        from datetime import datetime
        from databricks.sdk import WorkspaceClient

        SMTP_EMAIL = "{CFG_SMTP_EMAIL}"
        SMTP_DISPLAY_NAME = "GSK Compliance Agent"
        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        w = WorkspaceClient()

        # Fetch SMTP password from secrets
        smtp_password = ""
        try:
            resp = w.secrets.get_secret(scope="gsk-compliance", key="smtp-app-password")
            if resp.value:
                smtp_password = base64.b64decode(resp.value).decode("utf-8")
        except Exception:
            pass

        # Find workbook attachment
        xlsx_bytes = None
        xlsx_name = ""
        should_attach = attach_workbook.lower() in ("true", "1", "yes") if attach_workbook else True
        if should_attach and run_id and (project_dir or project_path):
            proj = project_dir or project_path
            run_base = f"{PROJECTS_BASE_PATH}/{proj}/runs/{run_id}"
            try:
                items = list(w.files.list_directory_contents(run_base))
                xlsx_items = [it for it in items if it.path and it.path.endswith(".xlsx") and "_completed_" in it.path]
                if xlsx_items:
                    xlsx_items.sort(key=lambda x: x.path, reverse=True)
                    chosen = xlsx_items[0].path
                    xlsx_name = chosen.rstrip("/").split("/")[-1]
                    xlsx_bytes = w.files.download(chosen).contents.read()
            except Exception:
                pass

        # Build HTML email
        body_lines = body.strip().splitlines()
        body_html_parts = []
        for line in body_lines:
            stripped = line.strip()
            if not stripped:
                body_html_parts.append('<div style="height:12px"></div>')
            elif stripped.startswith("# "):
                body_html_parts.append(f'<h2 style="color:#1a1a2e;margin:18px 0 8px;font-size:18px;border-bottom:1px solid #e5e7eb;padding-bottom:6px">{stripped[2:]}</h2>')
            elif stripped.startswith("## "):
                body_html_parts.append(f'<h3 style="color:#333;margin:14px 0 6px;font-size:15px">{stripped[3:]}</h3>')
            elif stripped.startswith("- ") or stripped.startswith("\\u2022 "):
                body_html_parts.append(f'<div style="padding:2px 0 2px 16px;color:#374151">&#8226; {stripped[2:]}</div>')
            else:
                m = re.match(r"^\\*\\*(.+?)\\*\\*:?\\s*(.*)$", stripped)
                if m:
                    body_html_parts.append(f'<div style="padding:2px 0;color:#374151"><strong>{m.group(1)}</strong>: {m.group(2)}</div>')
                else:
                    body_html_parts.append(f'<div style="padding:2px 0;color:#374151">{stripped}</div>')

        body_html = "\\n".join(body_html_parts)
        body_html = re.sub(r"\\*\\*(.+?)\\*\\*", r"<strong>\\1</strong>", body_html)

        report_section = ""
        if report_url:
            report_section = f'<div style="margin:24px 0;text-align:center"><a href="{report_url}" style="display:inline-block;background:#1a73e8;color:#ffffff;padding:12px 28px;border-radius:6px;text-decoration:none;font-weight:600;font-size:14px">View Full Report &rarr;</a></div>'

        attachment_note = ""
        if xlsx_name:
            attachment_note = f'<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:10px 14px;margin:16px 0;font-size:13px;color:#166534">Completed workbook attached: <strong>{xlsx_name}</strong></div>'

        html_body = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"/></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif">
<div style="max-width:640px;margin:0 auto;background:#ffffff;border-radius:8px;overflow:hidden;margin-top:20px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,0.1)">
<div style="background:linear-gradient(135deg,#1a1a2e 0%,#16213e 100%);padding:24px 32px">
<div style="color:#ffffff;font-size:20px;font-weight:700">GSK Controls Evidence Review</div>
<div style="color:#94a3b8;font-size:12px;margin-top:4px">Automated Compliance Assessment</div></div>
<div style="padding:24px 32px;font-size:14px;line-height:1.65;color:#374151">{body_html}{report_section}{attachment_note}</div>
<div style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:16px 32px;font-size:11px;color:#9ca3af;text-align:center">This is an automated notification from the GSK Compliance Agent.</div>
</div></body></html>'''

        if smtp_password and SMTP_EMAIL:
            msg = MIMEMultipart("mixed")
            msg["From"] = f"{SMTP_DISPLAY_NAME} <{SMTP_EMAIL}>"
            msg["To"] = to
            msg["Subject"] = subject
            if cc:
                msg["Cc"] = cc
            if importance == "high":
                msg["X-Priority"] = "1"
            msg.attach(MIMEText(html_body, "html"))
            if xlsx_bytes and xlsx_name:
                part = MIMEApplication(xlsx_bytes, Name=xlsx_name)
                part["Content-Disposition"] = f'attachment; filename="{xlsx_name}"'
                msg.attach(part)
            all_recipients = [a.strip() for a in to.split(",")]
            if cc:
                all_recipients.extend(a.strip() for a in cc.split(",") if a.strip())
            try:
                with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
                    server.starttls()
                    server.login(SMTP_EMAIL, smtp_password)
                    server.sendmail(SMTP_EMAIL, all_recipients, msg.as_string())
                return json.dumps({"status": "sent", "to": to, "subject": subject, "method": "gmail_smtp", "report_url": report_url, "attached_workbook": xlsx_name or None}, indent=2)
            except Exception as e:
                return json.dumps({"status": "error", "to": to, "subject": subject, "method": "gmail_smtp", "error": str(e), "report_url": report_url}, indent=2)
        else:
            # EML fallback — save to volume
            eml = EmailMessage()
            eml["From"] = f"{SMTP_DISPLAY_NAME} <{SMTP_EMAIL}>"
            eml["To"] = to
            eml["Subject"] = subject
            eml["Date"] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
            if cc:
                eml["Cc"] = cc
            eml.set_content(html_body, subtype="html")
            saved_path = None
            if run_id and (project_dir or project_path):
                proj = project_dir or project_path
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                vol_path = f"{PROJECTS_BASE_PATH}/{proj}/runs/{run_id}/sent_email_{timestamp}.eml"
                try:
                    w.files.upload(vol_path, io.BytesIO(eml.as_string().encode("utf-8")), overwrite=True)
                    saved_path = vol_path
                except Exception:
                    pass
            return json.dumps({"status": "not_sent", "to": to, "subject": subject, "method": "eml_fallback", "note": "SMTP not configured. Email saved as .eml.", "saved_path": saved_path, "report_url": report_url}, indent=2)
    """),
)


# ═══════════════════════════════════════════════════════════════════════
# TIER 6 — Batch tools (parallel execution)
# ═══════════════════════════════════════════════════════════════════════

TIER6 = {}

TIER6["batch_review_evidence"] = dict(
    params="evidence_files_json STRING, project_path STRING, control_context STRING",
    comment="Review ALL evidence files in parallel. Call INSTEAD of reviewing files one-by-one. Dispatches PDFs, images, and emails to the appropriate analysis logic concurrently. Args: evidence_files_json – JSON array of file objects with path, type, focus; project_path – project dir name; control_context – JSON with control_id, rules.",
    deps=["PyMuPDF", "databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import io
        import base64
        import time as _time
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from databricks.sdk import WorkspaceClient

        PROJECTS_BASE_PATH = "{CFG_PROJECTS_BASE_PATH}"
        FAST_LLM_ENDPOINT = "{CFG_FAST_LLM_ENDPOINT}"
        VISION_LLM_ENDPOINT = "{CFG_VISION_LLM_ENDPOINT}"
        MAX_WORKERS = 4

        evidence_files = json.loads(evidence_files_json) if isinstance(evidence_files_json, str) else evidence_files_json
        evidence_files = sorted(evidence_files, key=lambda f: f.get("path", ""))
        start = _time.time()

        def resolve_path(rel_path):
            if rel_path.startswith("/"):
                return rel_path
            return f"{PROJECTS_BASE_PATH}/{project_path}/{rel_path}"

        def review_one(file_info):
            w = WorkspaceClient()
            file_type = file_info.get("type", "pdf")
            rel_path = file_info.get("path", "")
            focus = file_info.get("focus", "")
            file_path = resolve_path(rel_path)

            resp = w.files.download(file_path)
            content = resp.contents.read()
            ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""

            if file_type == "email":
                import email as email_mod
                from email import policy as email_policy
                msg = email_mod.message_from_bytes(content, policy=email_policy.default)
                email_from = str(msg.get("From", ""))
                email_to = str(msg.get("To", ""))
                email_subject = str(msg.get("Subject", ""))
                email_date = str(msg.get("Date", ""))
                body_text = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == "text/plain":
                            body_text = part.get_content()
                            break
                else:
                    body_text = msg.get_content()
                ctx = control_context or "GSK FRMC control testing"
                prompt = (
                    f"You are analyzing an email for {ctx}.\\nFocus area: {focus or 'authorization and approval'}\\n\\n"
                    f"Email metadata:\\n- From: {email_from}\\n- To: {email_to}\\n- Subject: {email_subject}\\n- Date: {email_date}\\n\\n"
                    f"Email body:\\n{body_text[:5000]}\\n\\n"
                    f"Analyze: sender role, recipient, action, proper authorization, exceptions, compliance findings.\\n"
                )
                llm_resp = w.serving_endpoints.query(name=FAST_LLM_ENDPOINT, messages=[{"role": "user", "content": prompt}], max_tokens=4096, temperature=0)
                return {"file_path": file_path, "email_from": email_from, "email_to": email_to, "email_subject": email_subject, "review_focus": focus, "analysis": llm_resp.choices[0].message.content}

            elif file_type in ("screenshot", "image", "photo"):
                b64 = base64.b64encode(content).decode("utf-8")
                mime = "image/png" if ext == "png" else "image/jpeg"
                ctx = control_context or "GSK FRMC control testing"
                prompt = (
                    f"You are reviewing a screenshot/photo for {ctx}.\\nFocus area: {focus or 'general visual inspection'}\\n\\n"
                    f"Analyze: what it shows, key data, status indicators, exceptions, findings.\\n"
                )
                llm_resp = w.serving_endpoints.query(
                    name=VISION_LLM_ENDPOINT,
                    messages=[{"role": "user", "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                    ]}],
                    max_tokens=4096, temperature=0)
                return {"file_path": file_path, "image_type": ext, "review_focus": focus, "analysis": llm_resp.choices[0].message.content}

            else:
                text_content = None
                if ext == "pdf":
                    try:
                        import fitz
                        doc = fitz.open(stream=content, filetype="pdf")
                        text_content = ""
                        for page in doc:
                            text_content += page.get_text() + "\\n"
                        doc.close()
                    except ImportError:
                        text_content = f"[PDF at {file_path}, {len(content)} bytes]"
                else:
                    text_content = content.decode("utf-8", errors="replace")
                ctx = control_context or "GSK FRMC control testing"
                prompt = (
                    f"You are reviewing a document for {ctx}.\\nFocus area: {focus or 'general compliance review'}\\n\\n"
                    f"Extract: document type, key data, approvals, exceptions, thresholds, summary.\\n\\n"
                    f"Document text:\\n\\n{text_content[:10000]}"
                )
                llm_resp = w.serving_endpoints.query(name=FAST_LLM_ENDPOINT, messages=[{"role": "user", "content": prompt}], max_tokens=4096, temperature=0)
                return {"file_path": file_path, "document_type": ext, "review_focus": focus, "analysis": llm_resp.choices[0].message.content}

        reviews = [None] * len(evidence_files)
        errors = []

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            futures = {pool.submit(review_one, f): i for i, f in enumerate(evidence_files)}
            for future in as_completed(futures):
                idx = futures[future]
                try:
                    reviews[idx] = future.result()
                except Exception as e:
                    errors.append(f"{evidence_files[idx].get('path', '?')}: {e}")

        elapsed = round(_time.time() - start, 1)
        reviews = [r for r in reviews if r is not None]

        return json.dumps({"reviews": reviews, "files_reviewed": len(reviews), "errors": errors, "elapsed_seconds": elapsed, "parallel_workers": MAX_WORKERS}, indent=2)
    """),
)

TIER6["batch_execute_tests"] = dict(
    params="test_plan_json STRING, control_context STRING, evidence_summary STRING",
    comment="Execute ALL tests from the test plan in parallel. Call INSTEAD of calling execute_test one-by-one. Args: test_plan_json – full test_plan array from generate_test_plan; control_context – JSON with control_objective and rules; evidence_summary – combined evidence review summaries. Returns JSON with results list, pass/fail counts, and timing.",
    deps=["databricks-sdk"],
    body=textwrap.dedent("""\
        import json
        import time as _time
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from databricks.sdk import WorkspaceClient

        LLM_ENDPOINT = "{CFG_LLM_ENDPOINT}"
        MAX_WORKERS = 3

        test_plan = json.loads(test_plan_json) if isinstance(test_plan_json, str) else test_plan_json
        test_plan = sorted(test_plan, key=lambda e: (e.get("test_ref", ""), e.get("sample_item_json", "")))
        start = _time.time()

        def run_pre_checks(test_ref, attribute, sample_item, control_ctx_str):
            findings = []
            attr_lower = attribute.lower()
            try:
                ctx = json.loads(control_ctx_str) if isinstance(control_ctx_str, str) else control_ctx_str
            except Exception:
                ctx = {}
            rules = ctx.get("rules", {})

            preparer = sample_item.get("Preparer", "")
            approver = sample_item.get("Approver", "")
            if preparer and approver and ("self-approval" in attr_lower or "dual auth" in attr_lower):
                if preparer.strip().lower() == approver.strip().lower():
                    findings.append(f"DATA CHECK FAIL: Preparer ({preparer}) = Approver ({approver}). Self-approval detected.")
                else:
                    findings.append(f"DATA CHECK PASS: Preparer ({preparer}) != Approver ({approver}). Dual authorization verified.")

            threshold = rules.get("threshold_gbp", rules.get("threshold_usd", 0))
            if threshold and ("threshold" in attr_lower or "above" in attr_lower):
                for key in ("Amount_GBP", "Amount_in_GBP", "Amount"):
                    if key in sample_item:
                        try:
                            amt = float(str(sample_item[key]).replace(",", "").replace("\\u00a3", "").replace("$", ""))
                            if amt >= float(threshold):
                                findings.append(f"DATA CHECK: Amount ({amt:,.2f}) >= threshold ({threshold:,}).")
                            else:
                                findings.append(f"DATA CHECK: Amount ({amt:,.2f}) < threshold ({threshold:,}).")
                        except (ValueError, TypeError):
                            pass
                        break
            return "\\n".join(findings)

        def execute_one(entry):
            w = WorkspaceClient()
            ref = entry.get("test_ref", "?")
            attribute = entry.get("attribute", "")
            procedure = entry.get("procedure", "")
            sample_item_json = entry.get("sample_item_json", "{}")
            sample_item = json.loads(sample_item_json) if sample_item_json else {}

            pre_checks = run_pre_checks(ref, attribute, sample_item, control_context)
            evidence_with_checks = evidence_summary[:3000]
            if pre_checks:
                evidence_with_checks = f"**Automated Data Pre-Checks:**\\n{pre_checks}\\n\\n{evidence_with_checks}"

            prompt = (
                f"Execute testing attribute {ref} for the control described below.\\n\\n"
                f"**Control Context**:\\n{control_context[:2000]}\\n\\n"
                f"**Testing Attribute**: {attribute}\\n"
                f"**Testing Procedure**: {procedure}\\n\\n"
                f"**Sample Item Being Tested**:\\n{sample_item_json[:2000]}\\n\\n"
                f"**Available Evidence**:\\n{evidence_with_checks}\\n\\n"
                "Respond with a JSON object: result, narrative, exception, severity, confidence, confidence_rationale.\\n"
                "Respond ONLY with the JSON object."
            )

            max_retries = 4
            backoff_base = 8
            response = None
            for attempt in range(max_retries + 1):
                try:
                    response = w.serving_endpoints.query(
                        name=LLM_ENDPOINT,
                        messages=[{"role": "user", "content": prompt}],
                        max_tokens=4096, temperature=0)
                    break
                except Exception as e:
                    err_str = str(e)
                    is_rate_limit = "429" in err_str or "REQUEST_LIMIT_EXCEEDED" in err_str or "rate limit" in err_str.lower()
                    if is_rate_limit and attempt < max_retries:
                        _time.sleep(backoff_base * (2 ** attempt))
                        continue
                    raise

            return {"test_ref": ref, "sample_item": sample_item, "llm_analysis": response.choices[0].message.content}

        results = [None] * len(test_plan)
        errors = []

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            futures = {pool.submit(execute_one, e): i for i, e in enumerate(test_plan)}
            for future in as_completed(futures):
                idx = futures[future]
                try:
                    results[idx] = future.result()
                except Exception as e:
                    ref = test_plan[idx].get("test_ref", "?")
                    errors.append(f"Test {ref} #{idx}: {e}")

        elapsed = round(_time.time() - start, 1)
        results = [r for r in results if r is not None]

        pass_count = 0
        fail_count = 0
        for r in results:
            try:
                analysis = r.get("llm_analysis", "")
                parsed = json.loads(analysis.strip().strip("`").lstrip("json\\n"))
                if parsed.get("result", "").lower() == "pass":
                    pass_count += 1
                elif parsed.get("result", "").lower() == "fail":
                    fail_count += 1
            except Exception:
                pass

        return json.dumps({
            "results": results,
            "total_tests": len(results),
            "passed": pass_count,
            "failed": fail_count,
            "errors": errors,
            "elapsed_seconds": elapsed,
            "parallel_workers": MAX_WORKERS,
        }, indent=2)
    """),
)


# ═══════════════════════════════════════════════════════════════════════
# Registration engine
# ═══════════════════════════════════════════════════════════════════════

ALL_FUNCTIONS: dict[str, dict] = {}
ALL_FUNCTIONS.update(TIER1)
ALL_FUNCTIONS.update(TIER2)
ALL_FUNCTIONS.update(TIER3)
ALL_FUNCTIONS.update(TIER4)
ALL_FUNCTIONS.update(TIER5)
ALL_FUNCTIONS.update(TIER6)


def register_all(catalog: str, schema: str, cfg: dict, warehouse_id: str | None = None):
    from databricks.sdk import WorkspaceClient
    from databricks.sdk.service.sql import StatementState

    w = WorkspaceClient()

    if not warehouse_id:
        warehouses = list(w.warehouses.list())
        running = [wh for wh in warehouses if wh.state and wh.state.value == "RUNNING"]
        if running:
            running.sort(key=lambda x: x.name or "")
            warehouse_id = running[0].id
        else:
            starting = [wh for wh in warehouses if wh.state and wh.state.value in ("STARTING", "STOPPING")]
            if starting:
                warehouse_id = starting[0].id
            elif warehouses:
                warehouse_id = warehouses[0].id
            else:
                print("ERROR: No SQL warehouses found. Cannot register UC functions.")
                sys.exit(1)

    print(f"\n  Using warehouse: {warehouse_id}")
    print(f"  Registering {len(ALL_FUNCTIONS)} functions in `{catalog}`.`{schema}`\n")

    success = 0
    failed = 0

    for name, defn in ALL_FUNCTIONS.items():
        sql = _build_sql(catalog, schema, name, defn, cfg)
        print(f"  [{success + failed + 1:2d}/{len(ALL_FUNCTIONS)}] {name} ...", end=" ", flush=True)

        try:
            resp = w.statement_execution.execute_statement(
                warehouse_id=warehouse_id,
                statement=sql,
                wait_timeout="50s",
            )
            # Poll if still running (DDL can take longer than the initial wait)
            poll_limit = 24
            while (
                resp.status
                and resp.status.state in (StatementState.PENDING, StatementState.RUNNING)
                and poll_limit > 0
            ):
                _time.sleep(5)
                resp = w.statement_execution.get_statement(resp.statement_id)
                poll_limit -= 1
            if resp.status and resp.status.state == StatementState.FAILED:
                print(f"FAILED: {resp.status.error}")
                failed += 1
            elif resp.status and resp.status.state == StatementState.SUCCEEDED:
                print("OK")
                success += 1
            else:
                print(f"OK (state: {resp.status.state.value if resp.status else 'unknown'})")
                success += 1
        except Exception as e:
            print(f"ERROR: {e}")
            failed += 1

        _time.sleep(0.2)

    print(f"\n  Done: {success} registered, {failed} failed")
    return failed == 0


# ═══════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Register all compliance agent tools as UC Python functions.",
    )
    parser.add_argument("--catalog", required=True, help="Unity Catalog catalog name")
    parser.add_argument("--schema", default="gsk_compliance", help="UC schema name")
    parser.add_argument("--volume", default="evidence_files", help="UC volume name")
    parser.add_argument("--llm-endpoint", default="databricks-claude-sonnet-4-6", dest="llm_endpoint")
    parser.add_argument("--vision-llm-endpoint", default="databricks-claude-sonnet-4-6", dest="vision_llm_endpoint")
    parser.add_argument("--fast-llm-endpoint", default="databricks-claude-haiku-4-5", dest="fast_llm_endpoint")
    parser.add_argument("--smtp-email", default="", dest="smtp_email")
    parser.add_argument("--warehouse-id", default=None, dest="warehouse_id")

    args = parser.parse_args()

    cfg = {
        "projects_base_path": f"/Volumes/{args.catalog}/{args.schema}/{args.volume}/projects",
        "volume_path": f"/Volumes/{args.catalog}/{args.schema}/{args.volume}",
        "llm_endpoint": args.llm_endpoint,
        "vision_llm_endpoint": args.vision_llm_endpoint,
        "fast_llm_endpoint": args.fast_llm_endpoint,
        "smtp_email": args.smtp_email,
    }

    print("=" * 60)
    print("  UC Function Registration")
    print("=" * 60)
    print(f"  Catalog:     {args.catalog}")
    print(f"  Schema:      {args.schema}")
    print(f"  LLM:         {args.llm_endpoint}")
    print(f"  Fast LLM:    {args.fast_llm_endpoint}")
    print(f"  Vision LLM:  {args.vision_llm_endpoint}")

    ok = register_all(args.catalog, args.schema, cfg, args.warehouse_id)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
