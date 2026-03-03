"""
Generate synthetic data for 5 different FRMC control testing projects.

Each project lives in sample_data/projects/<control_id>/ and contains:
  - engagement.json          (GRC metadata, instructions, evidence list, rules)
  - engagement_workbook.xlsx (tabs vary per control)
  - evidence/                (PDFs, screenshots, emails -- varies per control)

Controls:
  1. P2P-028  - Payment Proposal Approval
  2. ITG-015  - User Access Review
  3. FIN-042  - Manual Journal Entry Review
  4. HR-003   - Segregation of Duties
  5. REV-019  - Revenue Recognition Cutoff
"""

import json
import os
import random
import textwrap
from datetime import datetime, timedelta
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError("pip install openpyxl")

try:
    from fpdf import FPDF
except ImportError:
    raise ImportError("pip install fpdf2")

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    raise ImportError("pip install Pillow")

random.seed(42)

BASE_DIR = Path(__file__).parent / "projects"

DARK_BLUE = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HDR_FONT = Font(name="Georgia", bold=True, color="FFFFFF", size=9)
NORM_FONT = Font(name="Georgia", size=9)
BOLD_FONT = Font(name="Georgia", bold=True, size=9)
TITLE_FONT = Font(name="Georgia", bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*(Side(style="thin") for _ in range(4)))


def _hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DARK_BLUE
        cell.font = HDR_FONT
        cell.alignment = WRAP
        cell.border = THIN


def _data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORM_FONT
        cell.alignment = WRAP
        cell.border = THIN


class StyledPDF(FPDF):
    def __init__(self, title="GSK FRMC"):
        super().__init__()
        self._doc_title = title

    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(0, 51, 102)
        self.cell(0, 8, f"GSK - {self._doc_title}", new_x="LMARGIN", new_y="NEXT", align="R")
        self.set_draw_color(0, 51, 102)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"CONFIDENTIAL - Page {self.page_no()}/{{nb}}", align="C")


def _rand_date(year=2024, month=None):
    m = month or random.randint(1, 12)
    return datetime(year, m, random.randint(1, 28))


def _make_screenshot(path: Path, lines: list[str], width=900, height=600):
    """Generate a simple screenshot-like PNG image with text."""
    img = Image.new("RGB", (width, height), color=(245, 245, 250))
    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, width, 40], fill=(0, 51, 102))
    draw.text((15, 10), "GSK Identity & Access Management Portal", fill="white")
    draw.rectangle([0, 41, width, 42], fill=(243, 111, 33))
    y = 60
    for line in lines:
        color = (180, 0, 0) if "EXCEPTION" in line or "OVERDUE" in line or "ACTIVE" in line.upper() and "terminated" in line.lower() else (30, 30, 30)
        draw.text((20, y), line, fill=color)
        y += 22
        if y > height - 30:
            break
    img.save(str(path))


def _make_delivery_photo(path: Path, text_lines: list[str]):
    """Generate a simple delivery confirmation photo."""
    img = Image.new("RGB", (600, 400), color=(255, 253, 245))
    draw = ImageDraw.Draw(img)
    draw.rectangle([20, 20, 580, 380], outline=(100, 100, 100), width=2)
    draw.text((30, 30), "DELIVERY CONFIRMATION", fill=(0, 51, 102))
    draw.line((30, 55, 300, 55), fill=(0, 51, 102), width=2)
    y = 70
    for line in text_lines:
        color = (180, 0, 0) if "NOT SIGNED" in line or "MISSING" in line else (30, 30, 30)
        draw.text((30, y), line, fill=color)
        y += 25
    img.save(str(path))


# ============================================================================
# PROJECT 1: P2P-028 - Payment Proposal Approval
# ============================================================================
def gen_p2p_028():
    proj = BASE_DIR / "p2p_028"
    evi = proj / "evidence"
    evi.mkdir(parents=True, exist_ok=True)

    # --- engagement.json ---
    engagement = {
        "number": "ENG-2024-P2P-001",
        "name": "P2P Payment Proposal Review - Above Country 2024",
        "type": "Fieldwork",
        "state": "Fieldwork",
        "tracking_id": "56063177",
        "record_type": "Finance",
        "engagement_lead": "Rohit Rego",
        "planned_start": "06-Sep-2024",
        "notification_emails": {
            "report_to": "sean.zhang@databricks.com",
            "exceptions_to": "sean.zhang@databricks.com",
        },
        "control_objective": {
            "control_id": "P2P-028",
            "control_name": "Payment Proposal Approval and Threshold Review",
            "domain": "Accounts Payable",
            "policy_reference": "GSK-FIN-POL-028 v3.2",
            "rules": {
                "threshold_gbp": 50000,
                "approval_matrix": {
                    "below_50k": "Automated approval",
                    "50k_to_250k": "Finance Director",
                    "250k_to_1m": "VP Finance",
                    "above_1m": "CFO",
                },
            },
        },
        "testing_attributes": [
            {"ref": "A", "name": "Payment proposal list obtained", "applies_to": "all"},
            {"ref": "B", "name": "Threshold check performed", "applies_to": "all"},
            {"ref": "C", "name": "Above-threshold approval verified", "applies_to": "above_threshold"},
            {"ref": "D", "name": "Supporting documents reviewed", "applies_to": "above_threshold"},
            {"ref": "E", "name": "Blocked payments reviewed", "applies_to": "control_level"},
            {"ref": "F", "name": "Prepayments/down payments reviewed", "applies_to": "control_level"},
        ],
        "instructions": (
            "Execute fieldwork for Control P2P-028. Test all 6 attributes (A-F) against "
            "the selected sample. Review supporting evidence documents. Document findings "
            "in the Issue template."
        ),
        "evidence_files": [
            {"path": "evidence/payment_proposal_list.pdf", "type": "pdf", "focus": "proposal_list"},
            {"path": "evidence/approval_evidence.pdf", "type": "pdf", "focus": "approvals"},
            {"path": "evidence/supporting_docs.pdf", "type": "pdf", "focus": "supporting_docs"},
            {"path": "evidence/policy_reference.pdf", "type": "pdf", "focus": "policy"},
        ],
    }
    (proj / "engagement.json").write_text(json.dumps(engagement, indent=2))

    # --- Workbook ---
    companies = ["PS01", "TR03", "GX04", "UK07", "BE02"]
    currencies = ["GBP", "EUR", "USD"]
    fx = {"GBP": 1.0, "EUR": 0.86, "USD": 0.79}
    pop = []
    for i in range(200):
        ccy = random.choice(currencies)
        amt = round(random.uniform(-5000, 500000) if random.random() > 0.3 else random.uniform(-50000, 50000), 2)
        gbp = round(amt * fx[ccy], 2)
        pop.append({
            "Row_No": i + 1,
            "Company_Code": random.choice(companies),
            "Document_No": str(random.randint(70000000, 73999999)),
            "Amount": amt,
            "Amount_in_GBP": gbp,
            "Currency": ccy,
            "Threshold_Status": "Check" if abs(gbp) >= 50000 else "Under threshold",
        })

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Population determination"
    h1 = list(pop[0].keys())
    for c, h in enumerate(h1, 1):
        ws1.cell(row=1, column=c, value=h)
    _hdr(ws1, 1, len(h1))
    for r, row in enumerate(pop, 2):
        for c, h in enumerate(h1, 1):
            ws1.cell(row=r, column=c, value=row[h])
        _data(ws1, r, len(h1))

    above = [p for p in pop if p["Threshold_Status"] == "Check"]
    below = [p for p in pop if p["Threshold_Status"] == "Under threshold"]
    sample = random.sample(above, min(10, len(above))) + random.sample(below, min(15, len(below)))
    random.shuffle(sample)

    ws2 = wb.create_sheet("Sampling")
    ws2.cell(row=1, column=1, value="Sampling Methodology").font = TITLE_FONT
    for r, (k, v) in enumerate([("Population Size", str(len(pop))), ("Sample Size", str(len(sample))), ("Method", "Stratified random")], 3):
        ws2.cell(row=r, column=1, value=k).font = BOLD_FONT
        ws2.cell(row=r, column=2, value=v).font = NORM_FONT
    ws2.cell(row=7, column=1, value="Selected Sample").font = TITLE_FONT
    sh = ["Sample_No", "Document_No", "Amount_in_GBP", "Threshold_Status"]
    for c, h in enumerate(sh, 1):
        ws2.cell(row=8, column=c, value=h)
    _hdr(ws2, 8, len(sh))
    for i, s in enumerate(sample):
        r = 9 + i
        ws2.cell(row=r, column=1, value=i + 1)
        ws2.cell(row=r, column=2, value=s["Document_No"])
        ws2.cell(row=r, column=3, value=s["Amount_in_GBP"])
        ws2.cell(row=r, column=4, value=s["Threshold_Status"])

    ws3 = wb.create_sheet("Testing Table")
    ws3.cell(row=1, column=1, value="Control Testing Attributes").font = TITLE_FONT
    for c, h in enumerate(["Ref", "Attribute", "Procedure", "Answer"], 1):
        ws3.cell(row=2, column=c, value=h)
    _hdr(ws3, 2, 4)
    attrs = engagement["testing_attributes"]
    for i, a in enumerate(attrs):
        ws3.cell(row=3 + i, column=1, value=a["ref"])
        ws3.cell(row=3 + i, column=2, value=a["name"])
        ws3.cell(row=3 + i, column=3, value="See engagement instructions")
        ws3.cell(row=3 + i, column=4, value="")

    ws4 = wb.create_sheet("Issue template")
    ws4.cell(row=1, column=1, value="Issue / Exception Report").font = TITLE_FONT
    for c, h in enumerate(["Issue_ID", "Testing_Attribute", "Severity", "Description", "Affected_Samples", "Root_Cause", "Remediation", "Owner", "Due_Date", "Status"], 1):
        ws4.cell(row=3, column=c, value=h)
    _hdr(ws4, 3, 10)

    wb.save(proj / "engagement_workbook.xlsx")

    # --- Evidence PDFs ---
    pdf = StyledPDF("Payment Proposal List")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Payment Proposal List", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Population: {len(pop)} transactions | Sample: {len(sample)} items", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    for i, s in enumerate(sample):
        pdf.cell(0, 5, f"  {i+1}. Doc {s['Document_No']} | GBP {s['Amount_in_GBP']:,.2f} | {s['Threshold_Status']}", new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(evi / "payment_proposal_list.pdf"))

    above_sample = [s for s in sample if s["Threshold_Status"] == "Check"]
    pdf2 = StyledPDF("Approval Evidence")
    pdf2.alias_nb_pages()
    pdf2.add_page()
    pdf2.set_font("Helvetica", "B", 14)
    pdf2.cell(0, 10, "Approval Evidence Pack", new_x="LMARGIN", new_y="NEXT")
    pdf2.set_font("Helvetica", "", 9)
    for i, s in enumerate(above_sample):
        amt = abs(s["Amount_in_GBP"])
        req = "Finance Director" if amt < 250000 else "VP Finance" if amt < 1000000 else "CFO"
        if i == 2:
            pdf2.set_text_color(180, 0, 0)
            pdf2.cell(0, 5, f"  Doc {s['Document_No']} | GBP {amt:,.2f} | Required: {req} | Actual: Finance Director | ** EXCEPTION: Wrong level **", new_x="LMARGIN", new_y="NEXT")
            pdf2.set_text_color(0, 0, 0)
        elif i == 5 and len(above_sample) > 5:
            pdf2.set_text_color(180, 0, 0)
            pdf2.cell(0, 5, f"  Doc {s['Document_No']} | GBP {amt:,.2f} | Required: {req} | Actual: [NOT FOUND] | ** EXCEPTION: Missing approval **", new_x="LMARGIN", new_y="NEXT")
            pdf2.set_text_color(0, 0, 0)
        else:
            actual = {"Finance Director": "D. Chen", "VP Finance": "S. Mitchell", "CFO": "J. Wright"}.get(req, "Auto")
            pdf2.cell(0, 5, f"  Doc {s['Document_No']} | GBP {amt:,.2f} | Required: {req} | Actual: {actual} | Approved", new_x="LMARGIN", new_y="NEXT")
    pdf2.output(str(evi / "approval_evidence.pdf"))

    pdf3 = StyledPDF("Supporting Documentation")
    pdf3.alias_nb_pages()
    pdf3.add_page()
    pdf3.set_font("Helvetica", "B", 14)
    pdf3.cell(0, 10, "Supporting Documentation Pack", new_x="LMARGIN", new_y="NEXT")
    pdf3.set_font("Helvetica", "", 9)
    for i, s in enumerate(above_sample):
        if i == 7 and len(above_sample) > 7:
            pdf3.set_text_color(180, 0, 0)
            pdf3.cell(0, 5, f"  Doc {s['Document_No']} | PO: [NOT FOUND] | GR: [NOT FOUND] | ** EXCEPTION **", new_x="LMARGIN", new_y="NEXT")
            pdf3.set_text_color(0, 0, 0)
        else:
            pdf3.cell(0, 5, f"  Doc {s['Document_No']} | PO: PO-{random.randint(400000,499999)} | GR: GR-{random.randint(500000,599999)} | Verified", new_x="LMARGIN", new_y="NEXT")
    pdf3.output(str(evi / "supporting_docs.pdf"))

    pdf4 = StyledPDF("Policy Reference")
    pdf4.alias_nb_pages()
    pdf4.add_page()
    pdf4.set_font("Helvetica", "B", 14)
    pdf4.cell(0, 10, "GSK-FIN-POL-028 v3.2 - Payment Controls Policy", new_x="LMARGIN", new_y="NEXT")
    pdf4.set_font("Helvetica", "", 9)
    pdf4.multi_cell(0, 5, "Threshold: GBP 50,000\nBelow 50K: Auto | 50-250K: Finance Director | 250K-1M: VP Finance | >1M: CFO\nAll above-threshold require supporting docs (PO + GR).")
    pdf4.output(str(evi / "policy_reference.pdf"))

    print(f"  P2P-028: {len(pop)} pop, {len(sample)} sample, 4 PDFs")


# ============================================================================
# PROJECT 2: ITG-015 - User Access Review
# ============================================================================
def gen_itg_015():
    proj = BASE_DIR / "itg_015"
    evi = proj / "evidence"
    evi.mkdir(parents=True, exist_ok=True)

    engagement = {
        "number": "ENG-2024-ITG-001",
        "name": "IT General Controls - User Access Review Q1-Q4 2024",
        "type": "Fieldwork",
        "state": "Fieldwork",
        "tracking_id": "56063200",
        "record_type": "IT Controls",
        "engagement_lead": "Priya Nair",
        "planned_start": "15-Oct-2024",
        "notification_emails": {
            "report_to": "sean.zhang@databricks.com",
            "exceptions_to": "sean.zhang@databricks.com",
        },
        "control_objective": {
            "control_id": "ITG-015",
            "control_name": "User Access Review and Recertification",
            "domain": "IT General Controls",
            "policy_reference": "GSK-IT-POL-015 v2.1",
            "rules": {
                "review_cadence": "Quarterly",
                "removal_sla_days": 30,
                "privileged_access_justification_required": True,
            },
        },
        "testing_attributes": [
            {"ref": "A", "name": "Complete user access list obtained from system", "applies_to": "all"},
            {"ref": "B", "name": "Quarterly access reviews completed on time", "applies_to": "all"},
            {"ref": "C", "name": "Terminated users removed within 30-day SLA", "applies_to": "terminated_users"},
            {"ref": "D", "name": "Privileged access has documented justification", "applies_to": "privileged_users"},
        ],
        "instructions": (
            "Execute fieldwork for Control ITG-015. Verify that user access reviews were "
            "performed quarterly, terminated users were removed within SLA, and privileged "
            "access is justified. Review screenshots from the IAM portal and access reports."
        ),
        "evidence_files": [
            {"path": "evidence/access_review_q1.png", "type": "screenshot", "focus": "quarterly_review_completion"},
            {"path": "evidence/access_review_q2.png", "type": "screenshot", "focus": "quarterly_review_completion"},
            {"path": "evidence/access_report.pdf", "type": "pdf", "focus": "user_access_listing"},
            {"path": "evidence/access_policy.pdf", "type": "pdf", "focus": "policy"},
        ],
    }
    (proj / "engagement.json").write_text(json.dumps(engagement, indent=2))

    roles = ["Read Only", "Standard User", "Power User", "Admin", "Service Account"]
    systems = ["SAP ERP", "Concur", "Veeva CRM", "LIMS", "Active Directory"]
    statuses = ["Active", "Active", "Active", "Active", "Terminated"]
    users = []
    for i in range(150):
        status = random.choices(statuses, weights=[4, 4, 4, 4, 1], k=1)[0]
        term_date = _rand_date().strftime("%d/%m/%Y") if status == "Terminated" else ""
        users.append({
            "User_ID": f"USR{10000 + i}",
            "Name": f"User {i+1}",
            "Department": random.choice(["Finance", "IT", "HR", "Sales", "Manufacturing"]),
            "System": random.choice(systems),
            "Role": random.choice(roles),
            "Status": status,
            "Last_Review_Date": _rand_date().strftime("%d/%m/%Y"),
            "Termination_Date": term_date,
        })

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "User Access List"
    h1 = list(users[0].keys())
    for c, h in enumerate(h1, 1):
        ws1.cell(row=1, column=c, value=h)
    _hdr(ws1, 1, len(h1))
    for r, row in enumerate(users, 2):
        for c, h in enumerate(h1, 1):
            ws1.cell(row=r, column=c, value=row[h])

    term_users = [u for u in users if u["Status"] == "Terminated"]
    priv_users = [u for u in users if u["Role"] in ("Admin", "Service Account")]
    sample = random.sample(term_users, min(10, len(term_users))) + random.sample(priv_users, min(10, len(priv_users)))

    ws2 = wb.create_sheet("Sampling")
    ws2.cell(row=1, column=1, value="Sampling").font = TITLE_FONT
    ws2.cell(row=3, column=1, value="Population").font = BOLD_FONT
    ws2.cell(row=3, column=2, value=str(len(users))).font = NORM_FONT
    ws2.cell(row=4, column=1, value="Sample Size").font = BOLD_FONT
    ws2.cell(row=4, column=2, value=str(len(sample))).font = NORM_FONT
    ws2.cell(row=6, column=1, value="Selected Sample").font = TITLE_FONT
    sh = ["Sample_No", "User_ID", "Role", "Status", "System"]
    for c, h in enumerate(sh, 1):
        ws2.cell(row=7, column=c, value=h)
    _hdr(ws2, 7, len(sh))
    for i, s in enumerate(sample):
        ws2.cell(row=8 + i, column=1, value=i + 1)
        ws2.cell(row=8 + i, column=2, value=s["User_ID"])
        ws2.cell(row=8 + i, column=3, value=s["Role"])
        ws2.cell(row=8 + i, column=4, value=s["Status"])
        ws2.cell(row=8 + i, column=5, value=s["System"])

    ws3 = wb.create_sheet("Testing Table")
    ws3.cell(row=1, column=1, value="Testing Attributes").font = TITLE_FONT
    for c, h in enumerate(["Ref", "Attribute", "Procedure", "Answer"], 1):
        ws3.cell(row=2, column=c, value=h)
    _hdr(ws3, 2, 4)
    for i, a in enumerate(engagement["testing_attributes"]):
        ws3.cell(row=3 + i, column=1, value=a["ref"])
        ws3.cell(row=3 + i, column=2, value=a["name"])
        ws3.cell(row=3 + i, column=3, value="See engagement instructions")

    ws4 = wb.create_sheet("Issue template")
    ws4.cell(row=1, column=1, value="Issue / Exception Report").font = TITLE_FONT
    for c, h in enumerate(["Issue_ID", "Testing_Attribute", "Severity", "Description", "Affected_Samples", "Root_Cause", "Remediation", "Owner", "Due_Date", "Status"], 1):
        ws4.cell(row=3, column=c, value=h)
    _hdr(ws4, 3, 10)

    wb.save(proj / "engagement_workbook.xlsx")

    # --- Screenshots ---
    _make_screenshot(evi / "access_review_q1.png", [
        "Access Review Dashboard - Q1 2024",
        "",
        "Review Period: Jan 1 - Mar 31, 2024",
        "Status: COMPLETED | Completed Date: March 28, 2024",
        "Total Users Reviewed: 148 | Approved: 145 | Flagged: 3",
        "",
        "Reviewer: Priya Nair (IT Security Manager)",
        "Sign-off: Digital signature verified",
    ])
    _make_screenshot(evi / "access_review_q2.png", [
        "Access Review Dashboard - Q2 2024",
        "",
        "Review Period: Apr 1 - Jun 30, 2024",
        "Status: OVERDUE | Due Date: June 30, 2024 | Completed: July 22, 2024",
        "** EXCEPTION: Review completed 22 days late **",
        "",
        "Total Users Reviewed: 152 | Approved: 149 | Flagged: 3",
        "",
        "Flagged Users:",
        "  USR10045 - Admin access - terminated on 15/04/2024 - STILL ACTIVE in SAP ERP",
        "  ** EXCEPTION: Terminated user still has active access (>30 days) **",
        "  USR10098 - Admin access - no justification documented",
        "  ** EXCEPTION: Unjustified privileged access **",
    ])

    # --- PDFs ---
    pdf = StyledPDF("User Access Report")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "User Access Report - 2024", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Total users: {len(users)} | Terminated: {len(term_users)} | Privileged: {len(priv_users)}", new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(evi / "access_report.pdf"))

    pdf2 = StyledPDF("Access Policy")
    pdf2.alias_nb_pages()
    pdf2.add_page()
    pdf2.set_font("Helvetica", "B", 14)
    pdf2.cell(0, 10, "GSK-IT-POL-015 v2.1 - Access Management Policy", new_x="LMARGIN", new_y="NEXT")
    pdf2.set_font("Helvetica", "", 9)
    pdf2.multi_cell(0, 5, "Quarterly access reviews required.\nTerminated users must be removed within 30 days.\nPrivileged access requires documented business justification.")
    pdf2.output(str(evi / "access_policy.pdf"))

    print(f"  ITG-015: {len(users)} users, {len(sample)} sample, 2 screenshots + 2 PDFs")


# ============================================================================
# PROJECT 3: FIN-042 - Manual Journal Entry Review
# ============================================================================
def gen_fin_042():
    proj = BASE_DIR / "fin_042"
    evi = proj / "evidence"
    evi.mkdir(parents=True, exist_ok=True)

    engagement = {
        "number": "ENG-2024-FIN-001",
        "name": "Financial Controls - Manual Journal Entry Review 2024",
        "type": "Fieldwork",
        "state": "Fieldwork",
        "tracking_id": "56063300",
        "record_type": "Finance",
        "engagement_lead": "Sarah Mitchell",
        "planned_start": "01-Nov-2024",
        "notification_emails": {
            "report_to": "sean.zhang@databricks.com",
            "exceptions_to": "sean.zhang@databricks.com",
        },
        "control_objective": {
            "control_id": "FIN-042",
            "control_name": "Manual Journal Entry Review and Approval",
            "domain": "Financial Reporting",
            "policy_reference": "GSK-FIN-POL-042 v1.8",
            "rules": {
                "threshold_gbp": 100000,
                "no_self_approval": True,
                "posting_window_days": 5,
            },
        },
        "testing_attributes": [
            {"ref": "A", "name": "Complete JE listing obtained", "applies_to": "all"},
            {"ref": "B", "name": "Above-threshold JEs reviewed by second level", "applies_to": "above_threshold"},
            {"ref": "C", "name": "Dual authorization verified (no self-approval)", "applies_to": "all"},
            {"ref": "D", "name": "Supporting documentation attached", "applies_to": "all"},
            {"ref": "E", "name": "Posted in correct accounting period", "applies_to": "all"},
        ],
        "instructions": (
            "Execute fieldwork for Control FIN-042. Verify manual journal entries have "
            "proper dual authorization (preparer != approver), above-threshold entries have "
            "second-level review, and all entries are posted in the correct period. "
            "Review approval emails for authorization evidence."
        ),
        "evidence_files": [
            {"path": "evidence/je_listing.pdf", "type": "pdf", "focus": "journal_entry_listing"},
            {"path": "evidence/approval_email_001.eml", "type": "email", "focus": "authorization"},
            {"path": "evidence/approval_email_002.eml", "type": "email", "focus": "authorization"},
            {"path": "evidence/approval_email_003.eml", "type": "email", "focus": "authorization"},
            {"path": "evidence/je_policy.pdf", "type": "pdf", "focus": "policy"},
        ],
    }
    (proj / "engagement.json").write_text(json.dumps(engagement, indent=2))

    accounts = ["4100-Sales Revenue", "5200-COGS", "6100-SGA Expense", "7200-R&D Expense", "8100-Other Income"]
    preparers = ["Alice Brown", "Bob Carter", "Carol Davis", "David Evans"]
    approvers = ["Sarah Mitchell", "James Wright", "Emily Foster"]
    entries = []
    for i in range(180):
        preparer = random.choice(preparers)
        amt = round(random.uniform(5000, 500000), 2)
        entries.append({
            "JE_Number": f"JE-2024-{i+1:04d}",
            "Posting_Date": _rand_date().strftime("%d/%m/%Y"),
            "Account": random.choice(accounts),
            "Description": f"Manual adjustment - {random.choice(['reclassification','accrual','provision','correction'])}",
            "Amount_GBP": amt,
            "Preparer": preparer,
            "Approver": random.choice(approvers) if i != 42 else preparer,  # Exception: self-approval
            "Period": f"P{random.randint(1,12):02d}",
        })

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Journal Entries"
    h1 = list(entries[0].keys())
    for c, h in enumerate(h1, 1):
        ws1.cell(row=1, column=c, value=h)
    _hdr(ws1, 1, len(h1))
    for r, row in enumerate(entries, 2):
        for c, h in enumerate(h1, 1):
            ws1.cell(row=r, column=c, value=row[h])

    above_thresh = [e for e in entries if e["Amount_GBP"] >= 100000]
    sample = random.sample(entries, min(25, len(entries)))

    ws2 = wb.create_sheet("Sampling")
    ws2.cell(row=1, column=1, value="Sampling").font = TITLE_FONT
    ws2.cell(row=3, column=1, value="Population").font = BOLD_FONT
    ws2.cell(row=3, column=2, value=str(len(entries))).font = NORM_FONT
    ws2.cell(row=4, column=1, value="Above threshold (100K)").font = BOLD_FONT
    ws2.cell(row=4, column=2, value=str(len(above_thresh))).font = NORM_FONT
    ws2.cell(row=6, column=1, value="Selected Sample").font = TITLE_FONT
    sh = ["Sample_No", "JE_Number", "Amount_GBP", "Preparer", "Approver"]
    for c, h in enumerate(sh, 1):
        ws2.cell(row=7, column=c, value=h)
    _hdr(ws2, 7, len(sh))
    for i, s in enumerate(sample):
        ws2.cell(row=8 + i, column=1, value=i + 1)
        ws2.cell(row=8 + i, column=2, value=s["JE_Number"])
        ws2.cell(row=8 + i, column=3, value=s["Amount_GBP"])
        ws2.cell(row=8 + i, column=4, value=s["Preparer"])
        ws2.cell(row=8 + i, column=5, value=s["Approver"])

    ws3 = wb.create_sheet("Testing Table")
    ws3.cell(row=1, column=1, value="Testing Attributes").font = TITLE_FONT
    for c, h in enumerate(["Ref", "Attribute", "Procedure", "Answer"], 1):
        ws3.cell(row=2, column=c, value=h)
    _hdr(ws3, 2, 4)
    for i, a in enumerate(engagement["testing_attributes"]):
        ws3.cell(row=3 + i, column=1, value=a["ref"])
        ws3.cell(row=3 + i, column=2, value=a["name"])

    ws4 = wb.create_sheet("Issue template")
    ws4.cell(row=1, column=1, value="Issue / Exception Report").font = TITLE_FONT
    for c, h in enumerate(["Issue_ID", "Testing_Attribute", "Severity", "Description", "Affected_Samples", "Root_Cause", "Remediation", "Owner", "Due_Date", "Status"], 1):
        ws4.cell(row=3, column=c, value=h)
    _hdr(ws4, 3, 10)
    wb.save(proj / "engagement_workbook.xlsx")

    # --- Emails ---
    def _write_eml(path, from_addr, to_addr, subject, body, date_str):
        msg = EmailMessage()
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg["Subject"] = subject
        msg["Date"] = date_str
        msg.set_content(body)
        path.write_text(msg.as_string())

    _write_eml(evi / "approval_email_001.eml",
        "alice.brown@gsk.com", "sarah.mitchell@gsk.com",
        "JE Approval Request: JE-2024-0001 - GBP 245,000",
        "Hi Sarah,\n\nPlease approve JE-2024-0001 for GBP 245,000.\nAccount: 4100-Sales Revenue\nDescription: Manual adjustment - accrual\n\nRegards,\nAlice",
        "Mon, 15 Jan 2024 09:30:00 +0000")

    _write_eml(evi / "approval_email_002.eml",
        "sarah.mitchell@gsk.com", "alice.brown@gsk.com",
        "Re: JE Approval Request: JE-2024-0001 - GBP 245,000",
        "Hi Alice,\n\nApproved. Please proceed with posting.\n\nRegards,\nSarah Mitchell\nFinance Director",
        "Mon, 15 Jan 2024 11:15:00 +0000")

    # Exception: self-approval email
    _write_eml(evi / "approval_email_003.eml",
        "bob.carter@gsk.com", "bob.carter@gsk.com",
        "JE Approval: JE-2024-0043 - GBP 187,500 - SELF-APPROVED",
        "Journal entry JE-2024-0043 approved.\nPreparer: Bob Carter\nApprover: Bob Carter\n\n** NOTE: This is a SELF-APPROVAL which violates GSK-FIN-POL-042 dual authorization requirement **",
        "Wed, 20 Mar 2024 14:00:00 +0000")

    # --- PDFs ---
    pdf = StyledPDF("Journal Entry Listing")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Manual Journal Entry Listing - 2024", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Total entries: {len(entries)} | Above GBP 100K: {len(above_thresh)}", new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(evi / "je_listing.pdf"))

    pdf2 = StyledPDF("JE Policy")
    pdf2.alias_nb_pages()
    pdf2.add_page()
    pdf2.set_font("Helvetica", "B", 14)
    pdf2.cell(0, 10, "GSK-FIN-POL-042 v1.8 - Journal Entry Controls", new_x="LMARGIN", new_y="NEXT")
    pdf2.set_font("Helvetica", "", 9)
    pdf2.multi_cell(0, 5, "Threshold: GBP 100,000 for second-level review.\nNo self-approval: Preparer and approver must be different individuals.\nPosting window: Within 5 business days of period end.")
    pdf2.output(str(evi / "je_policy.pdf"))

    print(f"  FIN-042: {len(entries)} JEs, {len(sample)} sample, 3 emails + 2 PDFs")


# ============================================================================
# PROJECT 4: HR-003 - Segregation of Duties
# ============================================================================
def gen_hr_003():
    proj = BASE_DIR / "hr_003"
    evi = proj / "evidence"
    evi.mkdir(parents=True, exist_ok=True)

    engagement = {
        "number": "ENG-2024-HR-001",
        "name": "HR Controls - Segregation of Duties Monitoring 2024",
        "type": "Fieldwork",
        "state": "Fieldwork",
        "tracking_id": "56063400",
        "record_type": "IT Controls",
        "engagement_lead": "Michael Chen",
        "planned_start": "01-Dec-2024",
        "notification_emails": {
            "report_to": "sean.zhang@databricks.com",
            "exceptions_to": "sean.zhang@databricks.com",
        },
        "control_objective": {
            "control_id": "HR-003",
            "control_name": "Segregation of Duties Monitoring",
            "domain": "HR / IT Controls",
            "policy_reference": "GSK-HR-POL-003 v2.0",
            "rules": {
                "conflicting_pairs": [
                    ["Create Purchase Order", "Approve Purchase Order"],
                    ["Create Vendor", "Process Payment"],
                    ["Post Journal Entry", "Approve Journal Entry"],
                    ["Maintain User Access", "Approve Access Changes"],
                ],
                "compensating_control_required": True,
            },
        },
        "testing_attributes": [
            {"ref": "A", "name": "SoD conflict matrix obtained and complete", "applies_to": "all"},
            {"ref": "B", "name": "All SoD conflicts identified and documented", "applies_to": "all"},
            {"ref": "C", "name": "Compensating controls documented for all conflicts", "applies_to": "conflicts"},
        ],
        "instructions": (
            "Execute fieldwork for Control HR-003. Obtain the SoD conflict matrix, "
            "verify all conflicts are identified, and confirm compensating controls "
            "are documented for each conflict. Review SAP role configuration screenshots."
        ),
        "evidence_files": [
            {"path": "evidence/sod_conflict_report.pdf", "type": "pdf", "focus": "conflict_identification"},
            {"path": "evidence/sap_role_screenshot.png", "type": "screenshot", "focus": "role_configuration"},
            {"path": "evidence/remediation_tracker.pdf", "type": "pdf", "focus": "compensating_controls"},
        ],
    }
    (proj / "engagement.json").write_text(json.dumps(engagement, indent=2))

    functions = ["Create PO", "Approve PO", "Create Vendor", "Process Payment", "Post JE", "Approve JE", "Maintain Access", "Approve Access"]
    assignments = []
    for i in range(120):
        user_funcs = random.sample(functions, random.randint(1, 3))
        assignments.append({
            "User_ID": f"USR{20000 + i}",
            "Name": f"User {i+1}",
            "Department": random.choice(["Finance", "Procurement", "IT", "HR"]),
            "Assigned_Functions": ", ".join(user_funcs),
            "Has_Conflict": "Yes" if any(
                set(user_funcs) >= {a, b}
                for a, b in [("Create PO", "Approve PO"), ("Create Vendor", "Process Payment"), ("Post JE", "Approve JE")]
            ) else "No",
            "Compensating_Control": "",
        })

    conflicts = [a for a in assignments if a["Has_Conflict"] == "Yes"]
    for i, c in enumerate(conflicts):
        if i == 2:
            c["Compensating_Control"] = ""  # Exception: missing compensating control
        else:
            c["Compensating_Control"] = f"CC-{random.randint(100,999)}: Manager review of transactions"

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Role Assignments"
    h1 = list(assignments[0].keys())
    for c, h in enumerate(h1, 1):
        ws1.cell(row=1, column=c, value=h)
    _hdr(ws1, 1, len(h1))
    for r, row in enumerate(assignments, 2):
        for c, h in enumerate(h1, 1):
            ws1.cell(row=r, column=c, value=row[h])

    ws2 = wb.create_sheet("Sampling")
    ws2.cell(row=1, column=1, value="All conflicts sampled (100%)").font = TITLE_FONT
    ws2.cell(row=3, column=1, value="Total conflicts").font = BOLD_FONT
    ws2.cell(row=3, column=2, value=str(len(conflicts))).font = NORM_FONT

    ws3 = wb.create_sheet("Testing Table")
    ws3.cell(row=1, column=1, value="Testing Attributes").font = TITLE_FONT
    for c, h in enumerate(["Ref", "Attribute", "Procedure", "Answer"], 1):
        ws3.cell(row=2, column=c, value=h)
    _hdr(ws3, 2, 4)
    for i, a in enumerate(engagement["testing_attributes"]):
        ws3.cell(row=3 + i, column=1, value=a["ref"])
        ws3.cell(row=3 + i, column=2, value=a["name"])

    ws4 = wb.create_sheet("Issue template")
    ws4.cell(row=1, column=1, value="Issue / Exception Report").font = TITLE_FONT
    for c, h in enumerate(["Issue_ID", "Testing_Attribute", "Severity", "Description", "Affected_Samples", "Root_Cause", "Remediation", "Owner", "Due_Date", "Status"], 1):
        ws4.cell(row=3, column=c, value=h)
    _hdr(ws4, 3, 10)
    wb.save(proj / "engagement_workbook.xlsx")

    # --- Screenshot ---
    _make_screenshot(evi / "sap_role_screenshot.png", [
        "SAP Role Configuration - Conflict Analysis",
        "",
        f"Total Users Analyzed: {len(assignments)}",
        f"Users with SoD Conflicts: {len(conflicts)}",
        "",
        "Conflict Details:",
    ] + [
        f"  {c['User_ID']} ({c['Department']}): {c['Assigned_Functions']}"
        + (" | CC: " + c["Compensating_Control"] if c["Compensating_Control"] else " | ** EXCEPTION: No compensating control **")
        for c in conflicts[:8]
    ])

    # --- PDFs ---
    pdf = StyledPDF("SoD Conflict Report")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Segregation of Duties Conflict Report", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Total conflicts: {len(conflicts)}", new_x="LMARGIN", new_y="NEXT")
    for c in conflicts:
        color = (180, 0, 0) if not c["Compensating_Control"] else (0, 0, 0)
        pdf.set_text_color(*color)
        cc = c["Compensating_Control"] or "** MISSING **"
        pdf.cell(0, 5, f"  {c['User_ID']}: {c['Assigned_Functions']} | CC: {cc}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.output(str(evi / "sod_conflict_report.pdf"))

    pdf2 = StyledPDF("Remediation Tracker")
    pdf2.alias_nb_pages()
    pdf2.add_page()
    pdf2.set_font("Helvetica", "B", 14)
    pdf2.cell(0, 10, "SoD Remediation Tracker", new_x="LMARGIN", new_y="NEXT")
    pdf2.set_font("Helvetica", "", 9)
    for c in conflicts:
        status = "Remediated" if c["Compensating_Control"] else "OPEN - No action taken"
        pdf2.cell(0, 5, f"  {c['User_ID']}: {status}", new_x="LMARGIN", new_y="NEXT")
    pdf2.output(str(evi / "remediation_tracker.pdf"))

    print(f"  HR-003: {len(assignments)} users, {len(conflicts)} conflicts, 1 screenshot + 2 PDFs")


# ============================================================================
# PROJECT 5: REV-019 - Revenue Recognition Cutoff
# ============================================================================
def gen_rev_019():
    proj = BASE_DIR / "rev_019"
    evi = proj / "evidence"
    evi.mkdir(parents=True, exist_ok=True)

    engagement = {
        "number": "ENG-2024-REV-001",
        "name": "Revenue Controls - Cutoff Testing Q4 2024",
        "type": "Fieldwork",
        "state": "Fieldwork",
        "tracking_id": "56063500",
        "record_type": "Finance",
        "engagement_lead": "James Wright",
        "planned_start": "15-Jan-2025",
        "notification_emails": {
            "report_to": "sean.zhang@databricks.com",
            "exceptions_to": "sean.zhang@databricks.com",
        },
        "control_objective": {
            "control_id": "REV-019",
            "control_name": "Revenue Recognition Cutoff Testing",
            "domain": "Revenue / Financial Reporting",
            "policy_reference": "GSK-REV-POL-019 v1.5",
            "rules": {
                "cutoff_window_days": 5,
                "delivery_confirmation_required": True,
                "period_end": "31-Dec-2024",
            },
        },
        "testing_attributes": [
            {"ref": "A", "name": "Cutoff population obtained (orders +/- 5 days of period end)", "applies_to": "all"},
            {"ref": "B", "name": "Revenue recognized in correct accounting period", "applies_to": "all"},
            {"ref": "C", "name": "Delivery confirmed before revenue recognition", "applies_to": "all"},
            {"ref": "D", "name": "Credit notes near period end reviewed", "applies_to": "credit_notes"},
        ],
        "instructions": (
            "Execute fieldwork for Control REV-019. Test revenue cutoff by examining "
            "sales orders shipped within +/- 5 days of December 31, 2024. Verify delivery "
            "was confirmed before revenue was recognized. Review delivery photos and "
            "shipping documents. Check credit notes issued near period end."
        ),
        "evidence_files": [
            {"path": "evidence/shipping_manifest.pdf", "type": "pdf", "focus": "shipment_dates"},
            {"path": "evidence/delivery_photo_001.jpg", "type": "screenshot", "focus": "delivery_confirmation"},
            {"path": "evidence/delivery_photo_002.jpg", "type": "screenshot", "focus": "delivery_confirmation"},
            {"path": "evidence/credit_notes.pdf", "type": "pdf", "focus": "credit_notes"},
        ],
    }
    (proj / "engagement.json").write_text(json.dumps(engagement, indent=2))

    period_end = datetime(2024, 12, 31)
    orders = []
    for i in range(100):
        ship_delta = random.randint(-5, 5)
        ship_date = period_end + timedelta(days=ship_delta)
        delivery_delta = random.randint(1, 3)
        delivery_date = ship_date + timedelta(days=delivery_delta)
        rev_date = ship_date if random.random() > 0.1 else ship_date - timedelta(days=random.randint(1, 3))
        orders.append({
            "Order_No": f"SO-2024-{i+1:04d}",
            "Customer": f"Customer {random.randint(1, 50)}",
            "Amount_GBP": round(random.uniform(10000, 500000), 2),
            "Ship_Date": ship_date.strftime("%d/%m/%Y"),
            "Delivery_Date": delivery_date.strftime("%d/%m/%Y") if i != 15 else "",
            "Revenue_Date": rev_date.strftime("%d/%m/%Y"),
            "Period": "P12" if rev_date <= period_end else "P01-2025",
            "Delivery_Confirmed": "Yes" if i != 15 else "No",
        })

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Cutoff Population"
    h1 = list(orders[0].keys())
    for c, h in enumerate(h1, 1):
        ws1.cell(row=1, column=c, value=h)
    _hdr(ws1, 1, len(h1))
    for r, row in enumerate(orders, 2):
        for c, h in enumerate(h1, 1):
            ws1.cell(row=r, column=c, value=row[h])

    sample = random.sample(orders, min(25, len(orders)))

    ws2 = wb.create_sheet("Sampling")
    ws2.cell(row=1, column=1, value="Sampling").font = TITLE_FONT
    ws2.cell(row=3, column=1, value="Population").font = BOLD_FONT
    ws2.cell(row=3, column=2, value=str(len(orders))).font = NORM_FONT
    ws2.cell(row=5, column=1, value="Selected Sample").font = TITLE_FONT
    sh = ["Sample_No", "Order_No", "Amount_GBP", "Ship_Date", "Revenue_Date", "Delivery_Confirmed"]
    for c, h in enumerate(sh, 1):
        ws2.cell(row=6, column=c, value=h)
    _hdr(ws2, 6, len(sh))
    for i, s in enumerate(sample):
        ws2.cell(row=7 + i, column=1, value=i + 1)
        ws2.cell(row=7 + i, column=2, value=s["Order_No"])
        ws2.cell(row=7 + i, column=3, value=s["Amount_GBP"])
        ws2.cell(row=7 + i, column=4, value=s["Ship_Date"])
        ws2.cell(row=7 + i, column=5, value=s["Revenue_Date"])
        ws2.cell(row=7 + i, column=6, value=s["Delivery_Confirmed"])

    ws3 = wb.create_sheet("Testing Table")
    ws3.cell(row=1, column=1, value="Testing Attributes").font = TITLE_FONT
    for c, h in enumerate(["Ref", "Attribute", "Procedure", "Answer"], 1):
        ws3.cell(row=2, column=c, value=h)
    _hdr(ws3, 2, 4)
    for i, a in enumerate(engagement["testing_attributes"]):
        ws3.cell(row=3 + i, column=1, value=a["ref"])
        ws3.cell(row=3 + i, column=2, value=a["name"])

    ws4 = wb.create_sheet("Issue template")
    ws4.cell(row=1, column=1, value="Issue / Exception Report").font = TITLE_FONT
    for c, h in enumerate(["Issue_ID", "Testing_Attribute", "Severity", "Description", "Affected_Samples", "Root_Cause", "Remediation", "Owner", "Due_Date", "Status"], 1):
        ws4.cell(row=3, column=c, value=h)
    _hdr(ws4, 3, 10)
    wb.save(proj / "engagement_workbook.xlsx")

    # --- Delivery photos ---
    _make_delivery_photo(evi / "delivery_photo_001.jpg", [
        "Order: SO-2024-0001",
        "Customer: Customer 12",
        "Delivery Date: 02/01/2025",
        "Received by: John Smith",
        "Signature: [SIGNED]",
        "Condition: Good",
    ])
    _make_delivery_photo(evi / "delivery_photo_002.jpg", [
        "Order: SO-2024-0016",
        "Customer: Customer 33",
        "Delivery Date: MISSING",
        "Received by: N/A",
        "Signature: NOT SIGNED",
        "** EXCEPTION: No delivery confirmation **",
    ])

    # --- PDFs ---
    pdf = StyledPDF("Shipping Manifest")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Shipping Manifest - Dec 2024 / Jan 2025", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    for o in orders[:30]:
        status = "Delivered" if o["Delivery_Confirmed"] == "Yes" else "** UNCONFIRMED **"
        pdf.cell(0, 5, f"  {o['Order_No']} | Ship: {o['Ship_Date']} | GBP {o['Amount_GBP']:,.2f} | {status}", new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(evi / "shipping_manifest.pdf"))

    pdf2 = StyledPDF("Credit Notes")
    pdf2.alias_nb_pages()
    pdf2.add_page()
    pdf2.set_font("Helvetica", "B", 14)
    pdf2.cell(0, 10, "Credit Notes - Period End Review", new_x="LMARGIN", new_y="NEXT")
    pdf2.set_font("Helvetica", "", 9)
    for i in range(8):
        amt = round(random.uniform(5000, 50000), 2)
        pdf2.cell(0, 5, f"  CN-2024-{i+1:03d} | GBP {amt:,.2f} | Reason: {'Product return' if i % 2 == 0 else 'Pricing adjustment'}", new_x="LMARGIN", new_y="NEXT")
    pdf2.output(str(evi / "credit_notes.pdf"))

    print(f"  REV-019: {len(orders)} orders, {len(sample)} sample, 2 photos + 2 PDFs")


# ============================================================================
# PROJECT 6: ENV-007 - Environmental Compliance Inspection
# ============================================================================
def gen_env_007():
    proj = BASE_DIR / "env_007"
    evi = proj / "evidence"
    evi.mkdir(parents=True, exist_ok=True)

    engagement = {
        "number": "ENG-2024-ENV-001",
        "name": "EHS Controls - Environmental Compliance Inspection 2024",
        "type": "Fieldwork",
        "state": "Fieldwork",
        "tracking_id": "56063600",
        "record_type": "Environmental Health & Safety",
        "engagement_lead": "Dr. Karen Phillips",
        "planned_start": "01-Feb-2025",
        "notification_emails": {
            "report_to": "sean.zhang@databricks.com",
            "exceptions_to": "sean.zhang@databricks.com",
        },
        "control_objective": {
            "control_id": "ENV-007",
            "control_name": "Environmental Compliance Inspection and Monitoring",
            "domain": "Environmental Health & Safety",
            "policy_reference": "GSK-EHS-POL-007 v4.0",
            "rules": {
                "pm25_limit_ugm3": 25,
                "nox_limit_mgm3": 200,
                "inspection_cadence": "Monthly",
                "corrective_action_sla_days": 45,
            },
        },
        "testing_attributes": [
            {"ref": "A", "name": "Inspection schedule obtained and complete", "applies_to": "all"},
            {"ref": "B", "name": "Each inspection has photographic/screenshot evidence attached", "applies_to": "all"},
            {"ref": "C", "name": "Exceedances of emission thresholds flagged and escalated", "applies_to": "exceedances"},
            {"ref": "D", "name": "Corrective actions documented and tracked to closure", "applies_to": "corrective_actions"},
        ],
        "instructions": (
            "Execute fieldwork for Control ENV-007. Verify that monthly environmental "
            "inspections were performed, each has photographic evidence (screenshots of "
            "monitoring dashboards pasted into the workbook), emission exceedances were "
            "properly flagged, and corrective actions were tracked to closure. "
            "IMPORTANT: Evidence screenshots are EMBEDDED INSIDE the Excel workbook, "
            "not in separate files. Use extract_workbook_images to analyze them. "
            "After completing the review, email the report to the EHS Manager."
        ),
        "evidence_files": [
            {"path": "EMBEDDED_IN_WORKBOOK", "type": "embedded_image", "focus": "monitoring_dashboard_screenshots"},
            {"path": "EMBEDDED_IN_WORKBOOK", "type": "embedded_image", "focus": "equipment_condition_photo"},
            {"path": "evidence/ehs_policy.pdf", "type": "pdf", "focus": "policy"},
        ],
    }
    (proj / "engagement.json").write_text(json.dumps(engagement, indent=2))

    # --- Build workbook WITH embedded images ---
    from openpyxl.drawing.image import Image as XlImage

    sites = ["Stevenage UK", "Ware UK", "Barnard Castle UK", "Worthing UK", "Montrose UK"]
    inspectors = ["J. Adams", "M. Brown", "S. Clark", "P. Davis"]
    inspections = []
    for i in range(60):
        month = (i % 12) + 1
        pm25 = round(random.uniform(8, 35), 1)
        nox = round(random.uniform(80, 260), 1)
        inspections.append({
            "Inspection_ID": f"INS-2024-{i+1:03d}",
            "Date": datetime(2024, month, random.randint(1, 28)).strftime("%d/%m/%Y"),
            "Site": random.choice(sites),
            "Inspector": random.choice(inspectors),
            "PM2.5_ugm3": pm25,
            "NOx_mgm3": nox,
            "PM25_Status": "EXCEEDANCE" if pm25 > 25 else "Compliant",
            "NOx_Status": "EXCEEDANCE" if nox > 200 else "Compliant",
            "Evidence_Attached": "Yes" if i != 22 else "No",
        })

    exceedances = [ins for ins in inspections if ins["PM25_Status"] == "EXCEEDANCE" or ins["NOx_Status"] == "EXCEEDANCE"]
    corrective = []
    for j, exc in enumerate(exceedances):
        corrective.append({
            "CA_ID": f"CA-2024-{j+1:03d}",
            "Inspection_ID": exc["Inspection_ID"],
            "Site": exc["Site"],
            "Issue": f"{'PM2.5' if exc['PM25_Status'] == 'EXCEEDANCE' else 'NOx'} exceedance",
            "Action": "Filter replacement" if j % 2 == 0 else "Stack recalibration",
            "Due_Date": datetime(2024, min(12, int(exc["Date"].split("/")[1]) + 1), 15).strftime("%d/%m/%Y"),
            "Status": "Closed" if j != 3 else "Open - OVERDUE",
            "Escalated": "Yes" if j != 1 else "No",
        })

    wb = openpyxl.Workbook()

    # Tab 1: Inspection Log with embedded dashboard screenshots
    ws1 = wb.active
    ws1.title = "Inspection Log"
    h1 = list(inspections[0].keys())
    for c, h in enumerate(h1, 1):
        ws1.cell(row=1, column=c, value=h)
    _hdr(ws1, 1, len(h1))
    for r, row in enumerate(inspections, 2):
        for c, h in enumerate(h1, 1):
            ws1.cell(row=r, column=c, value=row[h])

    # Create dashboard screenshot images and embed them
    gap_row = len(inspections) + 4
    ws1.cell(row=gap_row, column=1, value="Dashboard Evidence (Embedded Screenshots)").font = TITLE_FONT

    # Dashboard 1: normal readings
    dash1 = _make_dashboard_image(
        "Air Quality Monitoring Dashboard - Stevenage UK - June 2024",
        [
            ("PM2.5", "18.3 ug/m3", "COMPLIANT", True),
            ("NOx", "145.2 mg/m3", "COMPLIANT", True),
            ("SO2", "12.1 mg/m3", "COMPLIANT", True),
            ("CO", "2.3 mg/m3", "COMPLIANT", True),
        ],
    )
    dash1_path = evi / "_temp_dash1.png"
    dash1.save(str(dash1_path))
    img1 = XlImage(str(dash1_path))
    img1.width = 500
    img1.height = 300
    ws1.add_image(img1, f"A{gap_row + 2}")

    # Dashboard 2: exceedance visible
    dash2 = _make_dashboard_image(
        "Air Quality Monitoring Dashboard - Ware UK - September 2024",
        [
            ("PM2.5", "31.7 ug/m3", "EXCEEDANCE", False),
            ("NOx", "210.5 mg/m3", "EXCEEDANCE", False),
            ("SO2", "11.8 mg/m3", "COMPLIANT", True),
            ("CO", "2.1 mg/m3", "COMPLIANT", True),
        ],
    )
    dash2_path = evi / "_temp_dash2.png"
    dash2.save(str(dash2_path))
    img2 = XlImage(str(dash2_path))
    img2.width = 500
    img2.height = 300
    ws1.add_image(img2, f"A{gap_row + 20}")

    # Tab 2: Sampling
    sample = random.sample(inspections, min(20, len(inspections)))
    ws2 = wb.create_sheet("Sampling")
    ws2.cell(row=1, column=1, value="Sampling").font = TITLE_FONT
    ws2.cell(row=3, column=1, value="Population").font = BOLD_FONT
    ws2.cell(row=3, column=2, value=str(len(inspections))).font = NORM_FONT
    ws2.cell(row=4, column=1, value="Exceedances").font = BOLD_FONT
    ws2.cell(row=4, column=2, value=str(len(exceedances))).font = NORM_FONT
    ws2.cell(row=6, column=1, value="Selected Sample").font = TITLE_FONT
    sh = ["Sample_No", "Inspection_ID", "Site", "PM2.5_ugm3", "NOx_mgm3", "Evidence_Attached"]
    for c, h in enumerate(sh, 1):
        ws2.cell(row=7, column=c, value=h)
    _hdr(ws2, 7, len(sh))
    for i, s in enumerate(sample):
        ws2.cell(row=8 + i, column=1, value=i + 1)
        ws2.cell(row=8 + i, column=2, value=s["Inspection_ID"])
        ws2.cell(row=8 + i, column=3, value=s["Site"])
        ws2.cell(row=8 + i, column=4, value=s["PM2.5_ugm3"])
        ws2.cell(row=8 + i, column=5, value=s["NOx_mgm3"])
        ws2.cell(row=8 + i, column=6, value=s["Evidence_Attached"])

    # Tab 3: Testing Table
    ws3 = wb.create_sheet("Testing Table")
    ws3.cell(row=1, column=1, value="Testing Attributes").font = TITLE_FONT
    for c, h in enumerate(["Ref", "Attribute", "Procedure", "Answer"], 1):
        ws3.cell(row=2, column=c, value=h)
    _hdr(ws3, 2, 4)
    for i, a in enumerate(engagement["testing_attributes"]):
        ws3.cell(row=3 + i, column=1, value=a["ref"])
        ws3.cell(row=3 + i, column=2, value=a["name"])

    # Tab 4: Corrective Actions with embedded equipment photo
    ws4 = wb.create_sheet("Corrective Actions")
    ws4.cell(row=1, column=1, value="Corrective Action Tracker").font = TITLE_FONT
    h4 = list(corrective[0].keys()) if corrective else []
    for c, h in enumerate(h4, 1):
        ws4.cell(row=3, column=c, value=h)
    _hdr(ws4, 3, len(h4))
    for r, row in enumerate(corrective, 4):
        for c, h in enumerate(h4, 1):
            ws4.cell(row=r, column=c, value=row[h])

    # Embed equipment condition photo
    equip_photo = _make_equipment_photo()
    equip_path = evi / "_temp_equip.png"
    equip_photo.save(str(equip_path))
    equip_img = XlImage(str(equip_path))
    equip_img.width = 400
    equip_img.height = 250
    ws4.add_image(equip_img, f"A{len(corrective) + 6}")

    # Issue template tab
    ws5 = wb.create_sheet("Issue template")
    ws5.cell(row=1, column=1, value="Issue / Exception Report").font = TITLE_FONT
    for c, h in enumerate(["Issue_ID", "Testing_Attribute", "Severity", "Description", "Affected_Samples", "Root_Cause", "Remediation", "Owner", "Due_Date", "Status"], 1):
        ws5.cell(row=3, column=c, value=h)
    _hdr(ws5, 3, 10)

    wb.save(proj / "engagement_workbook.xlsx")

    # Clean up temp image files
    for tmp in evi.glob("_temp_*.png"):
        tmp.unlink()

    # --- Policy PDF ---
    pdf = StyledPDF("EHS Policy")
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "GSK-EHS-POL-007 v4.0 - Environmental Compliance Policy", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5,
        "PM2.5 limit: 25 ug/m3\n"
        "NOx limit: 200 mg/m3\n"
        "Inspection cadence: Monthly at each manufacturing site\n"
        "All inspections must include photographic evidence of monitoring dashboards\n"
        "Exceedances must be escalated within 24 hours\n"
        "Corrective actions must be closed within 45 days"
    )
    pdf.output(str(evi / "ehs_policy.pdf"))

    print(f"  ENV-007: {len(inspections)} inspections, {len(exceedances)} exceedances, "
          f"3 embedded images + 1 PDF")


def _make_dashboard_image(title: str, readings: list[tuple[str, str, str, bool]]) -> Image.Image:
    """Generate a monitoring dashboard screenshot with readings."""
    img = Image.new("RGB", (700, 400), color=(20, 25, 35))
    draw = ImageDraw.Draw(img)

    # Header bar
    draw.rectangle([0, 0, 700, 50], fill=(0, 51, 102))
    draw.text((15, 15), title, fill="white")
    draw.rectangle([0, 50, 700, 52], fill=(243, 111, 33))

    y = 70
    for param, value, status, ok in readings:
        bg = (20, 80, 20) if ok else (120, 20, 20)
        draw.rectangle([20, y, 680, y + 65], fill=bg, outline=(60, 60, 60))
        draw.text((30, y + 8), param, fill=(200, 200, 200))
        draw.text((200, y + 5), value, fill="white")
        status_color = (100, 255, 100) if ok else (255, 80, 80)
        draw.text((450, y + 8), status, fill=status_color)
        y += 75

    draw.text((20, 370), f"Last updated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", fill=(100, 100, 100))
    return img


def _make_equipment_photo() -> Image.Image:
    """Generate a simulated equipment condition photo showing non-compliance."""
    img = Image.new("RGB", (500, 350), color=(200, 195, 185))
    draw = ImageDraw.Draw(img)

    # Equipment outline
    draw.rectangle([50, 50, 450, 280], outline=(80, 80, 80), width=3)
    draw.text((60, 60), "STACK EMISSION FILTER UNIT - Bay 3", fill=(0, 51, 102))
    draw.line((60, 80, 300, 80), fill=(0, 51, 102), width=2)

    # Damage indicators
    draw.rectangle([100, 100, 250, 200], outline=(150, 0, 0), width=3)
    draw.text((105, 105), "FILTER DEGRADATION", fill=(180, 0, 0))
    draw.text((105, 130), "Visible particulate", fill=(180, 0, 0))
    draw.text((105, 150), "buildup on housing", fill=(180, 0, 0))

    draw.rectangle([280, 120, 430, 240], outline=(180, 130, 0), width=2)
    draw.text((285, 125), "WARNING", fill=(180, 130, 0))
    draw.text((285, 150), "Seal integrity", fill=(180, 130, 0))
    draw.text((285, 170), "compromised", fill=(180, 130, 0))

    draw.text((60, 290), "Photo taken: 15/09/2024 | Inspector: M. Brown", fill=(80, 80, 80))
    draw.text((60, 310), "** NON-COMPLIANT: Requires immediate corrective action **", fill=(180, 0, 0))
    return img


# ============================================================================
# Main
# ============================================================================
if __name__ == "__main__":
    print("Generating 6 FRMC control testing projects...\n")

    print("[1/6] P2P-028 - Payment Proposal Approval")
    gen_p2p_028()

    print("\n[2/6] ITG-015 - User Access Review")
    gen_itg_015()

    print("\n[3/6] FIN-042 - Manual Journal Entry Review")
    gen_fin_042()

    print("\n[4/6] HR-003 - Segregation of Duties")
    gen_hr_003()

    print("\n[5/6] REV-019 - Revenue Recognition Cutoff")
    gen_rev_019()

    print("\n[6/6] ENV-007 - Environmental Compliance Inspection")
    gen_env_007()

    print("\n--- Summary ---")
    for d in sorted(BASE_DIR.iterdir()):
        if d.is_dir():
            files = list(d.rglob("*"))
            file_count = sum(1 for f in files if f.is_file())
            print(f"  {d.name}/: {file_count} files")
    print("\nDone!")
