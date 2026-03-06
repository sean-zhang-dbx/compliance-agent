"""
Tools for the GSK Controls Evidence Review Agent.

Fully general-purpose: no control-specific logic. The agent reads
engagement.json for each project and figures out what to do.

Tools:
  - list_projects            -> Discover available project directories
  - load_engagement          -> Read engagement metadata and instructions
  - parse_workbook           -> Read all tabs from the workbook dynamically
  - extract_workbook_images  -> Extract + analyze images embedded in Excel
  - review_document          -> Analyze PDFs with the LLM
  - review_screenshot        -> Analyze screenshots/photos with vision LLM
  - analyze_email            -> Parse and analyze .eml email files
  - execute_test             -> Run a specific test attribute against a sample item
  - compile_results          -> Produce the final assessment report
  - save_report              -> Persist the report to the project directory
  - send_email               -> Send email via Gmail SMTP (or save as .eml)
  - ask_user                 -> Ask the user for clarification
"""

from __future__ import annotations

import base64
import io
import json
import os
import zipfile
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Optional

import mlflow
from langchain_core.tools import tool

from agent.config import VOLUME_PATH, PROJECTS_BASE_PATH, PROJECTS_LOCAL_PATH

SAMPLE_DATA_DIR = Path(__file__).parent.parent / "sample_data"


@mlflow.trace(span_type="RETRIEVER", name="read_file")
def _read_file_bytes(file_path: str) -> bytes:
    """Read a file from local path, bundled sample_data, or UC volume.

    Resolution order (deterministic, path-aware):
      1. Exact local path
      2. Path-substituted local fallback (preserves directory structure)
      3. UC Volume SDK download
      4. Project-aware search in sample_data (matches parent dirs)
      5. Relative-path UC Volume lookup
    """
    p = Path(file_path)

    # 1. Exact local path
    if p.exists():
        return p.read_bytes()

    # 2. Path-substituted local fallback — preserves directory structure
    #    e.g. /Volumes/.../projects/fin_042/engagement.json
    #      -> .../sample_data/projects/fin_042/engagement.json
    local_fallback = file_path.replace(VOLUME_PATH, str(SAMPLE_DATA_DIR))
    fb = Path(local_fallback)
    if fb.exists():
        return fb.read_bytes()

    # 3. UC Volume SDK download (authoritative source)
    if file_path.startswith("/Volumes/"):
        try:
            from databricks.sdk import WorkspaceClient
            w = WorkspaceClient()
            resp = w.files.download(file_path)
            return resp.contents.read()
        except Exception:
            pass

    # 4. Project-aware search in sample_data and local cache
    #    Extract directory hints from the path to avoid cross-project matches
    path_parts = set(Path(file_path).parts)

    for search_root in (SAMPLE_DATA_DIR, Path(PROJECTS_LOCAL_PATH)):
        if not search_root.exists():
            continue
        best_match: Path | None = None
        best_score = -1
        for candidate in search_root.rglob(p.name):
            if not candidate.is_file():
                continue
            shared = len(set(candidate.parts) & path_parts)
            if shared > best_score:
                best_score = shared
                best_match = candidate
        if best_match is not None:
            return best_match.read_bytes()

    # 5. Relative-path UC Volume lookup
    if not file_path.startswith("/"):
        for prefix in (PROJECTS_BASE_PATH, VOLUME_PATH):
            vol_candidate = f"{prefix}/{file_path}"
            try:
                from databricks.sdk import WorkspaceClient
                w = WorkspaceClient()
                resp = w.files.download(vol_candidate)
                return resp.contents.read()
            except Exception:
                continue

    raise FileNotFoundError(
        f"Cannot find '{file_path}'. Checked: {p}, {fb}, UC Volume SDK, "
        f"sample_data (project-aware), PROJECTS_LOCAL_PATH"
    )


def _resolve_project_file(project_path: str, relative_path: str) -> str:
    """Resolve a file path relative to a project directory."""
    local_base = Path(PROJECTS_LOCAL_PATH) / project_path
    local_file = local_base / relative_path
    if local_file.exists():
        return str(local_file)
    return f"{PROJECTS_BASE_PATH}/{project_path}/{relative_path}"


@mlflow.trace(span_type="PARSER", name="parse_excel_tabs")
def _read_excel_tabs(file_path: str) -> dict:
    """Read all tabs from an Excel workbook into structured lists of rows."""
    import openpyxl
    content = _read_file_bytes(file_path)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        result[sheet_name] = rows
    return result


# =========================================================================
# Project discovery
# =========================================================================

@tool
def list_projects() -> str:
    """List all available control testing projects.
    Each project is a directory containing engagement.json, a workbook,
    and evidence files. Call this FIRST to see what projects are available.

    Returns:
        JSON with a list of projects, each with name, control_id,
        control_name, and domain from the engagement metadata.
    """
    projects = []

    local_base = Path(PROJECTS_LOCAL_PATH)
    if local_base.exists():
        for d in sorted(local_base.iterdir()):
            if d.is_dir():
                eng_file = d / "engagement.json"
                info = {"project_dir": d.name, "source": "local"}
                if eng_file.exists():
                    try:
                        eng = json.loads(eng_file.read_text())
                        co = eng.get("control_objective", {})
                        info.update({
                            "engagement_number": eng.get("number", ""),
                            "engagement_name": eng.get("name", ""),
                            "control_id": co.get("control_id", ""),
                            "control_name": co.get("control_name", ""),
                            "domain": co.get("domain", ""),
                        })
                    except Exception:
                        pass
                projects.append(info)

    try:
        from databricks.sdk import WorkspaceClient
        w = WorkspaceClient()
        items = w.files.list_directory_contents(PROJECTS_BASE_PATH)
        for item in items:
            if item.is_directory:
                name = item.path.rstrip("/").split("/")[-1]
                if not any(p["project_dir"] == name for p in projects):
                    projects.append({"project_dir": name, "source": "uc_volume"})
    except Exception:
        pass

    return json.dumps({"projects": projects, "count": len(projects)}, indent=2)


# =========================================================================
# Engagement + workbook loading
# =========================================================================

@tool
def load_engagement(project_path: str) -> str:
    """Load the engagement metadata and instructions from a project.

    Args:
        project_path: Project directory name (e.g. "p2p_028") or full path
                      to the engagement.json file.

    Returns:
        JSON string with engagement metadata including control_objective
        (with rules), testing_attributes, instructions, and evidence_files.
    """
    if project_path.endswith(".json"):
        file_path = project_path
    else:
        file_path = _resolve_project_file(project_path, "engagement.json")

    content = _read_file_bytes(file_path)
    engagement = json.loads(content)
    mlflow.update_current_trace(tags={
        "engagement_number": engagement.get("number", ""),
        "control_id": engagement.get("control_objective", {}).get("control_id", ""),
        "project": project_path,
    })
    return json.dumps(engagement, indent=2)


@tool
def parse_workbook(file_path: str) -> str:
    """Parse the engagement workbook (XLSX) and extract all tabs dynamically.
    Also detects if any worksheet contains embedded images (screenshots pasted
    into Excel). If so, the returned JSON includes has_embedded_images=true
    and per-tab image counts so you know to call extract_workbook_images.

    Args:
        file_path: Path to the workbook (.xlsx), or a project directory name.

    Returns:
        JSON with tab_names, tabs data, sampling_config, selected_sample,
        testing_attributes, and has_embedded_images with image_counts_by_tab.
    """
    if not file_path.endswith(".xlsx"):
        file_path = _resolve_project_file(file_path, "engagement_workbook.xlsx")

    tabs = _read_excel_tabs(file_path)

    # Detect embedded images
    import openpyxl
    wb_content = _read_file_bytes(file_path)
    wb = openpyxl.load_workbook(io.BytesIO(wb_content))
    image_counts = {}
    total_images = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = len(ws._images)
        if count > 0:
            image_counts[sheet_name] = count
            total_images += count

    # Also check the ZIP archive for images openpyxl might miss
    zip_image_count = 0
    try:
        with zipfile.ZipFile(io.BytesIO(wb_content), "r") as zf:
            zip_image_count = sum(1 for n in zf.namelist() if n.startswith("xl/media/"))
    except Exception:
        pass

    result = {
        "tab_names": list(tabs.keys()),
        "tabs": {},
        "sampling_config": {},
        "selected_sample": [],
        "testing_attributes": [],
        "has_embedded_images": total_images > 0 or zip_image_count > 0,
        "image_counts_by_tab": image_counts,
        "total_embedded_images": max(total_images, zip_image_count),
    }

    for tab_name, rows in tabs.items():
        if not rows:
            result["tabs"][tab_name] = {"headers": [], "row_count": 0, "data_preview": []}
            continue

        first_data_row = 0
        for i, row in enumerate(rows):
            non_empty = [c for c in row if c and c.strip()]
            if len(non_empty) >= 2:
                first_data_row = i
                break

        headers = rows[first_data_row] if first_data_row < len(rows) else []
        data_rows = rows[first_data_row + 1:] if first_data_row + 1 < len(rows) else []
        non_empty_rows = [r for r in data_rows if any(c.strip() for c in r)]

        result["tabs"][tab_name] = {
            "headers": headers,
            "row_count": len(non_empty_rows),
            "data_preview": non_empty_rows[:30],
        }

    for tab_name, rows in tabs.items():
        lower = tab_name.lower()
        if "sampling" in lower or "sample" in lower:
            in_config = True
            in_sample = False
            sample_headers = []

            for row in rows:
                if row[0] and "Selected Sample" in row[0]:
                    in_config = False
                    in_sample = True
                    continue

                if in_config and row[0] and row[0].strip() and row[1] and row[1].strip():
                    key = row[0].strip()
                    if key not in ("Sampling", "Sampling Methodology"):
                        result["sampling_config"][key] = row[1].strip()

                if in_sample and not sample_headers and any(c.strip() for c in row):
                    sample_headers = [h for h in row if h.strip()]
                    continue

                if in_sample and sample_headers and any(c.strip() for c in row):
                    item = {}
                    for i, h in enumerate(sample_headers):
                        if i < len(row):
                            item[h] = row[i]
                    if any(v.strip() for v in item.values()):
                        result["selected_sample"].append(item)

        if "testing" in lower and "table" in lower:
            for row in rows:
                if len(row) >= 2 and row[0].strip() and len(row[0].strip()) <= 2:
                    ref = row[0].strip()
                    if ref.isalpha() and ref.isupper():
                        result["testing_attributes"].append({
                            "ref": ref,
                            "attribute": row[1].strip() if len(row) > 1 else "",
                            "procedure": row[2].strip() if len(row) > 2 else "",
                            "answer": row[3].strip() if len(row) > 3 else "",
                        })

    return json.dumps(result, indent=2)


# =========================================================================
# Embedded image extraction from Excel
# =========================================================================

@tool
def extract_workbook_images(file_path: str, context: str = "") -> str:
    """Extract and analyze images embedded inside an Excel workbook.
    Some controls have screenshots or photos pasted directly into Excel
    tabs rather than stored as separate files. This tool extracts them
    and sends each to the vision LLM for analysis.

    Call this AFTER parse_workbook if has_embedded_images is true.

    Args:
        file_path: Path to the workbook (.xlsx) or project directory name.
        context: Control context (control_id, rules) for the LLM.

    Returns:
        JSON with a list of extracted images, each with sheet name,
        approximate cell anchor, image format, and the LLM's analysis.
    """
    if not file_path.endswith(".xlsx"):
        file_path = _resolve_project_file(file_path, "engagement_workbook.xlsx")

    from databricks_langchain import ChatDatabricks
    from langchain_core.messages import HumanMessage
    from agent.config import VISION_LLM_ENDPOINT

    wb_bytes = _read_file_bytes(file_path)
    images_found = []

    with mlflow.start_span(name="extract_workbook_images", span_type="PARSER") as span:
        span.set_inputs({"file_path": file_path})

        # Method 1: openpyxl _images (anchored images)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for idx, img in enumerate(ws._images):
                try:
                    anchor_cell = f"{img.anchor._from.col}:{img.anchor._from.row}" if hasattr(img.anchor, '_from') else "unknown"
                except Exception:
                    anchor_cell = "unknown"
                img_data = img._data()
                images_found.append({
                    "sheet": sheet_name,
                    "anchor": anchor_cell,
                    "source": "openpyxl",
                    "data": img_data,
                    "index": idx,
                })

        # Method 2: ZIP archive xl/media/ (catches images openpyxl misses)
        if not images_found:
            try:
                with zipfile.ZipFile(io.BytesIO(wb_bytes), "r") as zf:
                    for name in sorted(zf.namelist()):
                        if name.startswith("xl/media/"):
                            images_found.append({
                                "sheet": "archive",
                                "anchor": name,
                                "source": "zip_media",
                                "data": zf.read(name),
                                "index": len(images_found),
                            })
            except Exception:
                pass

        span.set_outputs({"images_found": len(images_found)})

    ctx = context or "GSK FRMC control testing"
    llm = ChatDatabricks(endpoint=VISION_LLM_ENDPOINT, temperature=0)
    analyses = []

    for img_info in images_found:
        with mlflow.start_span(name=f"analyze_embedded_image_{img_info['index']}", span_type="TOOL") as span:
            b64 = base64.b64encode(img_info["data"]).decode("utf-8")
            # Detect image type from header bytes
            header = img_info["data"][:8]
            if header[:4] == b'\x89PNG':
                mime = "image/png"
            elif header[:2] == b'\xff\xd8':
                mime = "image/jpeg"
            else:
                mime = "image/png"

            prompt = (
                f"You are analyzing an image extracted from an Excel workbook for {ctx}.\n"
                f"This image was found in sheet '{img_info['sheet']}' at position '{img_info['anchor']}'.\n\n"
                f"Analyze the image and report:\n"
                f"1. What does this image show? (dashboard, screenshot, photo, chart)\n"
                f"2. Key data visible (readings, statuses, dates, measurements)\n"
                f"3. Any EXCEPTIONS (values exceeding limits, red flags, non-compliance)\n"
                f"4. Relevance to the control being tested\n"
            )

            message = HumanMessage(content=[
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
            ])

            response = llm.invoke([message])
            span.set_outputs({"response_length": len(response.content)})

            entry = {
                "sheet": img_info["sheet"],
                "anchor": img_info["anchor"],
                "image_format": mime,
                "size_bytes": len(img_info["data"]),
                "analysis": response.content,
            }
            if len(img_info["data"]) < 500_000:
                entry["preview_data_uri"] = f"data:{mime};base64,{b64}"
            analyses.append(entry)

    return json.dumps({
        "file_path": file_path,
        "images_extracted": len(analyses),
        "analyses": analyses,
    }, indent=2)


# =========================================================================
# Evidence review tools
# =========================================================================

@tool
def review_document(file_path: str, context: str = "", focus_area: Optional[str] = None) -> str:
    """Review a supporting document (PDF) using the LLM.

    Args:
        file_path: Path to the PDF document.
        context: Control context from the engagement.
        focus_area: What to look for.

    Returns:
        JSON string with file_path, document_type, review_focus, and analysis.
    """
    from databricks_langchain import ChatDatabricks
    from agent.config import FAST_LLM_ENDPOINT

    content = _read_file_bytes(file_path)
    ext = Path(file_path).suffix.lower()
    text_content: str | None = None

    with mlflow.start_span(name="extract_document_text", span_type="PARSER") as span:
        span.set_inputs({"file_path": file_path, "ext": ext, "size_bytes": len(content)})
        if ext == ".pdf":
            try:
                import fitz
                doc = fitz.open(stream=content, filetype="pdf")
                text_content = ""
                for page in doc:
                    text_content += page.get_text() + "\n"
                doc.close()
                span.set_outputs({"chars_extracted": len(text_content), "method": "PyMuPDF"})
            except ImportError:
                text_content = f"[PDF at {file_path}, {len(content)} bytes]"
                span.set_outputs({"chars_extracted": 0, "method": "fallback"})
        else:
            text_content = content.decode("utf-8", errors="replace")
            span.set_outputs({"chars_extracted": len(text_content), "method": "raw_text"})

    focus = focus_area or "general compliance review"
    ctx = context or "GSK FRMC control testing"
    prompt = (
        f"You are reviewing a document for {ctx}.\n"
        f"Focus area: {focus}\n\n"
        f"Extract and report:\n"
        f"1. Document type and purpose\n"
        f"2. Key data points (amounts, dates, document numbers, user IDs)\n"
        f"3. Any approvals, sign-offs, or authorizations visible\n"
        f"4. Any EXCEPTIONS noted (policy violations, missing items, overdue actions)\n"
        f"5. Any thresholds or limits mentioned and whether they are breached\n"
        f"6. Summary of findings relevant to {focus}\n\n"
        f"Document text content:\n\n{text_content[:10000]}"
    )

    llm = ChatDatabricks(endpoint=FAST_LLM_ENDPOINT, temperature=0)
    response = llm.invoke(prompt)

    return json.dumps({
        "file_path": file_path,
        "document_type": ext,
        "file_size_bytes": len(content),
        "review_focus": focus,
        "analysis": response.content,
    }, indent=2)


@tool
def review_screenshot(file_path: str, context: str = "", focus_area: Optional[str] = None) -> str:
    """Review a screenshot or photo using the vision LLM.

    Args:
        file_path: Path to the image (PNG, JPG, JPEG).
        context: Control context from the engagement.
        focus_area: What to look for.

    Returns:
        JSON string with file_path, review_focus, and analysis.
    """
    from databricks_langchain import ChatDatabricks
    from langchain_core.messages import HumanMessage
    from agent.config import VISION_LLM_ENDPOINT

    content = _read_file_bytes(file_path)
    ext = Path(file_path).suffix.lower()

    with mlflow.start_span(name="review_screenshot", span_type="PARSER") as span:
        span.set_inputs({"file_path": file_path, "ext": ext, "size_bytes": len(content)})
        b64 = base64.b64encode(content).decode("utf-8")
        mime = "image/png" if ext == ".png" else "image/jpeg"

        focus = focus_area or "general visual inspection"
        ctx = context or "GSK FRMC control testing"
        prompt = (
            f"You are reviewing a screenshot/photo for {ctx}.\n"
            f"Focus area: {focus}\n\n"
            f"Analyze the image and report:\n"
            f"1. What the screenshot/photo shows\n"
            f"2. Key data visible (dates, statuses, user IDs, names, quantities)\n"
            f"3. Any status indicators (completed, overdue, pending, flagged)\n"
            f"4. Any EXCEPTIONS visible (red flags, overdue items, policy violations)\n"
            f"5. Summary of findings relevant to {focus}\n"
        )

        message = HumanMessage(content=[
            {"type": "text", "text": prompt},
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
        ])

        llm = ChatDatabricks(endpoint=VISION_LLM_ENDPOINT, temperature=0)
        response = llm.invoke([message])
        span.set_outputs({"response_length": len(response.content)})

    result = {
        "file_path": file_path,
        "image_type": ext,
        "file_size_bytes": len(content),
        "review_focus": focus,
        "analysis": response.content,
    }
    if len(content) < 500_000:
        result["preview_data_uri"] = f"data:{mime};base64,{b64}"
    return json.dumps(result, indent=2)


@tool
def analyze_email(file_path: str, context: str = "", focus_area: Optional[str] = None) -> str:
    """Parse and analyze an email file (.eml) for compliance evidence.

    Args:
        file_path: Path to the .eml email file.
        context: Control context from the engagement.
        focus_area: What to look for.

    Returns:
        JSON string with email metadata and analysis.
    """
    from databricks_langchain import ChatDatabricks
    from agent.config import FAST_LLM_ENDPOINT
    import email
    from email import policy as email_policy

    content = _read_file_bytes(file_path)

    with mlflow.start_span(name="parse_email", span_type="PARSER") as span:
        span.set_inputs({"file_path": file_path, "size_bytes": len(content)})

        msg = email.message_from_bytes(content, policy=email_policy.default)
        email_from = str(msg.get("From", ""))
        email_to = str(msg.get("To", ""))
        email_subject = str(msg.get("Subject", ""))
        email_date = str(msg.get("Date", ""))

        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    body = part.get_content()
                    break
        else:
            body = msg.get_content()

        span.set_outputs({"from": email_from, "to": email_to, "subject": email_subject})

    focus = focus_area or "authorization and approval"
    ctx = context or "GSK FRMC control testing"
    prompt = (
        f"You are analyzing an email for {ctx}.\n"
        f"Focus area: {focus}\n\n"
        f"Email metadata:\n"
        f"- From: {email_from}\n"
        f"- To: {email_to}\n"
        f"- Subject: {email_subject}\n"
        f"- Date: {email_date}\n\n"
        f"Email body:\n{body[:5000]}\n\n"
        f"Analyze and report:\n"
        f"1. Who is the sender and their role/authority level?\n"
        f"2. Who is the recipient?\n"
        f"3. What action is being requested or confirmed?\n"
        f"4. Is this a proper authorization/approval?\n"
        f"5. Any EXCEPTIONS: self-approval (From == To), unauthorized approver, "
        f"missing information, policy violations\n"
        f"6. Summary of compliance findings\n"
    )

    llm = ChatDatabricks(endpoint=FAST_LLM_ENDPOINT, temperature=0)
    response = llm.invoke(prompt)

    return json.dumps({
        "file_path": file_path,
        "email_from": email_from,
        "email_to": email_to,
        "email_subject": email_subject,
        "email_date": email_date,
        "review_focus": focus,
        "analysis": response.content,
    }, indent=2)


# =========================================================================
# Test plan generation (deterministic)
# =========================================================================

@tool
def generate_test_plan(
    engagement_json: str,
    workbook_json: str,
) -> str:
    """Generate a deterministic test plan from the engagement and workbook data.

    Call this AFTER load_engagement and parse_workbook. It computes the exact
    list of (attribute, sample_item) pairs that must be tested. Then execute
    each entry in the plan using execute_test.

    Args:
        engagement_json: The full engagement JSON returned by load_engagement.
        workbook_json: The full workbook JSON returned by parse_workbook.

    Returns:
        JSON with the ordered test_plan (list of tests to execute) and
        total_tests count. Each entry has test_ref, attribute, procedure,
        applies_to, and sample_item_json.
    """
    engagement = json.loads(engagement_json) if isinstance(engagement_json, str) else engagement_json
    workbook = json.loads(workbook_json) if isinstance(workbook_json, str) else workbook_json

    attributes = engagement.get("testing_attributes", [])
    rules = engagement.get("control_objective", {}).get("rules", {})
    threshold = rules.get("threshold_gbp", rules.get("threshold_usd", rules.get("threshold", 0)))

    selected_sample = workbook.get("selected_sample", [])
    wb_attributes = workbook.get("testing_attributes", [])

    attr_procedures = {}
    for wa in wb_attributes:
        attr_procedures[wa.get("ref", "")] = wa.get("procedure", wa.get("attribute", ""))

    test_plan = []

    for attr in attributes:
        ref = attr.get("ref", "?")
        name = attr.get("name", "")
        applies_to = attr.get("applies_to", "all")
        procedure = attr_procedures.get(ref, name)

        if applies_to == "control_level":
            test_plan.append({
                "test_ref": ref,
                "attribute": name,
                "procedure": procedure,
                "applies_to": applies_to,
                "sample_item_json": json.dumps({
                    "_type": "population_level",
                    "population_size": workbook.get("sampling_config", {}).get("Population Size",
                                       workbook.get("sampling_config", {}).get("Total Population", "unknown")),
                }),
            })
        elif applies_to in ("all",):
            for item in selected_sample:
                test_plan.append({
                    "test_ref": ref,
                    "attribute": name,
                    "procedure": procedure,
                    "applies_to": applies_to,
                    "sample_item_json": json.dumps(item),
                })
            if not selected_sample:
                test_plan.append({
                    "test_ref": ref,
                    "attribute": name,
                    "procedure": procedure,
                    "applies_to": applies_to,
                    "sample_item_json": json.dumps({"_type": "no_sample_available"}),
                })
        else:
            matched_items = []
            for item in selected_sample:
                amount_str = ""
                for k, v in item.items():
                    kl = k.lower()
                    if any(w in kl for w in ["amount", "value", "total", "gbp", "usd"]):
                        amount_str = str(v).replace(",", "").replace("£", "").replace("$", "").strip()
                        break

                if applies_to == "above_threshold" and threshold:
                    try:
                        if float(amount_str) >= float(threshold):
                            matched_items.append(item)
                    except (ValueError, TypeError):
                        matched_items.append(item)
                else:
                    matched_items.append(item)

            if matched_items:
                for item in matched_items:
                    test_plan.append({
                        "test_ref": ref,
                        "attribute": name,
                        "procedure": procedure,
                        "applies_to": applies_to,
                        "sample_item_json": json.dumps(item),
                    })
            else:
                test_plan.append({
                    "test_ref": ref,
                    "attribute": name,
                    "procedure": procedure,
                    "applies_to": applies_to,
                    "sample_item_json": json.dumps({"_type": "no_applicable_items", "filter": applies_to}),
                })

    return json.dumps({
        "test_plan": test_plan,
        "total_tests": len(test_plan),
        "attributes_count": len(attributes),
        "sample_size": len(selected_sample),
        "instruction": "Execute EVERY entry in test_plan using execute_test. Do NOT skip any.",
    }, indent=2)


# =========================================================================
# Test execution + report
# =========================================================================

def _run_pre_checks(test_ref: str, attribute: str, sample_item: dict, control_context_str: str) -> str:
    """Run deterministic data-level checks before the LLM analysis.

    Returns a string of pre-check findings to inject into the LLM prompt.
    """
    findings = []
    attr_lower = attribute.lower()

    try:
        ctx = json.loads(control_context_str) if isinstance(control_context_str, str) else control_context_str
    except Exception:
        ctx = {}
    rules = ctx.get("rules", {})

    preparer = sample_item.get("Preparer", "")
    approver = sample_item.get("Approver", "")
    if preparer and approver and ("self-approval" in attr_lower or "dual auth" in attr_lower):
        if preparer.strip().lower() == approver.strip().lower():
            findings.append(f"DATA CHECK FAIL: Preparer ({preparer}) = Approver ({approver}). Self-approval detected.")
        else:
            findings.append(f"DATA CHECK PASS: Preparer ({preparer}) != Approver ({approver}). Dual authorization verified.")

    threshold = rules.get("threshold_gbp", rules.get("threshold_usd", 0))
    if threshold and ("threshold" in attr_lower or "above" in attr_lower):
        for key in ("Amount_GBP", "Amount_in_GBP", "Amount"):
            if key in sample_item:
                try:
                    amt = float(str(sample_item[key]).replace(",", "").replace("£", "").replace("$", ""))
                    if amt >= float(threshold):
                        findings.append(f"DATA CHECK: Amount ({amt:,.2f}) >= threshold ({threshold:,}). Finance Director review required.")
                    else:
                        findings.append(f"DATA CHECK: Amount ({amt:,.2f}) < threshold ({threshold:,}). Below threshold — attribute may be N/A.")
                except (ValueError, TypeError):
                    pass
                break

    if "supporting doc" in attr_lower or "documentation" in attr_lower:
        sup = sample_item.get("Supporting_Doc", sample_item.get("supporting_doc", ""))
        if isinstance(sup, str) and not sup.strip():
            findings.append("DATA CHECK FAIL: Supporting_Doc field is empty — no documentation attached.")
        elif sup:
            findings.append(f"DATA CHECK PASS: Supporting document referenced: {sup}")

    if "period" in attr_lower or "posting" in attr_lower:
        posting_date = sample_item.get("Posting_Date", "")
        period = sample_item.get("Period", "")
        if posting_date and period:
            findings.append(f"DATA CHECK: Posting_Date={posting_date}, Period={period}. Verify period assignment is correct.")

    return "\n".join(findings) if findings else ""


@tool
def execute_test(
    test_ref: str,
    attribute: str,
    procedure: str,
    control_context: str,
    sample_item_json: str,
    evidence_summary: str = "",
) -> str:
    """Execute a specific testing attribute against a SINGLE sample item.

    Args:
        test_ref: The testing attribute reference (A, B, C, D, etc.).
        attribute: The testing attribute question/name.
        procedure: The testing procedure to follow.
        control_context: JSON string with control_id, control_name, domain, and rules.
        sample_item_json: JSON string of the single sample item to test.
        evidence_summary: Summary of reviewed supporting documents.

    Returns:
        JSON string with test_ref, sample_item, and llm_analysis.
    """
    from databricks_langchain import ChatDatabricks
    from agent.config import LLM_ENDPOINT
    from agent.prompts import TEST_EXECUTION_PROMPT

    import time as _time

    sample_item = json.loads(sample_item_json) if sample_item_json else {}

    pre_checks = _run_pre_checks(test_ref, attribute, sample_item, control_context)

    with mlflow.start_span(name=f"test_{test_ref}", span_type="TOOL") as span:
        span.set_inputs({"test_ref": test_ref, "attribute": attribute, "sample_item": sample_item, "pre_checks": pre_checks})

        evidence_with_checks = evidence_summary[:3000]
        if pre_checks:
            evidence_with_checks = f"**Automated Data Pre-Checks:**\n{pre_checks}\n\n{evidence_with_checks}"

        prompt = TEST_EXECUTION_PROMPT.format(
            ref=test_ref,
            control_context=control_context[:2000],
            attribute=attribute,
            procedure=procedure,
            sample_item=sample_item_json[:2000],
            evidence_summary=evidence_with_checks,
        )

        llm = ChatDatabricks(endpoint=LLM_ENDPOINT, temperature=0)

        max_retries = 4
        backoff_base = 8
        response = None
        for attempt in range(max_retries + 1):
            try:
                response = llm.invoke(prompt)
                break
            except Exception as e:
                err_str = str(e)
                is_rate_limit = "429" in err_str or "REQUEST_LIMIT_EXCEEDED" in err_str or "rate limit" in err_str.lower()
                if is_rate_limit and attempt < max_retries:
                    wait = backoff_base * (2 ** attempt)
                    print(f"[execute_test] 429 on attempt {attempt+1} for {test_ref}, retrying in {wait}s...")
                    _time.sleep(wait)
                    continue
                raise

        span.set_outputs({"response_length": len(response.content), "pre_checks": pre_checks})

    return json.dumps({
        "test_ref": test_ref,
        "sample_item": sample_item,
        "llm_analysis": response.content,
    }, indent=2)


@tool
def compile_results(
    control_id: str,
    control_name: str,
    engagement_number: str,
    domain: str,
    population_size: int,
    sample_size: int,
    testing_attributes_json: str,
    test_results_json: str,
    rules_json: str = "{}",
) -> str:
    """Compile all test results into the final assessment report.

    Args:
        control_id: The control identifier.
        control_name: The control name.
        engagement_number: The engagement reference number.
        domain: The control domain.
        population_size: Total items in the population.
        sample_size: Number of sampled items tested.
        testing_attributes_json: JSON of testing attributes.
        test_results_json: JSON of ALL test results.
        rules_json: JSON of control-specific rules.

    Returns:
        Formatted markdown report.
    """
    from databricks_langchain import ChatDatabricks
    from agent.config import LLM_ENDPOINT
    from agent.prompts import REPORT_GENERATION_PROMPT

    import time as _time

    with mlflow.start_span(name="generate_report", span_type="TOOL") as span:
        span.set_inputs({
            "control_id": control_id,
            "engagement_number": engagement_number,
            "domain": domain,
            "population_size": population_size,
            "sample_size": sample_size,
        })

        prompt = REPORT_GENERATION_PROMPT.format(
            control_id=control_id,
            control_name=control_name,
            engagement_number=engagement_number,
            domain=domain,
            population_size=population_size,
            sample_size=sample_size,
            testing_attributes=testing_attributes_json[:3000],
            test_results=test_results_json[:8000],
            rules=rules_json[:2000],
        )

        llm = ChatDatabricks(endpoint=LLM_ENDPOINT, temperature=0)

        max_retries = 4
        backoff_base = 8
        response = None
        for attempt in range(max_retries + 1):
            try:
                response = llm.invoke(prompt)
                break
            except Exception as e:
                err_str = str(e)
                is_rate_limit = "429" in err_str or "REQUEST_LIMIT_EXCEEDED" in err_str or "rate limit" in err_str.lower()
                if is_rate_limit and attempt < max_retries:
                    wait = backoff_base * (2 ** attempt)
                    print(f"[compile_results] 429 on attempt {attempt+1}, retrying in {wait}s...")
                    _time.sleep(wait)
                    continue
                raise

        span.set_outputs({
            "report_length": len(response.content),
            "has_exceptions": "ISS-" in response.content,
        })

    return response.content


@tool
def save_report(
    project_path: str,
    report_content: str,
    report_format: str = "markdown",
    control_id: str = "",
    control_name: str = "",
) -> str:
    """Save the final report to the project directory.

    Args:
        project_path: Project directory name.
        report_content: The full report content (markdown).
        report_format: "markdown" or "both" (markdown + JSON summary).
        control_id: Control ID for the report filename (e.g. "CTRL-FIN-042").
        control_name: Control name for the report filename.

    Returns:
        JSON with saved file paths, volume URL, accessible report_url, and status.
        The report_url is a clickable link that anyone with app access can open.
        ALWAYS include report_url in email notifications.
    """
    import re as _re
    from agent.run_context import get_report_url, get_run_id, get_project_dir
    from agent import volume_store as vs

    with mlflow.start_span(name="save_report", span_type="TOOL") as span:
        span.set_inputs({"project_path": project_path, "format": report_format})

        saved_files = []
        volume_url = None
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        name_slug = _re.sub(r"[^a-zA-Z0-9]+", "_", (control_name or "report")).strip("_").lower()[:50]
        ctrl_prefix = (control_id or "report").replace("-", "_").upper()
        base_name = f"{ctrl_prefix}_{name_slug}_{timestamp}"

        run_id = get_run_id()
        proj_dir = get_project_dir() or project_path

        if run_id and proj_dir:
            vol = vs.save_run_artifact(proj_dir, run_id, f"{base_name}.md", report_content)
            if vol:
                saved_files.append(vol)
                volume_url = vol

        local_dir = Path(PROJECTS_LOCAL_PATH) / project_path
        if local_dir.exists():
            report_file = local_dir / f"{base_name}.md"
            report_file.write_text(report_content)
            saved_files.append(str(report_file))

        report_url = get_report_url()

        span.set_outputs({"saved_files": saved_files, "volume_url": volume_url, "report_url": report_url})

    return json.dumps({
        "status": "saved",
        "files": saved_files,
        "volume_url": volume_url,
        "report_url": report_url,
        "filename": f"{base_name}.md",
        "report_length": len(report_content),
    }, indent=2)


# =========================================================================
# Email sending via Gmail SMTP
# =========================================================================


def _find_workbook_for_attachment(project_path: str = "") -> tuple[bytes | None, str]:
    """Locate the most recent completed .xlsx in the current run folder.

    Volume-first: checks UC Volume, then falls back to local cache.
    Falls back to scanning all local project dirs if thread context is lost.
    Returns (file_bytes, filename) or (None, "") if not found.
    """
    from agent.run_context import get_run_id, get_project_dir
    from agent import volume_store as vs

    run_id = get_run_id()
    proj = get_project_dir() or project_path
    print(f"[_find_workbook] run_id={run_id!r}, proj={proj!r}, project_path={project_path!r}")

    if run_id and proj:
        run_base = vs.run_path(proj, run_id)
        items = vs.list_dir(run_base)
        print(f"[_find_workbook] Volume listing {run_base}: {len(items)} items")
        xlsx_items = [
            it for it in items
            if it.path and it.path.endswith(".xlsx") and "_completed_" in it.path
        ]
        if xlsx_items:
            xlsx_items.sort(key=lambda x: x.path, reverse=True)
            chosen = xlsx_items[0].path
            name = Path(chosen).name
            try:
                data = vs.download_bytes(chosen)
                print(f"[_find_workbook] Downloaded from volume: {name} ({len(data)} bytes)")
                return data, name
            except Exception as exc:
                print(f"[_find_workbook] Volume download failed for {chosen}: {exc}")

        local_run = Path(PROJECTS_LOCAL_PATH) / proj / "runs" / run_id
        print(f"[_find_workbook] Checking local: {local_run} (exists={local_run.exists()})")
        if local_run.exists():
            xlsx_files = sorted(local_run.glob("*_completed_*.xlsx"), reverse=True)
            if xlsx_files:
                print(f"[_find_workbook] Found local: {xlsx_files[0].name}")
                return xlsx_files[0].read_bytes(), xlsx_files[0].name

    # Broad fallback: scan ALL local project dirs for the most recent completed workbook.
    # Handles the case where thread context is lost (ToolNode ThreadPoolExecutor).
    print("[_find_workbook] Broad fallback: scanning all local project directories")
    all_xlsx: list[Path] = []
    base = Path(PROJECTS_LOCAL_PATH)
    if base.exists():
        all_xlsx = sorted(base.rglob("*_completed_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if all_xlsx:
        chosen = all_xlsx[0]
        print(f"[_find_workbook] Fallback found: {chosen.name} (mtime={chosen.stat().st_mtime})")
        return chosen.read_bytes(), chosen.name

    print("[_find_workbook] No workbook found anywhere")
    return None, ""


def _build_html_email(
    subject: str,
    body: str,
    report_url: str = "",
    xlsx_name: str = "",
) -> str:
    """Build a professional HTML email from the agent's body text."""
    import re as _re

    body_lines = body.strip().splitlines()
    body_html_parts: list[str] = []
    for line in body_lines:
        stripped = line.strip()
        if not stripped:
            body_html_parts.append('<div style="height:12px"></div>')
        elif stripped.startswith("# "):
            body_html_parts.append(
                f'<h2 style="color:#1a1a2e;margin:18px 0 8px;font-size:18px;border-bottom:1px solid #e5e7eb;padding-bottom:6px">'
                f'{stripped[2:]}</h2>'
            )
        elif stripped.startswith("## "):
            body_html_parts.append(
                f'<h3 style="color:#333;margin:14px 0 6px;font-size:15px">{stripped[3:]}</h3>'
            )
        elif stripped.startswith("- ") or stripped.startswith("• "):
            body_html_parts.append(
                f'<div style="padding:2px 0 2px 16px;color:#374151">&#8226; {stripped[2:]}</div>'
            )
        elif _re.match(r"^\*\*(.+?)\*\*:?\s*(.*)$", stripped):
            m = _re.match(r"^\*\*(.+?)\*\*:?\s*(.*)$", stripped)
            body_html_parts.append(
                f'<div style="padding:2px 0;color:#374151"><strong>{m.group(1)}</strong>: {m.group(2)}</div>'
            )
        else:
            body_html_parts.append(f'<div style="padding:2px 0;color:#374151">{stripped}</div>')

    body_html = "\n".join(body_html_parts)
    body_html = _re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", body_html)

    report_section = ""
    if report_url:
        report_section = f"""
        <div style="margin:24px 0;text-align:center">
          <a href="{report_url}"
             style="display:inline-block;background:#1a73e8;color:#ffffff;padding:12px 28px;
                    border-radius:6px;text-decoration:none;font-weight:600;font-size:14px">
            View Full Report &rarr;
          </a>
        </div>"""

    attachment_note = ""
    if xlsx_name:
        attachment_note = f"""
        <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:10px 14px;margin:16px 0;font-size:13px;color:#166534">
          📎 Completed workbook attached: <strong>{xlsx_name}</strong>
        </div>"""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"/></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif">
  <div style="max-width:640px;margin:0 auto;background:#ffffff;border-radius:8px;overflow:hidden;margin-top:20px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,0.1)">
    <!-- Header -->
    <div style="background:linear-gradient(135deg,#1a1a2e 0%,#16213e 100%);padding:24px 32px">
      <div style="color:#ffffff;font-size:20px;font-weight:700;letter-spacing:-0.3px">GSK Controls Evidence Review</div>
      <div style="color:#94a3b8;font-size:12px;margin-top:4px">Automated Compliance Assessment</div>
    </div>

    <!-- Body -->
    <div style="padding:24px 32px;font-size:14px;line-height:1.65;color:#374151">
      {body_html}
      {report_section}
      {attachment_note}
    </div>

    <!-- Footer -->
    <div style="background:#f9fafb;border-top:1px solid #e5e7eb;padding:16px 32px;font-size:11px;color:#9ca3af;text-align:center">
      This is an automated notification from the GSK Compliance Agent.
      Please do not reply to this email.
    </div>
  </div>
</body>
</html>"""


@tool
def send_email(
    to: str,
    subject: str,
    body: str,
    cc: str = "",
    importance: str = "normal",
    project_path: str = "",
    report_url: str = "",
    attach_workbook: bool = True,
) -> str:
    """Send a professionally formatted email notification with the report.

    The tool wraps your body text into a branded HTML email template
    automatically. Just pass clear, structured body text — no need to
    write HTML yourself.

    IMPORTANT: If save_report returned a report_url, pass it here so the
    email contains a clickable button to the full report.

    Args:
        to: Recipient email address (comma-separated for multiple).
        subject: Email subject line.
        body: Email body content. Use markdown-style formatting:
              **bold** for labels, - for bullet points, # for headings.
              The tool will convert this into professional HTML.
        cc: Optional CC recipients (comma-separated).
        importance: "low", "normal", or "high".
        project_path: Optional project dir for saving a copy.
        report_url: Optional URL to the report (rendered as a button).
        attach_workbook: If True, attach the completed .xlsx workbook.

    Returns:
        JSON with status ("sent" or "fallback"), recipient, and details.
    """
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from agent.config import SMTP_EMAIL, SMTP_DISPLAY_NAME, get_smtp_password
    from agent.run_context import get_report_url, get_run_id, get_project_dir

    smtp_email = SMTP_EMAIL
    smtp_password = get_smtp_password()

    url = report_url or get_report_url()

    with mlflow.start_span(name="send_email", span_type="TOOL") as span:
        span.set_inputs({"to": to, "subject": subject, "importance": importance, "report_url": url, "attach_workbook": attach_workbook})

        xlsx_bytes: bytes | None = None
        xlsx_name: str = ""
        if attach_workbook:
            xlsx_bytes, xlsx_name = _find_workbook_for_attachment(project_path)

        html_body = _build_html_email(subject, body, url, xlsx_name)

        if smtp_password:
            msg = MIMEMultipart("mixed")
            msg["From"] = f"{SMTP_DISPLAY_NAME} <{smtp_email}>"
            msg["To"] = to
            msg["Subject"] = subject
            if cc:
                msg["Cc"] = cc
            if importance == "high":
                msg["X-Priority"] = "1"

            msg.attach(MIMEText(html_body, "html"))

            if xlsx_bytes and xlsx_name:
                part = MIMEApplication(xlsx_bytes, Name=xlsx_name)
                part["Content-Disposition"] = f'attachment; filename="{xlsx_name}"'
                msg.attach(part)

            all_recipients = [a.strip() for a in to.split(",")]
            if cc:
                all_recipients.extend(a.strip() for a in cc.split(",") if a.strip())

            try:
                with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
                    server.starttls()
                    server.login(smtp_email, smtp_password)
                    server.sendmail(smtp_email, all_recipients, msg.as_string())

                span.set_outputs({"status": "sent", "report_url": url, "attached_xlsx": xlsx_name or None})
                return json.dumps({
                    "status": "sent",
                    "to": to,
                    "subject": subject,
                    "method": "gmail_smtp",
                    "report_url": url,
                    "attached_workbook": xlsx_name or None,
                }, indent=2)
            except Exception as e:
                span.set_outputs({"status": "smtp_error", "error": str(e)})
                return json.dumps({
                    "status": "error",
                    "to": to,
                    "subject": subject,
                    "method": "gmail_smtp",
                    "error": str(e),
                    "report_url": url,
                }, indent=2)

        else:
            eml = EmailMessage()
            eml["From"] = f"{SMTP_DISPLAY_NAME} <{smtp_email}>"
            eml["To"] = to
            eml["Subject"] = subject
            eml["Date"] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
            if cc:
                eml["Cc"] = cc
            eml.set_content(html_body, subtype="html")

            saved_path = None
            if project_path:
                local_dir = Path(PROJECTS_LOCAL_PATH) / project_path
                if local_dir.exists():
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    eml_file = local_dir / f"sent_email_{timestamp}.eml"
                    eml_file.write_text(eml.as_string())
                    saved_path = str(eml_file)

            span.set_outputs({"status": "no_smtp_credentials", "saved_path": saved_path, "report_url": url})
            return json.dumps({
                "status": "not_sent",
                "to": to,
                "subject": subject,
                "method": "eml_fallback",
                "note": "SMTP_APP_PASSWORD not configured. Email saved as .eml file.",
                "saved_path": saved_path,
                "report_url": url,
            }, indent=2)


# =========================================================================
# Plan announcement
# =========================================================================

@tool
def announce_plan(steps: str) -> str:
    """Announce the structured assessment plan to the user.

    Call this ONCE, immediately after loading the engagement, to declare the
    high-level steps you will follow. The frontend renders this as a live
    checklist that updates as each phase completes.

    Args:
        steps: JSON array of plan steps. Each element must have:
               - "id": short identifier (e.g. "load", "evidence", "test")
               - "label": human-readable description of the phase
               - "detail": (optional) extra context
    Returns:
        Confirmation JSON.
    """
    parsed = json.loads(steps)
    return json.dumps({"status": "plan_announced", "step_count": len(parsed), "steps": parsed})


# =========================================================================
# User interaction
# =========================================================================

@tool
def ask_user(question: str, options: Optional[str] = None) -> str:
    """Ask the user for clarification when uncertain.

    Args:
        question: The question to ask.
        options: Optional comma-separated suggested options.

    Returns:
        A message for the user; their response arrives as the next message.
    """
    result = {"type": "user_question", "question": question}
    if options:
        result["suggested_options"] = [o.strip() for o in options.split(",")]
    return json.dumps(result, indent=2)


# ---------------------------------------------------------------------------
# Deterministic aggregation — removes LLM from pass/fail counting
# ---------------------------------------------------------------------------

_SAMPLE_ID_FIELDS = [
    "JE_Number", "Order_No", "Location_ID", "User_ID", "Document_No",
    "Inspection_ID", "Sample_No", "Row_No", "Location_Name",
]


def _extract_sample_id(sample: dict) -> str:
    """Pull a human-readable ID from a sample item dict."""
    for field in _SAMPLE_ID_FIELDS:
        val = sample.get(field)
        if val:
            return str(val)
    return json.dumps(sample, sort_keys=True)[:80]


def _parse_llm_analysis(raw: str) -> dict:
    """Robustly parse the JSON that execute_test's LLM returns."""
    text = raw.strip()
    if text.startswith("```"):
        text = text.strip("`").lstrip("json").lstrip("\n")
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {"result": "Unknown", "narrative": raw[:500]}


@tool
def aggregate_test_results(
    batch_results_json: str,
) -> str:
    """Deterministically aggregate per-sample test results into per-attribute summaries.

    Takes the raw JSON output from batch_execute_tests and produces a
    structured test_results_json array suitable for fill_workbook and
    compile_results. Grouping, pass/fail determination, and exception
    counting are all done in Python — no LLM judgment involved.

    Rules:
      - One entry per ref letter.
      - A ref is "Pass" only if ALL its samples passed.
      - A ref is "Fail" if ANY sample failed.
      - A ref is "Not Applicable" only if ALL samples were N/A.
      - Otherwise "Partial".
      - Exceptions are merged into one per failing ref.

    Args:
        batch_results_json: The full JSON string returned by batch_execute_tests.

    Returns:
        JSON array of per-attribute results, ready for fill_workbook.
    """
    batch = json.loads(batch_results_json) if isinstance(batch_results_json, str) else batch_results_json
    raw_results = batch.get("results", batch) if isinstance(batch, dict) else batch

    by_ref: dict[str, dict] = {}

    for entry in raw_results:
        ref = entry.get("test_ref", "?").strip().upper()
        sample = entry.get("sample_item", {})
        if isinstance(sample, str):
            try:
                sample = json.loads(sample)
            except json.JSONDecodeError:
                sample = {}

        parsed = _parse_llm_analysis(entry.get("llm_analysis", "{}"))
        sample_id = _extract_sample_id(sample)
        result_lower = parsed.get("result", "Unknown").strip().lower()

        if ref not in by_ref:
            by_ref[ref] = {
                "results": [],
                "narratives": [],
                "sample_ids": [],
                "exceptions": [],
            }

        bucket = by_ref[ref]
        bucket["results"].append(result_lower)
        bucket["narratives"].append(f"[{sample_id}]: {parsed.get('narrative', '')}")
        bucket["sample_ids"].append(sample_id)

        if result_lower == "fail":
            bucket["exceptions"].append({
                "description": parsed.get("exception") or parsed.get("narrative", "Test failed"),
                "severity": parsed.get("severity", "Medium"),
                "sample_id": sample_id,
            })

    aggregated = []
    for ref in sorted(by_ref.keys()):
        bucket = by_ref[ref]
        results_set = set(bucket["results"])

        if results_set == {"pass"}:
            agg_result = "Pass"
        elif "fail" in results_set:
            agg_result = "Fail"
        elif results_set <= {"not applicable"}:
            agg_result = "Not Applicable"
        else:
            agg_result = "Partial"

        merged_exceptions = []
        if bucket["exceptions"]:
            all_affected = sorted({e["sample_id"] for e in bucket["exceptions"]})
            descriptions = []
            severity = "Medium"
            severity_rank = {"critical": 4, "high": 3, "medium": 2, "low": 1}
            for exc in bucket["exceptions"]:
                descriptions.append(exc["description"])
                exc_sev = exc.get("severity", "Medium")
                if severity_rank.get(exc_sev.lower(), 0) > severity_rank.get(severity.lower(), 0):
                    severity = exc_sev

            merged_exceptions = [{
                "description": "; ".join(dict.fromkeys(descriptions)),
                "severity": severity,
                "affected_samples": all_affected,
                "root_cause": "See test narrative",
                "remediation": "Review and remediate",
            }]

        aggregated.append({
            "ref": ref,
            "result": agg_result,
            "narrative": "\n".join(bucket["narratives"]),
            "sample_items_tested": bucket["sample_ids"],
            "exceptions": merged_exceptions,
        })

    return json.dumps(aggregated, indent=2)


@tool
def fill_workbook(
    project_path: str,
    test_results_json: str,
    control_id: str = "",
) -> str:
    """Fill in the engagement workbook with test results and exceptions.

    Opens the original engagement_workbook.xlsx, writes test outcomes into the
    Testing Table sheet (Answer column), and populates the Issue template sheet
    with any exceptions. Saves the completed workbook as a run artifact.

    Args:
        project_path: Project directory name (e.g. "fin_042").
        test_results_json: JSON array of test results. Each entry should have:
            - ref: The testing attribute ref letter (A, B, C, …)
            - result: "Pass", "Fail", "Not Applicable", or "Partial"
            - narrative: Explanation / finding narrative
            - sample_items_tested: (optional) list of sample item IDs tested
            - exceptions: (optional) list of exception dicts with description,
              severity, affected_samples, root_cause, remediation, owner
        control_id: Control ID for naming the output file.

    Returns:
        JSON with saved file paths, download URL, and status.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from agent.run_context import get_run_id, get_project_dir, get_app_base_url, get_artifact_url

    with mlflow.start_span(name="fill_workbook", span_type="TOOL") as span:
        span.set_inputs({"project_path": project_path, "control_id": control_id})

        wb_path = _resolve_project_file(project_path, "engagement_workbook.xlsx")
        wb_content = _read_file_bytes(wb_path)
        wb = openpyxl.load_workbook(io.BytesIO(wb_content))

        test_results = json.loads(test_results_json)

        results_by_ref = {}
        for tr in test_results:
            ref = tr.get("ref", "").strip().upper()
            if ref:
                if ref not in results_by_ref:
                    results_by_ref[ref] = tr
                else:
                    existing = results_by_ref[ref]
                    existing_narr = existing.get("narrative", "")
                    new_narr = tr.get("narrative", "")
                    existing["narrative"] = f"{existing_narr}\n{new_narr}".strip()
                    if tr.get("result", "").lower() == "fail":
                        existing["result"] = "Fail"
                    for exc in tr.get("exceptions", []):
                        existing.setdefault("exceptions", []).append(exc)

        pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        partial_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        header_font = Font(bold=True, size=10)
        wrap_align = Alignment(wrap_text=True, vertical="top")

        testing_sheet = None
        for sn in wb.sheetnames:
            if "testing" in sn.lower() and "table" in sn.lower():
                testing_sheet = wb[sn]
                break
        if not testing_sheet:
            for sn in wb.sheetnames:
                if "testing" in sn.lower():
                    testing_sheet = wb[sn]
                    break

        if testing_sheet:
            header_row = None
            ref_col = None
            answer_col = None
            procedure_col = None
            for row_idx, row in enumerate(testing_sheet.iter_rows(min_row=1, max_row=testing_sheet.max_row), start=1):
                for cell in row:
                    val = str(cell.value or "").lower().strip()
                    if val == "ref":
                        header_row = row_idx
                        ref_col = cell.column
                    elif val == "answer" or val == "result":
                        answer_col = cell.column
                    elif val == "procedure":
                        procedure_col = cell.column

            if header_row and ref_col:
                if not answer_col:
                    answer_col = (procedure_col or ref_col) + 2
                    testing_sheet.cell(row=header_row, column=answer_col, value="Answer").font = header_font

                result_col = answer_col + 1
                testing_sheet.cell(row=header_row, column=result_col, value="Result").font = header_font

                for row_idx in range(header_row + 1, testing_sheet.max_row + 1):
                    ref_val = str(testing_sheet.cell(row=row_idx, column=ref_col).value or "").strip().upper()
                    if ref_val in results_by_ref:
                        tr = results_by_ref[ref_val]
                        result_text = tr.get("result", "")
                        narrative = tr.get("narrative", "")

                        result_cell = testing_sheet.cell(row=row_idx, column=result_col)
                        result_cell.value = result_text
                        result_cell.alignment = wrap_align
                        if result_text.lower() == "pass":
                            result_cell.fill = pass_fill
                        elif result_text.lower() == "fail":
                            result_cell.fill = fail_fill
                        else:
                            result_cell.fill = partial_fill

                        answer_cell = testing_sheet.cell(row=row_idx, column=answer_col)
                        answer_cell.value = narrative
                        answer_cell.alignment = wrap_align

                testing_sheet.column_dimensions[openpyxl.utils.get_column_letter(answer_col)].width = 60
                testing_sheet.column_dimensions[openpyxl.utils.get_column_letter(result_col)].width = 15

        issue_sheet = None
        for sn in wb.sheetnames:
            if "issue" in sn.lower():
                issue_sheet = wb[sn]
                break

        if issue_sheet:
            def _cell_val(v):
                if isinstance(v, list):
                    return ", ".join(str(i) for i in v)
                if isinstance(v, dict):
                    return json.dumps(v)
                return v

            agg: dict[str, dict] = {}
            for tr in test_results:
                ref = tr.get("ref", "?").strip().upper()
                if tr.get("result", "").lower() not in ("fail", "partial"):
                    continue

                exceptions = tr.get("exceptions", [])
                raw_samples = tr.get("sample_items_tested", [])
                if isinstance(raw_samples, str):
                    raw_samples = [raw_samples]
                samples_set = set(str(s) for s in raw_samples if s)

                if ref not in agg:
                    best_exc = exceptions[0] if exceptions else {}
                    agg[ref] = {
                        "description": best_exc.get("description", tr.get("narrative", "Test failed")),
                        "severity": best_exc.get("severity", "Medium"),
                        "samples": samples_set,
                        "root_cause": best_exc.get("root_cause", "See test narrative"),
                        "remediation": best_exc.get("remediation", "Review and remediate"),
                        "owner": best_exc.get("owner", ""),
                    }
                else:
                    entry = agg[ref]
                    entry["samples"] |= samples_set
                    for exc in exceptions:
                        if exc.get("description"):
                            entry["description"] = f"{entry['description']}; {exc['description']}"

            issue_counter = 1
            next_row = 3
            for ref in sorted(agg.keys()):
                entry = agg[ref]
                issue_id = f"ISS-{control_id or 'CTRL'}-{issue_counter:03d}"
                issue_sheet.cell(row=next_row, column=1, value=issue_id)
                issue_sheet.cell(row=next_row, column=2, value=ref)
                issue_sheet.cell(row=next_row, column=3, value=_cell_val(entry.get("severity", "Medium")))
                desc_text = entry.get("description", "")
                if len(desc_text) > 32000:
                    desc_text = desc_text[:32000] + "..."
                desc_cell = issue_sheet.cell(row=next_row, column=4, value=_cell_val(desc_text))
                desc_cell.alignment = wrap_align
                issue_sheet.cell(row=next_row, column=5, value=_cell_val(", ".join(sorted(entry.get("samples", set())))))
                issue_sheet.cell(row=next_row, column=6, value=_cell_val(entry.get("root_cause", "")))
                issue_sheet.cell(row=next_row, column=7, value=_cell_val(entry.get("remediation", "")))
                issue_sheet.cell(row=next_row, column=8, value=_cell_val(entry.get("owner", "")))
                issue_sheet.cell(row=next_row, column=9, value="")
                issue_sheet.cell(row=next_row, column=10, value="Open")
                next_row += 1
                issue_counter += 1

            if issue_counter == 1:
                issue_sheet.cell(row=3, column=1, value="No exceptions identified")
                issue_sheet.cell(row=3, column=3, value="N/A")
                issue_sheet.cell(row=3, column=10, value="Closed")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ctrl_prefix = (control_id or "workbook").replace("-", "_").upper()
        output_filename = f"{ctrl_prefix}_completed_{timestamp}.xlsx"

        output_buf = io.BytesIO()
        wb.save(output_buf)
        output_bytes = output_buf.getvalue()

        from agent import volume_store as vs

        saved_files = []
        volume_url = None

        run_id = get_run_id()
        proj_dir = get_project_dir() or project_path

        print(f"[fill_workbook] run_id={run_id!r}, proj_dir={proj_dir!r}, filename={output_filename}")

        if run_id and proj_dir:
            vol = vs.save_run_artifact(proj_dir, run_id, output_filename, output_bytes)
            if vol:
                saved_files.append(vol)
                volume_url = vol
                print(f"[fill_workbook] Saved to volume: {vol}")
            else:
                print(f"[fill_workbook] WARNING: Volume upload returned empty for {output_filename}")
        else:
            print(f"[fill_workbook] WARNING: Missing run_id or proj_dir — skipping volume save")

        try:
            local_dir = Path(PROJECTS_LOCAL_PATH) / project_path
            if local_dir.exists():
                out_file = local_dir / output_filename
                out_file.write_bytes(output_bytes)
                saved_files.append(str(out_file))
            if run_id and proj_dir:
                runs_dir = Path(PROJECTS_LOCAL_PATH) / proj_dir / "runs" / run_id
                runs_dir.mkdir(parents=True, exist_ok=True)
                (runs_dir / output_filename).write_bytes(output_bytes)
                saved_files.append(str(runs_dir / output_filename))
                print(f"[fill_workbook] Saved local copy: {runs_dir / output_filename}")
        except Exception as e:
            print(f"[fill_workbook] Local save error: {e}")

        workbook_url = get_artifact_url(output_filename)
        attrs_filled = len(results_by_ref)
        exceptions_count = len(agg) if issue_sheet else sum(
            1 for tr in test_results if tr.get("result", "").lower() in ("fail", "partial")
        )

        span.set_outputs({
            "filename": output_filename,
            "attrs_filled": attrs_filled,
            "exceptions": exceptions_count,
            "workbook_url": workbook_url,
            "volume_url": volume_url,
        })

    return json.dumps({
        "status": "saved",
        "filename": output_filename,
        "files": saved_files,
        "volume_url": volume_url,
        "workbook_url": workbook_url,
        "attrs_filled": attrs_filled,
        "exceptions_logged": exceptions_count,
    }, indent=2)


# =========================================================================
# Batch tools — parallel evidence review and test execution
# =========================================================================

def _dispatch_evidence_review(
    file_info: dict,
    project_path: str,
    control_context: str,
) -> dict:
    """Internal dispatcher: review a single evidence file (no @tool decorator)."""
    file_type = file_info.get("type", "pdf")
    rel_path = file_info.get("path", "")
    focus = file_info.get("focus", "")
    file_path = _resolve_project_file(project_path, rel_path)

    if file_type == "email":
        return json.loads(analyze_email.invoke({
            "file_path": file_path,
            "context": control_context,
            "focus_area": focus,
        }))
    elif file_type in ("screenshot", "image", "photo"):
        return json.loads(review_screenshot.invoke({
            "file_path": file_path,
            "context": control_context,
            "focus_area": focus,
        }))
    else:
        return json.loads(review_document.invoke({
            "file_path": file_path,
            "context": control_context,
            "focus_area": focus,
        }))


def _execute_single_test(
    test_entry: dict,
    control_context: str,
    evidence_summary: str,
) -> dict:
    """Internal dispatcher: execute a single test (no @tool decorator)."""
    return json.loads(execute_test.invoke({
        "test_ref": test_entry["test_ref"],
        "attribute": test_entry["attribute"],
        "procedure": test_entry.get("procedure", ""),
        "control_context": control_context,
        "sample_item_json": test_entry.get("sample_item_json", "{}"),
        "evidence_summary": evidence_summary,
    }))


@tool
def batch_review_evidence(
    evidence_files_json: str,
    project_path: str,
    control_context: str,
) -> str:
    """Review ALL evidence files in parallel. Call this INSTEAD of reviewing
    files one-by-one. Dispatches PDFs to review_document, images to
    review_screenshot, and .eml files to analyze_email — all concurrently.

    Args:
        evidence_files_json: JSON array of evidence file objects from
            engagement.evidence_files. Each must have path, type, and focus.
        project_path: Project directory name (e.g. "fin_042").
        control_context: JSON string with control_id, control_name, rules.

    Returns:
        JSON with reviews list (one per file) and aggregate stats.
    """
    import time as _time
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from agent.config import MAX_PARALLEL_EVIDENCE
    from agent.run_context import snapshot_context, restore_context

    evidence_files = json.loads(evidence_files_json) if isinstance(evidence_files_json, str) else evidence_files_json
    evidence_files = sorted(evidence_files, key=lambda f: f.get("path", ""))
    ctx_snapshot = snapshot_context()

    with mlflow.start_span(name="batch_review_evidence", span_type="TOOL") as parent_span:
        parent_span.set_inputs({
            "file_count": len(evidence_files),
            "project": project_path,
            "max_workers": MAX_PARALLEL_EVIDENCE,
        })

        active_span = mlflow.get_current_active_span()
        request_id = active_span.request_id if active_span else None

        start = _time.time()
        reviews = [None] * len(evidence_files)
        errors = []

        def _worker(idx: int, file_info: dict) -> tuple[int, dict | None, str]:
            restore_context(ctx_snapshot)
            label = Path(file_info.get("path", "")).stem
            try:
                try:
                    span_ctx = mlflow.start_span(
                        name=f"review_{file_info.get('type', 'doc')}_{label}",
                        span_type="RETRIEVER",
                        request_id=request_id,
                    )
                    span_ctx.__enter__()
                    span_ctx.set_inputs({"file": file_info.get("path"), "type": file_info.get("type")})
                except Exception:
                    span_ctx = None

                result = _dispatch_evidence_review(file_info, project_path, control_context)

                if span_ctx:
                    try:
                        span_ctx.set_outputs({"analysis_length": len(result.get("analysis", ""))})
                        span_ctx.__exit__(None, None, None)
                    except Exception:
                        pass

                return idx, result, ""
            except Exception as e:
                import traceback
                print(f"[batch_review] Worker {idx} error: {e}\n{traceback.format_exc()}")
                return idx, None, f"{file_info.get('path', '?')}: {e}"

        with ThreadPoolExecutor(max_workers=MAX_PARALLEL_EVIDENCE) as pool:
            futures = {pool.submit(_worker, i, f): i for i, f in enumerate(evidence_files)}
            done_count = 0
            for future in as_completed(futures):
                idx, result, error = future.result()
                if result:
                    reviews[idx] = result
                if error:
                    errors.append(error)
                done_count += 1
                from agent.run_context import report_progress
                fname = evidence_files[idx].get("path", "").split("/")[-1] if idx < len(evidence_files) else ""
                report_progress(done_count, len(evidence_files), fname)

        elapsed = round(_time.time() - start, 1)
        reviews = [r for r in reviews if r is not None]

        parent_span.set_outputs({
            "files_reviewed": len(reviews),
            "errors": len(errors),
            "elapsed_seconds": elapsed,
            "parallel_workers": MAX_PARALLEL_EVIDENCE,
        })

    return json.dumps({
        "reviews": reviews,
        "files_reviewed": len(reviews),
        "errors": errors,
        "elapsed_seconds": elapsed,
        "parallel_workers": MAX_PARALLEL_EVIDENCE,
    }, indent=2)


@tool
def batch_execute_tests(
    test_plan_json: str,
    control_context: str,
    evidence_summary: str,
) -> str:
    """Execute ALL tests from the test plan in parallel. Call this INSTEAD
    of calling execute_test one-by-one.

    Args:
        test_plan_json: The full test_plan array from generate_test_plan output.
        control_context: JSON string with control_objective and rules.
        evidence_summary: Combined evidence review summaries from
            batch_review_evidence output.

    Returns:
        JSON with results list (one per test), pass/fail counts,
        and aggregate timing.
    """
    import time as _time
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from agent.config import MAX_PARALLEL_TESTS
    from agent.run_context import snapshot_context, restore_context

    test_plan = json.loads(test_plan_json) if isinstance(test_plan_json, str) else test_plan_json
    test_plan = sorted(test_plan, key=lambda e: (e.get("test_ref", ""), e.get("sample_item_json", "")))
    ctx_snapshot = snapshot_context()

    with mlflow.start_span(name="batch_execute_tests", span_type="TOOL") as parent_span:
        parent_span.set_inputs({
            "total_tests": len(test_plan),
            "max_workers": MAX_PARALLEL_TESTS,
        })

        active_span = mlflow.get_current_active_span()
        request_id = active_span.request_id if active_span else None

        start = _time.time()
        results = [None] * len(test_plan)
        errors = []

        def _worker(idx: int, entry: dict) -> tuple[int, dict | None, str]:
            restore_context(ctx_snapshot)
            ref = entry.get("test_ref", "?")
            try:
                try:
                    span_ctx = mlflow.start_span(
                        name=f"test_{ref}_{idx}",
                        span_type="TOOL",
                        request_id=request_id,
                    )
                    span_ctx.__enter__()
                    span_ctx.set_inputs({
                        "ref": ref,
                        "attribute": entry.get("attribute", ""),
                        "sample_item": entry.get("sample_item_json", "")[:200],
                    })
                except Exception:
                    span_ctx = None

                result = _execute_single_test(entry, control_context, evidence_summary)

                if span_ctx:
                    try:
                        llm_result = "unknown"
                        analysis = result.get("llm_analysis", "")
                        parsed = json.loads(analysis.strip().strip("`").lstrip("json\n"))
                        llm_result = parsed.get("result", "unknown")
                        span_ctx.set_outputs({"result": llm_result})
                        span_ctx.__exit__(None, None, None)
                    except Exception:
                        pass

                return idx, result, ""
            except Exception as e:
                import traceback
                print(f"[batch_tests] Worker {idx} error: {e}\n{traceback.format_exc()}")
                return idx, None, f"Test {ref} #{idx}: {e}"

        with ThreadPoolExecutor(max_workers=MAX_PARALLEL_TESTS) as pool:
            futures = {pool.submit(_worker, i, e): i for i, e in enumerate(test_plan)}
            done_count = 0
            for future in as_completed(futures):
                idx, result, error = future.result()
                if result:
                    results[idx] = result
                if error:
                    errors.append(error)
                done_count += 1
                from agent.run_context import report_progress
                ref = test_plan[idx].get("test_ref", "") if idx < len(test_plan) else ""
                report_progress(done_count, len(test_plan), ref)

        elapsed = round(_time.time() - start, 1)
        results = [r for r in results if r is not None]

        pass_count = 0
        fail_count = 0
        confidence_counts = {"High": 0, "Medium": 0, "Low": 0}
        low_confidence_refs = []
        for r in results:
            try:
                analysis = r.get("llm_analysis", "")
                parsed = json.loads(analysis.strip().strip("`").lstrip("json\n"))
                if parsed.get("result", "").lower() == "pass":
                    pass_count += 1
                elif parsed.get("result", "").lower() == "fail":
                    fail_count += 1
                conf = parsed.get("confidence", "")
                if conf in confidence_counts:
                    confidence_counts[conf] += 1
                if conf == "Low":
                    low_confidence_refs.append(r.get("test_ref", "?"))
            except Exception:
                pass

        parent_span.set_outputs({
            "total_tests": len(results),
            "passed": pass_count,
            "failed": fail_count,
            "confidence_counts": confidence_counts,
            "low_confidence_refs": low_confidence_refs,
            "errors": len(errors),
            "elapsed_seconds": elapsed,
            "parallel_workers": MAX_PARALLEL_TESTS,
        })

    return json.dumps({
        "results": results,
        "total_tests": len(results),
        "passed": pass_count,
        "failed": fail_count,
        "confidence_counts": confidence_counts,
        "low_confidence_refs": low_confidence_refs,
        "errors": errors,
        "elapsed_seconds": elapsed,
        "parallel_workers": MAX_PARALLEL_TESTS,
    }, indent=2)


