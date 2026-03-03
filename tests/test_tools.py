"""
Unit tests for every tool in agent/tools.py.

All external dependencies (Databricks SDK, LLM endpoints, MLflow tracing)
are mocked so tests run fast and offline.
"""

from __future__ import annotations

import io
import json
import os
import tempfile
from email.message import EmailMessage
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# ---------------------------------------------------------------------------
# Disable MLflow tracing before importing anything from agent
# ---------------------------------------------------------------------------
os.environ["MLFLOW_TRACKING_URI"] = "file:///tmp/mlflow_test"
os.environ["MLFLOW_ENABLE_SYSTEM_METRICS_LOGGING"] = "false"

import mlflow  # noqa: E402

mlflow.tracing.disable()

# ---------------------------------------------------------------------------
# Import tools
# ---------------------------------------------------------------------------
import agent.tools as tools_module  # noqa: E402
from agent.tools import (  # noqa: E402
    _read_file_bytes,
    _resolve_project_file,
    _read_excel_tabs,
    _run_pre_checks,
    _build_html_email,
)


# Patch targets: these are imported locally inside functions so we must
# patch them at the source module.
PATCH_CHAT = "databricks_langchain.ChatDatabricks"
PATCH_WS_CLIENT = "databricks.sdk.WorkspaceClient"


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def sample_data_dir():
    return Path(__file__).parent.parent / "sample_data"


@pytest.fixture
def fin042_dir(sample_data_dir):
    return sample_data_dir / "projects" / "fin_042"


@pytest.fixture
def engagement_json(fin042_dir):
    return json.loads((fin042_dir / "engagement.json").read_text())


@pytest.fixture
def tmp_project(tmp_path):
    """Create a minimal project structure in a temp directory."""
    proj = tmp_path / "test_proj"
    proj.mkdir()
    evidence = proj / "evidence"
    evidence.mkdir()

    engagement = {
        "number": "ENG-TEST-001",
        "name": "Test Control",
        "control_objective": {
            "control_id": "TST-001",
            "control_name": "Test Control",
            "domain": "Testing",
            "rules": {"threshold_gbp": 50000, "no_self_approval": True},
        },
        "testing_attributes": [
            {"ref": "A", "name": "Population completeness", "applies_to": "control_level"},
            {"ref": "B", "name": "Dual authorization", "applies_to": "all"},
        ],
        "evidence_files": [
            {"path": "evidence/doc.pdf", "type": "pdf", "focus": "policy"},
        ],
        "instructions": "Test the control.",
    }
    (proj / "engagement.json").write_text(json.dumps(engagement))

    (evidence / "doc.pdf").write_bytes(b"%PDF-1.4 fake pdf content")
    (evidence / "screenshot.png").write_bytes(b"\x89PNG fake image data")

    eml = EmailMessage()
    eml["From"] = "alice@example.com"
    eml["To"] = "bob@example.com"
    eml["Subject"] = "Approval for JE-001"
    eml["Date"] = "Mon, 01 Jan 2024 10:00:00 +0000"
    eml.set_content("I approve this journal entry.")
    (evidence / "approval.eml").write_bytes(eml.as_bytes())

    return proj


@pytest.fixture
def mock_llm():
    """A mock LLM that returns a plain string."""
    llm = MagicMock()
    llm.invoke = MagicMock(return_value=MagicMock(content="LLM mock response"))
    return llm


@pytest.fixture
def mock_llm_json():
    """A mock LLM that returns a JSON test result."""
    llm = MagicMock()
    llm.invoke = MagicMock(return_value=MagicMock(content=json.dumps({
        "result": "Pass",
        "narrative": "Yes, the control is effective.",
        "exception": None,
        "severity": None,
    })))
    return llm


# ============================================================================
# Helper: _read_file_bytes
# ============================================================================

class TestReadFileBytes:
    def test_reads_local_file(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_bytes(b"hello")
        assert _read_file_bytes(str(f)) == b"hello"

    def test_reads_from_sample_data(self, fin042_dir):
        content = _read_file_bytes(str(fin042_dir / "engagement.json"))
        data = json.loads(content)
        assert data["control_objective"]["control_id"] == "FIN-042"

    def test_raises_for_missing_file(self):
        with pytest.raises(FileNotFoundError, match="Cannot find"):
            _read_file_bytes("/nonexistent/path/to/file_xyz_abc.noext")

    def test_relative_path_tries_uc_volume(self):
        """Relative paths should attempt UC SDK download as last resort."""
        mock_resp = MagicMock()
        mock_resp.contents.read.return_value = b"uc content"
        mock_client = MagicMock()
        mock_client.files.download.return_value = mock_resp

        with patch(PATCH_WS_CLIENT, return_value=mock_client):
            with patch.object(tools_module, "SAMPLE_DATA_DIR", Path("/nonexistent_sample")):
                with patch.object(tools_module, "PROJECTS_LOCAL_PATH", "/nonexistent_local"):
                    with patch.object(tools_module, "PROJECTS_BASE_PATH", "/Volumes/test/vol/projects"):
                        result = _read_file_bytes("proj/evidence/file.pdf")
                        assert result == b"uc content"


# ============================================================================
# Helper: _resolve_project_file
# ============================================================================

class TestResolveProjectFile:
    def test_returns_local_if_exists(self, tmp_project):
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            result = _resolve_project_file("test_proj", "engagement.json")
            assert result.endswith("engagement.json")
            assert Path(result).exists()

    def test_falls_back_to_volume_path(self):
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", "/nonexistent"):
            result = _resolve_project_file("proj_x", "engagement.json")
            assert "/proj_x/engagement.json" in result


# ============================================================================
# Helper: _run_pre_checks
# ============================================================================

class TestRunPreChecks:
    def test_detects_self_approval(self):
        item = {"Preparer": "John Smith", "Approver": "john smith"}
        result = _run_pre_checks("C", "Dual authorization / self-approval check", item, "{}")
        assert "Self-approval detected" in result

    def test_passes_dual_auth(self):
        item = {"Preparer": "Alice", "Approver": "Bob"}
        result = _run_pre_checks("C", "Dual authorization / self-approval check", item, "{}")
        assert "Dual authorization verified" in result

    def test_threshold_above(self):
        ctx = json.dumps({"rules": {"threshold_gbp": 100000}})
        item = {"Amount_GBP": "150,000"}
        result = _run_pre_checks("B", "Above-threshold review", item, ctx)
        assert "Finance Director review required" in result

    def test_threshold_below(self):
        ctx = json.dumps({"rules": {"threshold_gbp": 100000}})
        item = {"Amount_GBP": "50000"}
        result = _run_pre_checks("B", "Above-threshold review", item, ctx)
        assert "Below threshold" in result

    def test_missing_supporting_doc(self):
        item = {"Supporting_Doc": ""}
        result = _run_pre_checks("D", "Supporting documentation check", item, "{}")
        assert "FAIL" in result and "empty" in result

    def test_supporting_doc_present(self):
        item = {"Supporting_Doc": "INV-2024-001"}
        result = _run_pre_checks("D", "Supporting documentation check", item, "{}")
        assert "PASS" in result

    def test_no_checks_for_unrelated_attribute(self):
        item = {"some_field": "value"}
        result = _run_pre_checks("X", "Unrelated attribute", item, "{}")
        assert result == ""


# ============================================================================
# Tool: list_projects
# ============================================================================

class TestListProjects:
    def test_lists_local_projects(self):
        with patch(PATCH_WS_CLIENT, side_effect=Exception("no sdk")):
            result_str = tools_module.list_projects.invoke({})
            result = json.loads(result_str)
            assert "projects" in result
            assert result["count"] >= 0
            for proj in result["projects"]:
                assert "project_dir" in proj

    def test_includes_engagement_metadata(self, sample_data_dir):
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(sample_data_dir / "projects")):
            with patch(PATCH_WS_CLIENT, side_effect=Exception("skip")):
                result = json.loads(tools_module.list_projects.invoke({}))
                fin = next((p for p in result["projects"] if p["project_dir"] == "fin_042"), None)
                assert fin is not None
                assert fin["control_id"] == "FIN-042"
                assert fin["domain"] == "Financial Reporting"


# ============================================================================
# Tool: load_engagement
# ============================================================================

class TestLoadEngagement:
    def test_loads_by_project_path(self, tmp_project):
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            result_str = tools_module.load_engagement.invoke({"project_path": "test_proj"})
            result = json.loads(result_str)
            assert result["control_objective"]["control_id"] == "TST-001"
            assert result["number"] == "ENG-TEST-001"

    def test_loads_by_direct_file_path(self, tmp_project):
        path = str(tmp_project / "engagement.json")
        result_str = tools_module.load_engagement.invoke({"project_path": path})
        result = json.loads(result_str)
        assert result["control_objective"]["control_id"] == "TST-001"


# ============================================================================
# Tool: parse_workbook
# ============================================================================

class TestParseWorkbook:
    def _make_xlsx(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sampling"
        ws.append(["Sampling Methodology", "Random"])
        ws.append(["Population Size", "100"])
        ws.append(["Sample Size", "5"])
        ws.append(["", ""])
        ws.append(["Selected Sample Items", ""])
        ws.append(["Item_ID", "Amount", "Description"])
        ws.append(["JE-001", "50000", "Test entry 1"])
        ws.append(["JE-002", "150000", "Test entry 2"])

        ws2 = wb.create_sheet("Testing Table")
        ws2.append(["Ref", "Attribute", "Procedure", "Answer"])
        ws2.append(["A", "Population complete", "Check listing", ""])
        ws2.append(["B", "Threshold check", "Verify amounts", ""])

        path = tmp_path / "workbook.xlsx"
        wb.save(path)
        return path

    def test_parses_tabs(self, tmp_path):
        wb_path = self._make_xlsx(tmp_path)
        result = json.loads(tools_module.parse_workbook.invoke({"file_path": str(wb_path)}))
        assert "Sampling" in result["tab_names"]
        assert "Testing Table" in result["tab_names"]

    def test_extracts_sampling_config(self, tmp_path):
        wb_path = self._make_xlsx(tmp_path)
        result = json.loads(tools_module.parse_workbook.invoke({"file_path": str(wb_path)}))
        assert result["sampling_config"].get("Population Size") == "100"

    def test_extracts_selected_sample(self, tmp_path):
        wb_path = self._make_xlsx(tmp_path)
        result = json.loads(tools_module.parse_workbook.invoke({"file_path": str(wb_path)}))
        assert len(result["selected_sample"]) >= 1

    def test_extracts_testing_attributes(self, tmp_path):
        wb_path = self._make_xlsx(tmp_path)
        result = json.loads(tools_module.parse_workbook.invoke({"file_path": str(wb_path)}))
        assert any(a["ref"] == "A" for a in result["testing_attributes"])

    def test_detects_no_embedded_images(self, tmp_path):
        wb_path = self._make_xlsx(tmp_path)
        result = json.loads(tools_module.parse_workbook.invoke({"file_path": str(wb_path)}))
        assert result["has_embedded_images"] is False

    def test_accepts_project_dir(self, tmp_project):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.save(tmp_project / "engagement_workbook.xlsx")
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            result = json.loads(tools_module.parse_workbook.invoke({"file_path": "test_proj"}))
            assert "tab_names" in result


# ============================================================================
# Tool: extract_workbook_images
# ============================================================================

class TestExtractWorkbookImages:
    def test_extracts_no_images_from_plain_workbook(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.save(tmp_path / "plain.xlsx")

        with patch(PATCH_CHAT) as MockLLM:
            result = json.loads(tools_module.extract_workbook_images.invoke({
                "file_path": str(tmp_path / "plain.xlsx"),
                "context": "test",
            }))
            assert result["images_extracted"] == 0
            MockLLM.assert_not_called()


# ============================================================================
# Tool: review_document
# ============================================================================

class TestReviewDocument:
    def test_reviews_pdf(self, tmp_project, mock_llm):
        pdf_path = str(tmp_project / "evidence" / "doc.pdf")
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = json.loads(tools_module.review_document.invoke({
                "file_path": pdf_path,
                "context": "test control",
                "focus_area": "policy compliance",
            }))
            assert result["file_path"] == pdf_path
            assert result["document_type"] == ".pdf"
            assert result["review_focus"] == "policy compliance"
            assert result["analysis"] == "LLM mock response"

    def test_reviews_text_file(self, tmp_path, mock_llm):
        txt = tmp_path / "doc.txt"
        txt.write_text("This is a text document for review.")
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = json.loads(tools_module.review_document.invoke({"file_path": str(txt)}))
            assert result["analysis"] == "LLM mock response"
            assert result["document_type"] == ".txt"

    def test_uses_default_focus(self, tmp_project, mock_llm):
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = json.loads(tools_module.review_document.invoke({
                "file_path": str(tmp_project / "evidence" / "doc.pdf"),
            }))
            assert result["review_focus"] == "general compliance review"


# ============================================================================
# Tool: review_screenshot
# ============================================================================

class TestReviewScreenshot:
    def test_reviews_png(self, tmp_project, mock_llm):
        img_path = str(tmp_project / "evidence" / "screenshot.png")
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = json.loads(tools_module.review_screenshot.invoke({
                "file_path": img_path,
                "context": "IT controls",
                "focus_area": "access management",
            }))
            assert result["image_type"] == ".png"
            assert result["review_focus"] == "access management"
            assert result["analysis"] == "LLM mock response"

    def test_uses_default_focus(self, tmp_project, mock_llm):
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = json.loads(tools_module.review_screenshot.invoke({
                "file_path": str(tmp_project / "evidence" / "screenshot.png"),
            }))
            assert result["review_focus"] == "general visual inspection"


# ============================================================================
# Tool: analyze_email
# ============================================================================

class TestAnalyzeEmail:
    def test_parses_eml(self, tmp_project, mock_llm):
        eml_path = str(tmp_project / "evidence" / "approval.eml")
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = json.loads(tools_module.analyze_email.invoke({
                "file_path": eml_path,
                "context": "JE approval",
                "focus_area": "authorization",
            }))
            assert result["email_from"] == "alice@example.com"
            assert result["email_to"] == "bob@example.com"
            assert result["email_subject"] == "Approval for JE-001"
            assert result["review_focus"] == "authorization"
            assert result["analysis"] == "LLM mock response"

    def test_uses_default_focus(self, tmp_project, mock_llm):
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = json.loads(tools_module.analyze_email.invoke({
                "file_path": str(tmp_project / "evidence" / "approval.eml"),
            }))
            assert result["review_focus"] == "authorization and approval"


# ============================================================================
# Tool: generate_test_plan
# ============================================================================

class TestGenerateTestPlan:
    def test_generates_plan_for_control_level(self):
        engagement = {
            "testing_attributes": [
                {"ref": "A", "name": "Population complete", "applies_to": "control_level"},
            ],
            "control_objective": {"rules": {}},
        }
        workbook = {
            "selected_sample": [{"Item_ID": "JE-001"}],
            "testing_attributes": [{"ref": "A", "procedure": "Check listing"}],
            "sampling_config": {"Population Size": "100"},
        }
        result = json.loads(tools_module.generate_test_plan.invoke({
            "engagement_json": json.dumps(engagement),
            "workbook_json": json.dumps(workbook),
        }))
        assert result["total_tests"] == 1
        assert result["test_plan"][0]["applies_to"] == "control_level"
        item = json.loads(result["test_plan"][0]["sample_item_json"])
        assert item["_type"] == "population_level"

    def test_generates_plan_for_all_samples(self):
        engagement = {
            "testing_attributes": [{"ref": "C", "name": "Dual auth", "applies_to": "all"}],
            "control_objective": {"rules": {}},
        }
        workbook = {
            "selected_sample": [{"Item_ID": "JE-001"}, {"Item_ID": "JE-002"}],
            "testing_attributes": [],
            "sampling_config": {},
        }
        result = json.loads(tools_module.generate_test_plan.invoke({
            "engagement_json": json.dumps(engagement),
            "workbook_json": json.dumps(workbook),
        }))
        assert result["total_tests"] == 2
        assert result["sample_size"] == 2

    def test_threshold_filter(self):
        engagement = {
            "testing_attributes": [
                {"ref": "B", "name": "Above threshold", "applies_to": "above_threshold"},
            ],
            "control_objective": {"rules": {"threshold_gbp": 100000}},
        }
        workbook = {
            "selected_sample": [
                {"Item_ID": "JE-001", "Amount": "50000"},
                {"Item_ID": "JE-002", "Amount": "150000"},
                {"Item_ID": "JE-003", "Amount": "200000"},
            ],
            "testing_attributes": [],
            "sampling_config": {},
        }
        result = json.loads(tools_module.generate_test_plan.invoke({
            "engagement_json": json.dumps(engagement),
            "workbook_json": json.dumps(workbook),
        }))
        assert result["total_tests"] == 2

    def test_empty_sample_generates_placeholder(self):
        engagement = {
            "testing_attributes": [{"ref": "A", "name": "Check", "applies_to": "all"}],
            "control_objective": {"rules": {}},
        }
        workbook = {"selected_sample": [], "testing_attributes": [], "sampling_config": {}}
        result = json.loads(tools_module.generate_test_plan.invoke({
            "engagement_json": json.dumps(engagement),
            "workbook_json": json.dumps(workbook),
        }))
        assert result["total_tests"] == 1
        item = json.loads(result["test_plan"][0]["sample_item_json"])
        assert item["_type"] == "no_sample_available"

    def test_multiple_attributes_cross_product(self):
        engagement = {
            "testing_attributes": [
                {"ref": "A", "name": "Check A", "applies_to": "all"},
                {"ref": "B", "name": "Check B", "applies_to": "all"},
            ],
            "control_objective": {"rules": {}},
        }
        workbook = {
            "selected_sample": [{"Item_ID": "1"}, {"Item_ID": "2"}, {"Item_ID": "3"}],
            "testing_attributes": [],
            "sampling_config": {},
        }
        result = json.loads(tools_module.generate_test_plan.invoke({
            "engagement_json": json.dumps(engagement),
            "workbook_json": json.dumps(workbook),
        }))
        assert result["total_tests"] == 6  # 2 attrs * 3 samples


# ============================================================================
# Tool: execute_test
# ============================================================================

class TestExecuteTest:
    def test_executes_and_returns_result(self, mock_llm_json):
        with patch(PATCH_CHAT, return_value=mock_llm_json):
            result = json.loads(tools_module.execute_test.invoke({
                "test_ref": "A",
                "attribute": "Population complete",
                "procedure": "Check listing",
                "control_context": json.dumps({"control_id": "TST-001", "rules": {}}),
                "sample_item_json": json.dumps({"_type": "population_level"}),
                "evidence_summary": "The listing contains 100 entries.",
            }))
            assert result["test_ref"] == "A"
            assert "llm_analysis" in result

    def test_injects_pre_checks(self, mock_llm_json):
        with patch(PATCH_CHAT, return_value=mock_llm_json):
            tools_module.execute_test.invoke({
                "test_ref": "C",
                "attribute": "Dual authorization / self-approval check",
                "procedure": "Verify preparer != approver",
                "control_context": "{}",
                "sample_item_json": json.dumps({"Preparer": "Alice", "Approver": "Alice"}),
            })
            call_args = mock_llm_json.invoke.call_args[0][0]
            assert "Self-approval detected" in call_args

    def test_retries_on_rate_limit(self):
        call_count = 0

        def rate_limit_then_succeed(prompt):
            nonlocal call_count
            call_count += 1
            if call_count <= 2:
                raise Exception("429 REQUEST_LIMIT_EXCEEDED")
            return MagicMock(content=json.dumps({
                "result": "Pass", "narrative": "ok", "exception": None, "severity": None,
            }))

        llm = MagicMock()
        llm.invoke = rate_limit_then_succeed
        with patch(PATCH_CHAT, return_value=llm):
            with patch("time.sleep"):
                result = json.loads(tools_module.execute_test.invoke({
                    "test_ref": "A", "attribute": "Test", "procedure": "Test",
                    "control_context": "{}", "sample_item_json": "{}",
                }))
                assert result["test_ref"] == "A"
                assert call_count == 3


# ============================================================================
# Tool: compile_results
# ============================================================================

class TestCompileResults:
    def test_produces_report(self, mock_llm):
        mock_llm.invoke.return_value = MagicMock(content="# Assessment Complete: TST-001\n\n**Pass**")
        with patch(PATCH_CHAT, return_value=mock_llm):
            result = tools_module.compile_results.invoke({
                "control_id": "TST-001",
                "control_name": "Test Control",
                "engagement_number": "ENG-001",
                "domain": "Testing",
                "population_size": 100,
                "sample_size": 5,
                "testing_attributes_json": json.dumps([{"ref": "A", "name": "Check"}]),
                "test_results_json": json.dumps([{"ref": "A", "result": "Pass"}]),
                "rules_json": "{}",
            })
            assert "Assessment Complete" in result
            mock_llm.invoke.assert_called_once()


# ============================================================================
# Tool: save_report
# ============================================================================

class TestSaveReport:
    def test_saves_to_local(self, tmp_project):
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            with patch("agent.tools.vs") as mock_vs:
                mock_vs.save_run_artifact.return_value = None
                with patch("agent.run_context.get_run_id", return_value=""):
                    with patch("agent.run_context.get_project_dir", return_value=""):
                        with patch("agent.run_context.get_report_url", return_value=""):
                            result = json.loads(tools_module.save_report.invoke({
                                "project_path": "test_proj",
                                "report_content": "# Test Report\nEverything passed.",
                                "control_id": "TST-001",
                            }))
                            assert result["status"] == "saved"
                            assert result["report_length"] > 0
                            saved = [f for f in result["files"] if f.endswith(".md")]
                            assert len(saved) >= 1

    def test_includes_report_url(self, tmp_project):
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            with patch("agent.tools.vs") as mock_vs:
                mock_vs.save_run_artifact.return_value = "/Volumes/test/artifact.md"
                with patch("agent.run_context.get_run_id", return_value="run123"):
                    with patch("agent.run_context.get_project_dir", return_value="test_proj"):
                        with patch("agent.run_context.get_report_url", return_value="http://app/api/artifacts/test_proj/run123/report.md"):
                            result = json.loads(tools_module.save_report.invoke({
                                "project_path": "test_proj",
                                "report_content": "# Report",
                            }))
                            assert result["report_url"] == "http://app/api/artifacts/test_proj/run123/report.md"


# ============================================================================
# Tool: fill_workbook
# ============================================================================

class TestFillWorkbook:
    def _make_workbook(self, path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Testing Table"
        ws.append(["Ref", "Attribute", "Procedure", "Answer"])
        ws.append(["A", "Population complete", "Check listing", ""])
        ws.append(["B", "Threshold check", "Verify amounts", ""])

        ws2 = wb.create_sheet("Issue Template")
        ws2.append(["Issue ID", "Ref", "Severity", "Description", "Affected",
                     "Root Cause", "Remediation", "Owner", "Due", "Status"])
        wb.save(path)

    def test_fills_workbook_with_results(self, tmp_project):
        wb_path = tmp_project / "engagement_workbook.xlsx"
        self._make_workbook(wb_path)

        test_results = [
            {"ref": "A", "result": "Pass", "narrative": "Population verified."},
            {"ref": "B", "result": "Fail", "narrative": "Missing approval.", "exceptions": [
                {"description": "No FD sign-off", "severity": "High"},
            ]},
        ]

        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            with patch("agent.tools.vs") as mock_vs:
                mock_vs.save_run_artifact.return_value = None
                with patch("agent.run_context.get_run_id", return_value=""):
                    with patch("agent.run_context.get_project_dir", return_value=""):
                        with patch("agent.run_context.get_app_base_url", return_value=""):
                            with patch("agent.run_context.get_artifact_url", return_value=""):
                                result = json.loads(tools_module.fill_workbook.invoke({
                                    "project_path": "test_proj",
                                    "test_results_json": json.dumps(test_results),
                                    "control_id": "TST-001",
                                }))
                                assert result["status"] == "saved"
                                assert result["attrs_filled"] == 2
                                assert result["exceptions_logged"] >= 1


# ============================================================================
# Tool: send_email
# ============================================================================

class TestSendEmail:
    def test_eml_fallback_without_smtp(self, tmp_project):
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            with patch("agent.config.get_smtp_password", return_value=""):
                with patch("agent.tools.get_smtp_password", return_value=""):
                    with patch("agent.run_context.get_report_url", return_value="http://example.com/report"):
                        with patch("agent.run_context.get_run_id", return_value=""):
                            with patch("agent.run_context.get_project_dir", return_value=""):
                                result = json.loads(tools_module.send_email.invoke({
                                    "to": "recipient@example.com",
                                    "subject": "Test Report",
                                    "body": "# Report\n\nAll tests passed.",
                                    "project_path": "test_proj",
                                }))
                                assert result["status"] == "not_sent"
                                assert result["method"] == "eml_fallback"
                                assert result["to"] == "recipient@example.com"

    def test_html_email_built(self):
        html = _build_html_email(
            subject="Test Report",
            body="**Control**: TST-001\n- Attribute A: Pass\n- Attribute B: Fail",
            report_url="http://example.com/report",
            xlsx_name="workbook.xlsx",
        )
        assert "GSK Controls Evidence Review" in html
        assert "http://example.com/report" in html
        assert "workbook.xlsx" in html
        assert "Attribute A" in html


# ============================================================================
# Tool: ask_user
# ============================================================================

class TestAskUser:
    def test_returns_question(self):
        result = json.loads(tools_module.ask_user.invoke({
            "question": "Which sample items should I retest?",
            "options": "All, Above threshold, None",
        }))
        assert result["type"] == "user_question"
        assert result["question"] == "Which sample items should I retest?"
        assert len(result["suggested_options"]) == 3

    def test_no_options(self):
        result = json.loads(tools_module.ask_user.invoke({
            "question": "Need clarification on this control.",
        }))
        assert result["type"] == "user_question"
        assert "suggested_options" not in result


# ============================================================================
# Tool: batch_review_evidence
# ============================================================================

class TestBatchReviewEvidence:
    def test_dispatches_to_correct_tools(self, tmp_project, mock_llm):
        evidence_files = [
            {"path": "evidence/doc.pdf", "type": "pdf", "focus": "policy"},
            {"path": "evidence/screenshot.png", "type": "screenshot", "focus": "access"},
            {"path": "evidence/approval.eml", "type": "email", "focus": "authorization"},
        ]

        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            with patch(PATCH_CHAT, return_value=mock_llm):
                with patch.object(tools_module, "MAX_PARALLEL_EVIDENCE", 2):
                    result = json.loads(tools_module.batch_review_evidence.invoke({
                        "evidence_files_json": json.dumps(evidence_files),
                        "project_path": "test_proj",
                        "control_context": json.dumps({"control_id": "TST-001"}),
                    }))
                    assert result["files_reviewed"] == 3
                    assert len(result["errors"]) == 0

    def test_handles_file_errors_gracefully(self, tmp_project, mock_llm):
        evidence_files = [
            {"path": "evidence/doc.pdf", "type": "pdf", "focus": "policy"},
            {"path": "evidence/missing_file_xyz.pdf", "type": "pdf", "focus": "missing"},
        ]
        with patch.object(tools_module, "PROJECTS_LOCAL_PATH", str(tmp_project.parent)):
            with patch(PATCH_CHAT, return_value=mock_llm):
                result = json.loads(tools_module.batch_review_evidence.invoke({
                    "evidence_files_json": json.dumps(evidence_files),
                    "project_path": "test_proj",
                    "control_context": "{}",
                }))
                assert result["files_reviewed"] >= 1
                assert len(result["errors"]) >= 1


# ============================================================================
# Tool: batch_execute_tests
# ============================================================================

class TestBatchExecuteTests:
    def test_executes_all_tests(self, mock_llm_json):
        test_plan = [
            {"test_ref": "A", "attribute": "Population", "procedure": "Check", "sample_item_json": "{}"},
            {"test_ref": "B", "attribute": "Threshold", "procedure": "Verify", "sample_item_json": "{}"},
        ]
        with patch(PATCH_CHAT, return_value=mock_llm_json):
            with patch.object(tools_module, "MAX_PARALLEL_TESTS", 2):
                result = json.loads(tools_module.batch_execute_tests.invoke({
                    "test_plan_json": json.dumps(test_plan),
                    "control_context": json.dumps({"control_id": "TST-001"}),
                    "evidence_summary": "All evidence reviewed.",
                }))
                assert result["total_tests"] == 2
                assert result["passed"] + result["failed"] >= 0


# ============================================================================
# Integration: _read_excel_tabs
# ============================================================================

class TestReadExcelTabs:
    def test_reads_all_tabs(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A", "B", "C"])
        ws.append(["1", "2", "3"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y"])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        result = _read_excel_tabs(str(path))
        assert "Sheet1" in result
        assert "Sheet2" in result
        assert result["Sheet1"][0] == ["A", "B", "C"]
        assert result["Sheet1"][1] == ["1", "2", "3"]
